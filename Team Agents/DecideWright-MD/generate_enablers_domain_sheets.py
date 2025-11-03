"""
Generate Complete Enablers Domain Sheets for VOC Predixtive Model
- Brand
- Culture
- People
- Technology
- Third Parties

Expert Senior Business Analyst - Strategy Execution & Risk Management
3-Level hierarchical structure for multivariate Bayesian analysis
"""

import openpyxl
from openpyxl import load_workbook

# Load workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("=" * 80)
print("GENERATING ENABLERS DOMAIN SHEETS")
print("=" * 80)
print()

# ============================================================================
# BRAND SHEET
# ============================================================================

print("1. BRAND SHEET")
print("-" * 80)

ws_brand = wb["Brand"]

# Clear existing content
for row in ws_brand.iter_rows(min_row=2, max_row=ws_brand.max_row):
    for cell in row:
        cell.value = None

brand_data = [
    # LEVEL 1: DIMENSION
    {
        "Level": 1,
        "Hierarchy": "Brand",
        "Description": "Brand identity, equity, and perception in the marketplace",
        "Business Drivers": "Brand Power & Recognition",
        "Business Drivers Description": "Strong brands command premium pricing, customer loyalty, and competitive advantage",
        "Performance Factors": "Brand Strength",
        "Performance Factors Description": "High awareness, positive associations, preference over alternatives, loyalty",
        "Risk Factors": "Brand Damage",
        "Risk Factors Description": "Reputational harm, loss of trust, or brand equity erosion",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of all brand elements - auto-calculated from Level 2"
    },

    # LEVEL 2: BRAND AWARENESS
    {
        "Level": 2,
        "Hierarchy": "Brand - Brand Awareness",
        "Description": "Extent to which target customers recognize and recall the brand",
        "Business Drivers": "Marketing Reach & Frequency",
        "Business Drivers Description": "Sustained visibility and presence in target markets builds awareness",
        "Performance Factors": "High Awareness",
        "Performance Factors Description": "Top-of-mind awareness in target segments, high aided/unaided recall",
        "Risk Factors": "Low Visibility",
        "Risk Factors Description": "Target customers unaware of brand, unable to compete effectively",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of brand awareness sub-elements"
    },

    # LEVEL 3: BRAND AWARENESS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Awareness - Unaided Brand Recall",
        "Description": "Percentage of target market who name your brand without prompting",
        "Business Drivers": "Brand Salience",
        "Business Drivers Description": "Top-of-mind awareness from repeated exposure and memorable experiences",
        "Performance Factors": "Strong Recall",
        "Performance Factors Description": "Brand comes to mind first when category or need mentioned",
        "Risk Factors": "Brand Obscurity",
        "Risk Factors Description": "Customers don't think of brand when considering category",
        "Metric": "Unaided Brand Recall %",
        "Metric Description": "% of target audience who name your brand when asked about category (no prompts)",
        "Unit": "Percentage",
        "Target": ">40% in core segments (leader), >20% (strong), >10% (emerging)",
        "Instructions": "Survey target customers: 'Name brands in [category]' before any prompts. Calculate %."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Awareness - Aided Brand Recognition",
        "Description": "Percentage who recognize brand when shown name/logo",
        "Business Drivers": "Brand Exposure",
        "Business Drivers Description": "Visual presence and consistent branding create recognition",
        "Performance Factors": "High Recognition",
        "Performance Factors Description": "Most target customers recognize brand when shown",
        "Risk Factors": "Brand Confusion",
        "Risk Factors Description": "Brand not recognized or confused with competitors",
        "Metric": "Aided Brand Recognition %",
        "Metric Description": "% who recognize brand when shown name, logo, or tagline",
        "Unit": "Percentage",
        "Target": ">70% in core segments, >50% in expansion segments",
        "Instructions": "Survey: Show brand assets, ask 'Do you recognize this brand?' Calculate recognition %."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Awareness - Share of Voice",
        "Description": "Brand's portion of total marketing/media presence in category",
        "Business Drivers": "Marketing Investment",
        "Business Drivers Description": "Higher marketing spend and effective campaigns increase voice share",
        "Performance Factors": "Voice Leadership",
        "Performance Factors Description": "Brand dominates conversations and mindshare in category",
        "Risk Factors": "Voice Deficit",
        "Risk Factors Description": "Competitors out-communicate and out-visible the brand",
        "Metric": "Share of Voice %",
        "Metric Description": "Your brand mentions / Total category brand mentions (media, social, search)",
        "Unit": "Percentage",
        "Target": "≥ Market share % (defensive), >2× market share (offensive growth)",
        "Instructions": "Track mentions across: paid media, earned media, social media, search. Calculate share."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Awareness - Brand Search Volume",
        "Description": "Volume of searches for brand name or branded terms",
        "Business Drivers": "Brand Interest",
        "Business Drivers Description": "Strong brands generate organic search interest",
        "Performance Factors": "Search Demand",
        "Performance Factors Description": "Growing search volume indicates rising awareness and interest",
        "Risk Factors": "Search Decline",
        "Risk Factors Description": "Falling search volume signals waning interest or relevance",
        "Metric": "Branded Search Volume & Trend",
        "Metric Description": "Monthly branded search volume + YoY % change",
        "Unit": "Count + Percentage change",
        "Target": "Growing ≥20% YoY (emerging), ≥10% YoY (growth), Stable (mature)",
        "Instructions": "Use Google Trends, SEMrush, or search analytics. Track branded terms monthly."
    },

    # LEVEL 2: BRAND PERCEPTION
    {
        "Level": 2,
        "Hierarchy": "Brand - Brand Perception",
        "Description": "How target audiences think and feel about the brand",
        "Business Drivers": "Brand Experience & Messaging",
        "Business Drivers Description": "Customer experiences and brand communications shape perceptions",
        "Performance Factors": "Positive Perception",
        "Performance Factors Description": "Brand associated with desirable attributes and positive emotions",
        "Risk Factors": "Negative Perception",
        "Risk Factors Description": "Brand associated with negative attributes or poor experiences",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of brand perception sub-elements"
    },

    # LEVEL 3: BRAND PERCEPTION SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Perception - Brand Attributes",
        "Description": "Characteristics and qualities associated with the brand",
        "Business Drivers": "Brand Positioning",
        "Business Drivers Description": "Deliberate positioning creates desired attribute associations",
        "Performance Factors": "Aligned Attributes",
        "Performance Factors Description": "Target attributes strongly associated with brand",
        "Risk Factors": "Attribute Mismatch",
        "Risk Factors Description": "Brand not associated with desired attributes or associated with negative ones",
        "Metric": "Brand Attribute Scores",
        "Metric Description": "Top 5 desired attributes: % associating each with your brand vs competitors",
        "Unit": "Percentage per attribute",
        "Target": ">60% association for core attributes, >10% lead vs top competitor",
        "Instructions": "Survey: Rate brands on attributes (quality, innovation, trust, etc.). Compare scores."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Perception - Brand Sentiment",
        "Description": "Emotional tone of conversations and mentions about the brand",
        "Business Drivers": "Customer Experience & PR",
        "Business Drivers Description": "Positive experiences and good PR drive positive sentiment",
        "Performance Factors": "Positive Sentiment",
        "Performance Factors Description": "Majority of brand mentions are positive vs neutral or negative",
        "Risk Factors": "Negative Sentiment",
        "Risk Factors Description": "Predominance of negative mentions damaging brand",
        "Metric": "Net Sentiment Score",
        "Metric Description": "(% Positive mentions - % Negative mentions) across all channels",
        "Unit": "Percentage (-100 to +100)",
        "Target": ">+40 excellent, >+20 good, >0 acceptable, <0 crisis",
        "Instructions": "Use sentiment analysis tools on: social media, reviews, news, surveys. Calculate net score."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Perception - Brand Trust",
        "Description": "Degree to which customers trust the brand to deliver on promises",
        "Business Drivers": "Consistent Delivery",
        "Business Drivers Description": "Repeatedly meeting expectations builds trust over time",
        "Performance Factors": "High Trust",
        "Performance Factors Description": "Brand trusted to deliver quality, honesty, customer care",
        "Risk Factors": "Trust Deficit",
        "Risk Factors Description": "Brand not trusted, creating hesitation and advocacy barriers",
        "Metric": "Brand Trust Score",
        "Metric Description": "Survey: Rate trust in brand on quality, honesty, customer care (1-10 scale average)",
        "Unit": "Score 1-10",
        "Target": ">8/10 trusted brand, >6/10 acceptable, <5/10 trust problem",
        "Instructions": "Survey customers: Rate trust on 3 dimensions (quality delivery, honest communications, customer care)."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Perception - Brand Differentiation",
        "Description": "Perceived uniqueness of brand vs competitors",
        "Business Drivers": "Unique Value Proposition",
        "Business Drivers Description": "Clear differentiation creates brand preference and pricing power",
        "Performance Factors": "Distinct Identity",
        "Performance Factors Description": "Brand seen as meaningfully different from alternatives",
        "Risk Factors": "Brand Commoditization",
        "Risk Factors Description": "Brand seen as interchangeable with competitors",
        "Metric": "Brand Differentiation Score",
        "Metric Description": "Survey: How different is this brand from competitors? (1-10 scale)",
        "Unit": "Score 1-10 (1=identical, 10=completely unique)",
        "Target": ">7/10 highly differentiated, >5/10 moderate, <4/10 commodity",
        "Instructions": "Survey: Rate brand uniqueness. Ask open-ended: What makes brand different? Analyze themes."
    },

    # LEVEL 2: BRAND PREFERENCE
    {
        "Level": 2,
        "Hierarchy": "Brand - Brand Preference",
        "Description": "Likelihood of choosing this brand over alternatives",
        "Business Drivers": "Brand Appeal",
        "Business Drivers Description": "Desirable brand attributes and positive experiences drive preference",
        "Performance Factors": "Preferred Brand",
        "Performance Factors Description": "First choice in consideration set, actively preferred",
        "Risk Factors": "Brand Rejection",
        "Risk Factors Description": "Not considered or actively rejected in favor of alternatives",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of brand preference sub-elements"
    },

    # LEVEL 3: BRAND PREFERENCE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Preference - Consideration Set Inclusion",
        "Description": "Likelihood brand is included when considering category",
        "Business Drivers": "Relevance & Awareness",
        "Business Drivers Description": "Relevant, known brands get considered; unknown or irrelevant brands don't",
        "Performance Factors": "Top Consideration",
        "Performance Factors Description": "Brand consistently included in customer consideration set",
        "Risk Factors": "Exclusion from Consideration",
        "Risk Factors Description": "Brand not considered, eliminating chance to compete",
        "Metric": "Consideration Set Inclusion %",
        "Metric Description": "% of target customers who include brand when actively considering category",
        "Unit": "Percentage",
        "Target": ">60% in core segments, >40% in expansion segments",
        "Instructions": "Survey in-market customers: Which brands did you consider? Calculate inclusion %."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Preference - Purchase Intent",
        "Description": "Stated intention to purchase brand in near future",
        "Business Drivers": "Brand Appeal & Satisfaction",
        "Business Drivers Description": "Appealing brands and satisfied customers show high purchase intent",
        "Performance Factors": "Strong Intent",
        "Performance Factors Description": "High percentage plan to purchase or repurchase",
        "Risk Factors": "Intent Weakness",
        "Risk Factors Description": "Low purchase intent signals brand not compelling",
        "Metric": "Purchase Intent Score",
        "Metric Description": "% likely/very likely to purchase in next [timeframe] (typically 6-12 months)",
        "Unit": "Percentage",
        "Target": ">50% for existing customers (retention), >30% for prospects (acquisition)",
        "Instructions": "Survey: How likely are you to purchase [brand] in next 6 months? (1-5 scale, % 4-5)"
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Preference - Win Rate vs Competitors",
        "Description": "Percentage of competitive situations won",
        "Business Drivers": "Competitive Advantage",
        "Business Drivers Description": "Superior value proposition or brand strength wins competitive deals",
        "Performance Factors": "Competitive Success",
        "Performance Factors Description": "Winning majority of head-to-head competitive situations",
        "Risk Factors": "Competitive Losses",
        "Risk Factors Description": "Losing most competitive deals to specific competitors",
        "Metric": "Competitive Win Rate %",
        "Metric Description": "Deals won / (Deals won + Deals lost to competitors) excluding 'no decision'",
        "Unit": "Percentage",
        "Target": ">50% (competitive), >60% (strong), >70% (dominant)",
        "Instructions": "Track sales pipeline: Win/loss analysis. Calculate win rate by competitor. Identify patterns."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Preference - Price Premium Tolerance",
        "Description": "Willingness to pay more for brand vs alternatives",
        "Business Drivers": "Brand Equity Value",
        "Business Drivers Description": "Strong brands command premium prices without demand destruction",
        "Performance Factors": "Premium Pricing Power",
        "Performance Factors Description": "Customers willing to pay X% more for brand",
        "Risk Factors": "Price Sensitivity",
        "Risk Factors Description": "Brand seen as commodity, no willingness to pay premium",
        "Metric": "Brand Price Premium %",
        "Metric Description": "Average % price premium achieved vs competitors (or % customers willing to pay)",
        "Unit": "Percentage premium",
        "Target": ">20% luxury/premium brands, >10% strong brands, >0% emerging brands",
        "Instructions": "Calculate: Your price vs competitor average. OR Survey: Max % more willing to pay for brand."
    },

    # LEVEL 2: BRAND LOYALTY
    {
        "Level": 2,
        "Hierarchy": "Brand - Brand Loyalty",
        "Description": "Strength of customer attachment and repeat behavior",
        "Business Drivers": "Customer Satisfaction & Switching Costs",
        "Business Drivers Description": "Satisfied customers with high switching costs exhibit strong loyalty",
        "Performance Factors": "Customer Loyalty",
        "Performance Factors Description": "High retention, repeat purchase, resistance to competitor offers",
        "Risk Factors": "Loyalty Weakness",
        "Risk Factors Description": "High churn, price sensitivity, easy switching to competitors",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of brand loyalty sub-elements"
    },

    # LEVEL 3: BRAND LOYALTY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Loyalty - Customer Retention Rate",
        "Description": "Percentage of customers retained year-over-year",
        "Business Drivers": "Ongoing Satisfaction",
        "Business Drivers Description": "Continued satisfaction and value delivery keeps customers",
        "Performance Factors": "High Retention",
        "Performance Factors Description": "Low churn, most customers renew or repurchase",
        "Risk Factors": "Customer Attrition",
        "Risk Factors Description": "High churn rate losing customers to competitors or non-consumption",
        "Metric": "Annual Customer Retention Rate",
        "Metric Description": "Customers at end of period / Customers at start of period (excl. new adds)",
        "Unit": "Percentage",
        "Target": ">90% B2B enterprise, >80% B2B SMB, >60% B2C subscription, varies by model",
        "Instructions": "Calculate: (Customers end - New customers) / Customers start. Track quarterly, report annually."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Loyalty - Repeat Purchase Rate",
        "Description": "Percentage of customers making multiple purchases",
        "Business Drivers": "Product Satisfaction",
        "Business Drivers Description": "Good first experience drives subsequent purchases",
        "Performance Factors": "Strong Repeat Business",
        "Performance Factors Description": "High percentage of customers return for additional purchases",
        "Risk Factors": "One-Time Transactions",
        "Risk Factors Description": "Customers buy once then never return",
        "Metric": "Repeat Purchase Rate",
        "Metric Description": "% of customers making 2+ purchases in defined period (typically 12 months)",
        "Unit": "Percentage",
        "Target": ">40% for transaction models, >60% for relationship models",
        "Instructions": "Calculate: Customers with 2+ purchases / Total customers in period. Track by cohort."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Loyalty - Net Promoter Score (NPS)",
        "Description": "Customer willingness to recommend brand to others",
        "Business Drivers": "Overall Brand Experience",
        "Business Drivers Description": "Entire customer journey and brand interaction drive advocacy",
        "Performance Factors": "Strong Advocacy",
        "Performance Factors Description": "Customers actively promote brand to friends and colleagues",
        "Risk Factors": "Detraction",
        "Risk Factors Description": "Customers warn others against brand, negative word-of-mouth",
        "Metric": "Net Promoter Score",
        "Metric Description": "% Promoters (9-10) minus % Detractors (0-6) on likelihood to recommend",
        "Unit": "Score -100 to +100",
        "Target": ">50 excellent, >30 good, >0 acceptable, <0 urgent action needed",
        "Instructions": "Survey: How likely to recommend 0-10? NPS = %Promoters - %Detractors. Track quarterly."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Loyalty - Customer Lifetime Value (CLV)",
        "Description": "Total profit generated from average customer relationship",
        "Business Drivers": "Retention & Expansion",
        "Business Drivers Description": "Loyal customers stay longer and spend more over time",
        "Performance Factors": "High CLV",
        "Performance Factors Description": "Strong brand loyalty drives increasing customer value",
        "Risk Factors": "CLV Decline",
        "Risk Factors Description": "Weakening loyalty reducing lifetime value",
        "Metric": "Average Customer Lifetime Value",
        "Metric Description": "Average revenue per customer × Gross margin % × Average relationship years",
        "Unit": "Currency",
        "Target": "CLV > 3× CAC (minimum), >5× CAC (good), >10× CAC (excellent)",
        "Instructions": "Calculate: (Avg annual revenue per customer × Margin %) / Churn rate. Compare to CAC."
    },

    # LEVEL 2: BRAND EQUITY
    {
        "Level": 2,
        "Hierarchy": "Brand - Brand Equity",
        "Description": "Overall brand value and financial contribution",
        "Business Drivers": "Accumulated Brand Strength",
        "Business Drivers Description": "Years of positive experiences and marketing build intangible asset value",
        "Performance Factors": "Strong Brand Equity",
        "Performance Factors Description": "Brand recognized as valuable asset with measurable financial contribution",
        "Risk Factors": "Brand Equity Erosion",
        "Risk Factors Description": "Brand value declining due to damage or neglect",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of brand equity sub-elements"
    },

    # LEVEL 3: BRAND EQUITY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Equity - Brand Valuation",
        "Description": "Financial value of the brand as an intangible asset",
        "Business Drivers": "Brand Strength Monetization",
        "Business Drivers Description": "Strong brands create measurable financial value beyond physical assets",
        "Performance Factors": "High Brand Value",
        "Performance Factors Description": "Brand contributes significant portion of enterprise value",
        "Risk Factors": "Brand Value Decline",
        "Risk Factors Description": "Brand asset value decreasing over time",
        "Metric": "Brand Value & Trend",
        "Metric Description": "Estimated brand value using relief from royalty or other method + YoY change",
        "Unit": "Currency + Percentage change",
        "Target": "Brand value >20% of enterprise value, stable or growing",
        "Instructions": "Use brand valuation method (Interbrand, BrandFinance) or relief from royalty. Track annually."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Equity - Brand Contribution to Revenue",
        "Description": "Revenue attributable specifically to brand strength",
        "Business Drivers": "Brand-Driven Sales",
        "Business Drivers Description": "Strong brands drive sales through preference and loyalty",
        "Performance Factors": "High Brand Pull",
        "Performance Factors Description": "Significant revenue directly attributable to brand",
        "Risk Factors": "Brand Irrelevance",
        "Risk Factors Description": "Brand contributes minimally to purchase decisions",
        "Metric": "Brand-Driven Revenue %",
        "Metric Description": "Estimated % of revenue attributable to brand vs product/price/other factors",
        "Unit": "Percentage",
        "Target": ">40% for consumer brands, >25% for B2B brands, varies by category",
        "Instructions": "Conjoint analysis or attribution modeling. Survey: How important was brand in decision? (1-10)"
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Equity - Marketing Efficiency",
        "Description": "Return on marketing investment in building brand",
        "Business Drivers": "Marketing Effectiveness",
        "Business Drivers Description": "Efficient marketing builds brand awareness and preference cost-effectively",
        "Performance Factors": "High Marketing ROI",
        "Performance Factors Description": "Marketing spend generates measurable brand strength and revenue",
        "Risk Factors": "Marketing Waste",
        "Risk Factors Description": "Marketing spend not translating to brand growth or revenue",
        "Metric": "Marketing ROI",
        "Metric Description": "Incremental revenue attributable to marketing / Marketing spend",
        "Unit": "Ratio (e.g., 3:1 = $3 revenue per $1 marketing)",
        "Target": ">3:1 for brand marketing, >5:1 for performance marketing, blended >4:1",
        "Instructions": "Marketing mix modeling or attribution. Calculate: Incremental revenue / Marketing spend."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Equity - Brand Portfolio Strength",
        "Description": "If multiple brands, overall portfolio health and synergy",
        "Business Drivers": "Portfolio Architecture",
        "Business Drivers Description": "Well-structured brand portfolio with clear roles and minimal conflict",
        "Performance Factors": "Portfolio Synergy",
        "Performance Factors Description": "Brands reinforce each other, clear positioning, efficient marketing",
        "Risk Factors": "Portfolio Confusion",
        "Risk Factors Description": "Brand overlap, cannibalization, or resource dilution",
        "Metric": "Brand Portfolio Assessment",
        "Metric Description": "Score: Clear architecture, Minimal overlap, Resource efficiency (each 1-10)",
        "Unit": "Composite score 3-30 (or N/A if single brand)",
        "Target": ">24/30 = strong portfolio management",
        "Instructions": "If multi-brand: Assess architecture clarity, cannibalization risk, marketing efficiency. Sum scores."
    },

    # LEVEL 2: BRAND CONSISTENCY
    {
        "Level": 2,
        "Hierarchy": "Brand - Brand Consistency",
        "Description": "Coherence of brand identity, message, and experience across touchpoints",
        "Business Drivers": "Brand Management Discipline",
        "Business Drivers Description": "Consistent brand execution across channels and geographies builds equity",
        "Performance Factors": "Brand Coherence",
        "Performance Factors Description": "Consistent look, feel, message, and experience everywhere",
        "Risk Factors": "Brand Fragmentation",
        "Risk Factors Description": "Inconsistent brand presentation confusing customers and diluting equity",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of brand consistency sub-elements"
    },

    # LEVEL 3: BRAND CONSISTENCY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Consistency - Brand Identity Compliance",
        "Description": "Adherence to brand standards (logo, colors, fonts, voice)",
        "Business Drivers": "Brand Guidelines Enforcement",
        "Business Drivers Description": "Clear brand standards and governance ensure consistent application",
        "Performance Factors": "Identity Discipline",
        "Performance Factors Description": "All brand touchpoints comply with identity standards",
        "Risk Factors": "Identity Violations",
        "Risk Factors Description": "Off-brand applications creating visual and message confusion",
        "Metric": "Brand Compliance Audit Score",
        "Metric Description": "% of reviewed touchpoints fully compliant with brand guidelines",
        "Unit": "Percentage",
        "Target": ">95% compliance for customer-facing, >90% overall",
        "Instructions": "Audit sample of touchpoints (website, social, ads, packaging, stores). Calculate compliance %."
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Consistency - Message Consistency",
        "Description": "Alignment of brand messaging across channels and campaigns",
        "Business Drivers": "Integrated Marketing",
        "Business Drivers Description": "Coordinated messaging reinforces key brand themes",
        "Performance Factors": "Message Coherence",
        "Performance Factors Description": "Core brand messages consistent across all communications",
        "Risk Factors": "Message Confusion",
        "Risk Factors Description": "Conflicting or contradictory messages across touchpoints",
        "Metric": "Message Alignment Score",
        "Metric Description": "Assessment: Core message present, Tone consistent, No contradictions (Yes/No each)",
        "Unit": "Score 0-3 (count of Yes) or 1-10 subjective",
        "Target": "3/3 on binary check or >8/10 on subjective rating",
        "Instructions": "Review messaging across channels. Check: Core themes present? Tone matches? Any conflicts?"
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Consistency - Customer Experience Alignment",
        "Description": "Consistency between brand promise and actual customer experience",
        "Business Drivers": "Brand Promise Delivery",
        "Business Drivers Description": "Brand is what you do, not what you say - experience must match promise",
        "Performance Factors": "Promise-Experience Fit",
        "Performance Factors Description": "Customer experience consistently delivers on brand promise",
        "Risk Factors": "Promise-Reality Gap",
        "Risk Factors Description": "Experience falls short of brand promise, eroding trust",
        "Metric": "Brand Promise Delivery Score",
        "Metric Description": "Customer rating: Experience matches brand promise (1-10 scale)",
        "Unit": "Score 1-10",
        "Target": ">8/10 = promise delivered, >6/10 acceptable, <5/10 trust problem",
        "Instructions": "Survey: Rate how well experience matched expectations set by brand (advertising, website, etc.)"
    },
    {
        "Level": 3,
        "Hierarchy": "Brand - Brand Consistency - Employee Brand Understanding",
        "Description": "Employee knowledge of and alignment with brand values",
        "Business Drivers": "Internal Brand Engagement",
        "Business Drivers Description": "Employees who understand and believe in brand deliver authentic experiences",
        "Performance Factors": "Brand Champions",
        "Performance Factors Description": "Employees knowledgeable about brand and embody it in actions",
        "Risk Factors": "Employee Brand Disconnect",
        "Risk Factors Description": "Employees don't understand or don't believe in brand",
        "Metric": "Employee Brand Alignment",
        "Metric Description": "Employee survey: Understand brand values, Believe in them, Act accordingly (% agree each)",
        "Unit": "Percentage agreement per question",
        "Target": ">80% on all three dimensions",
        "Instructions": "Employee survey: Do you understand brand values? Believe in them? Feel you embody them?"
    },
]

# Write Brand data
row_idx = 2
for item in brand_data:
    ws_brand.cell(row=row_idx, column=1, value=item["Level"])
    ws_brand.cell(row=row_idx, column=2, value=item["Hierarchy"])
    ws_brand.cell(row=row_idx, column=3, value=item["Description"])
    ws_brand.cell(row=row_idx, column=4, value=item["Business Drivers"])
    ws_brand.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
    ws_brand.cell(row=row_idx, column=6, value=item["Performance Factors"])
    ws_brand.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
    ws_brand.cell(row=row_idx, column=8, value=item["Risk Factors"])
    ws_brand.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
    ws_brand.cell(row=row_idx, column=10, value=item["Metric"])
    ws_brand.cell(row=row_idx, column=11, value=item["Metric Description"])
    ws_brand.cell(row=row_idx, column=12, value=item["Unit"])
    ws_brand.cell(row=row_idx, column=13, value=item["Target"])
    ws_brand.cell(row=row_idx, column=14, value=item["Instructions"])
    row_idx += 1

# Auto-adjust column widths
for column in ws_brand.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws_brand.column_dimensions[column_letter].width = adjusted_width

print("   Brand: {0} rows completed".format(len(brand_data)))
print("   - Level 1: 1 dimension")
print("   - Level 2: 6 elements")
print("   - Level 3: 24 sub-elements")
print()

wb.save(wb_path)

print("=" * 80)
print("BRAND SHEET COMPLETE - Part 1 of 5")
print("=" * 80)
print()
print("Due to length, remaining 4 sheets (Culture, People, Technology, Third Parties)")
print("will be generated in a second script...")
print()
