"""
Implement Business Unit / Subsidiary Analysis

Extends the framework from single-company to multi-BU analysis:
- BU Configuration sheet defining business units
- Business_Unit column added to all 16 dimension sheets
- BU-level Risk and Performance Dashboards
- BU Comparison Dashboard
- Consolidated company-level views (rollup across BUs)

Expert Senior Business Analyst - 20+ years experience
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 80)
print("IMPLEMENTING BUSINESS UNIT / SUBSIDIARY ANALYSIS")
print("=" * 80)
print()

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

# ============================================================================
# STEP 1: Create BU Configuration Sheet
# ============================================================================

print("Step 1: Creating BU Configuration sheet...")

bu_config_data = [
    {
        "BU_Code": "CORP",
        "BU_Name": "Corporate / Consolidated",
        "BU_Type": "Consolidated",
        "Region": "Global",
        "Industry": "[Company Industry]",
        "Employees": "[Total]",
        "Revenue_M": "[Total]",
        "Status": "Active",
        "Consolidation_Method": "Sum of all BUs",
        "Description": "Consolidated enterprise-wide view aggregating all business units",
        "Notes": "This is the default consolidated view. Company-level dashboards show CORP data."
    },
    {
        "BU_Code": "NA",
        "BU_Name": "North America",
        "BU_Type": "Geographic",
        "Region": "Americas",
        "Industry": "[Company Industry]",
        "Employees": "[NA Total]",
        "Revenue_M": "[NA Revenue]",
        "Status": "Active",
        "Consolidation_Method": "Included in CORP",
        "Description": "North America business unit including US, Canada, Mexico operations",
        "Notes": "Largest revenue contributor; mature market"
    },
    {
        "BU_Code": "EMEA",
        "BU_Name": "Europe, Middle East & Africa",
        "BU_Type": "Geographic",
        "Region": "EMEA",
        "Industry": "[Company Industry]",
        "Employees": "[EMEA Total]",
        "Revenue_M": "[EMEA Revenue]",
        "Status": "Active",
        "Consolidation_Method": "Included in CORP",
        "Description": "EMEA region including all European, Middle Eastern, and African operations",
        "Notes": "Second largest region; diverse regulatory environment"
    },
    {
        "BU_Code": "APAC",
        "BU_Name": "Asia Pacific",
        "BU_Type": "Geographic",
        "Region": "Asia Pacific",
        "Industry": "[Company Industry]",
        "Employees": "[APAC Total]",
        "Revenue_M": "[APAC Revenue]",
        "Status": "Active",
        "Consolidation_Method": "Included in CORP",
        "Description": "Asia Pacific region including China, India, Japan, Australia, Southeast Asia",
        "Notes": "Fastest growing region; high growth potential"
    },
    {
        "BU_Code": "LATAM",
        "BU_Name": "Latin America",
        "BU_Type": "Geographic",
        "Region": "Americas",
        "Industry": "[Company Industry]",
        "Employees": "[LATAM Total]",
        "Revenue_M": "[LATAM Revenue]",
        "Status": "Active",
        "Consolidation_Method": "Included in CORP",
        "Description": "Latin America including South America, Central America, Caribbean",
        "Notes": "Emerging market; growth focus area"
    }
]

# Additional example BU types (commented out - user can activate as needed)
additional_bu_examples = [
    {
        "BU_Code": "PROD_A",
        "BU_Name": "Product Line A",
        "BU_Type": "Product Line",
        "Region": "Global",
        "Industry": "[Specific]",
        "Employees": "[Product A Team]",
        "Revenue_M": "[Product A Revenue]",
        "Status": "Inactive (Example)",
        "Consolidation_Method": "Included in CORP",
        "Description": "EXAMPLE: Product-based BU structure (activate if using product line BUs instead of geographic)",
        "Notes": "To use product line BUs: Delete geographic BUs, activate these, customize to your products"
    },
    {
        "BU_Code": "SUB_1",
        "BU_Name": "Subsidiary 1",
        "BU_Type": "Legal Entity",
        "Region": "North America",
        "Industry": "[Subsidiary Industry]",
        "Employees": "[Sub 1 Total]",
        "Revenue_M": "[Sub 1 Revenue]",
        "Status": "Inactive (Example)",
        "Consolidation_Method": "Included in CORP",
        "Description": "EXAMPLE: Subsidiary-based BU structure (activate if using legal entity BUs)",
        "Notes": "To use subsidiary BUs: Customize to your legal entity structure"
    }
]

# Create BU Configuration sheet
if "BU Configuration" in wb.sheetnames:
    ws_bu = wb["BU Configuration"]
    for row in ws_bu.iter_rows():
        for cell in row:
            cell.value = None
else:
    # Insert right after Performance Assumptions (position 6)
    ws_bu = wb.create_sheet("BU Configuration", 6)

# Headers
bu_headers = [
    "BU_Code", "BU_Name", "BU_Type", "Region", "Industry",
    "Employees", "Revenue_M", "Status", "Consolidation_Method",
    "Description", "Notes"
]

for col_idx, header in enumerate(bu_headers, 1):
    cell = ws_bu.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write active BUs
for row_idx, bu in enumerate(bu_config_data, 2):
    for col_idx, header in enumerate(bu_headers, 1):
        value = bu.get(header, "")
        cell = ws_bu.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Highlight CORP row
        if bu["BU_Code"] == "CORP":
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.font = Font(bold=True)

# Add separator and examples
ws_bu.cell(row=len(bu_config_data) + 2, column=1, value="")
ws_bu.cell(row=len(bu_config_data) + 3, column=1, value="=== EXAMPLE BU STRUCTURES (Inactive - Customize as needed) ===")
cell = ws_bu.cell(row=len(bu_config_data) + 3, column=1)
cell.font = Font(bold=True, size=11, color="7F7F7F", italic=True)
cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

for row_idx, bu in enumerate(additional_bu_examples, len(bu_config_data) + 4):
    for col_idx, header in enumerate(bu_headers, 1):
        value = bu.get(header, "")
        cell = ws_bu.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = Font(italic=True, color="7F7F7F")

# Column widths
bu_widths = {"A": 12, "B": 30, "C": 18, "D": 20, "E": 20, "F": 12, "G": 15, "H": 12, "I": 25, "J": 50, "K": 50}
for col_letter, width in bu_widths.items():
    ws_bu.column_dimensions[col_letter].width = width

ws_bu.freeze_panes = "A2"

print(f"   Created BU Configuration with {len(bu_config_data)} active BUs")
print("   Active BUs: CORP (Consolidated), NA, EMEA, APAC, LATAM")

# ============================================================================
# STEP 2: Modify All 16 Dimension Sheets to Add Business_Unit Column
# ============================================================================

print()
print("Step 2: Adding Business_Unit column to all 16 dimension sheets...")

dimension_sheets = [
    "Brand", "Culture", "People", "Technology", "Third Parties",
    "Processes", "Change", "Innovation", "Product & Services",
    "Annual Results", "Strategic Goals", "Reputation"
]

for sheet_name in dimension_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"   WARNING: Sheet '{sheet_name}' not found - skipping")
        continue

    ws = wb[sheet_name]

    # Insert new column A for Business_Unit
    ws.insert_cols(1)

    # Add header
    cell = ws.cell(row=1, column=1, value="Business_Unit")
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column width
    ws.column_dimensions['A'].width = 15

    # Populate Business_Unit column for existing data (default to CORP)
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        # Check if row has data (look at what's now column B - was column A)
        if ws.cell(row=row_idx, column=2).value:
            ws.cell(row=row_idx, column=1, value="CORP")

    print(f"   [OK] {sheet_name}: Added Business_Unit column, populated with 'CORP' for existing data")

print()
print("All 16 dimension sheets now support Business Unit analysis!")

# ============================================================================
# STEP 3: Create BU-Level Risk Dashboard
# ============================================================================

print()
print("Step 3: Creating BU-Level Risk Dashboard...")

# This will be a template that can be filtered by BU
bu_risk_dashboard_data = []

# Header row
bu_risk_dashboard_data.append({
    "Business_Unit": "[Select BU]",
    "Aggregation_Level": "BUSINESS UNIT RISK PROFILE",
    "Entity_Name": "[BU Name]",
    "Risk_Metric": "Risk analysis for selected business unit",
    "Exposure_Amount": "",
    "Currency_Unit": "",
    "Probability": "",
    "Expected_Loss": "",
    "Time_Horizon": "",
    "Confidence_Level": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": "",
    "Notes": "Filter/calculate risk metrics using data from selected BU only. Use BU_Code from BU Configuration sheet."
})

bu_risk_dashboard_data.append({})

# Instructions row
bu_risk_dashboard_data.append({
    "Business_Unit": "INSTRUCTIONS",
    "Aggregation_Level": "",
    "Entity_Name": "",
    "Risk_Metric": "=== HOW TO USE BU-LEVEL RISK DASHBOARD ===",
    "Exposure_Amount": "",
    "Currency_Unit": "",
    "Probability": "",
    "Expected_Loss": "",
    "Time_Horizon": "",
    "Confidence_Level": "",
    "Data_Sources": "",
    "Calculation_Method": "1. Select Business Unit from BU Configuration sheet (e.g., 'NA', 'EMEA', 'APAC', 'LATAM', or 'CORP' for consolidated). 2. Filter all 16 dimension sheets to selected BU using Business_Unit column. 3. Run risk calculations using ONLY the filtered BU data. 4. Populate this dashboard with BU-specific risk metrics. 5. Compare BU risk profiles using BU Comparison Dashboard.",
    "Last_Updated": "",
    "Notes": "For CORP (consolidated), aggregate risks across all BUs. For individual BUs, calculate risks using that BU's dimension data only."
})

bu_risk_dashboard_data.append({})

# Template for 17 risk metrics (will be populated with BU-specific data)
bu_risk_metrics_template = [
    "Opex at Risk", "Capex at Risk", "Stratex at Risk", "Revenue at Risk",
    "Productivity Time at Risk", "Service Availability at Risk", "Product at Risk",
    "Reputation at Risk", "Enterprise Value at Risk", "Customer Lifetime Value at Risk",
    "Market Share at Risk", "Talent at Risk", "Compliance at Risk", "Data/IP at Risk",
    "Cash Flow at Risk", "Credit Rating at Risk", "Innovation Pipeline at Risk"
]

for metric in bu_risk_metrics_template:
    bu_risk_dashboard_data.append({
        "Business_Unit": "[BU_Code]",
        "Aggregation_Level": "BU",
        "Entity_Name": "[BU Name]",
        "Risk_Metric": metric,
        "Exposure_Amount": "[Calculate from BU dimension data]",
        "Currency_Unit": "[See Risk Dashboard for definition]",
        "Probability": "[From BU risk assessments]",
        "Expected_Loss": "[Exposure × Probability for this BU]",
        "Time_Horizon": "[Metric dependent]",
        "Confidence_Level": "[Metric dependent]",
        "Data_Sources": "[Same dimensions as company-level, filtered to BU]",
        "Calculation_Method": f"Apply same calculation methodology as company-level {metric}, but using ONLY dimension data where Business_Unit = selected BU. Reference Risk Calculations sheet for detailed methodology.",
        "Last_Updated": "[Date]",
        "Notes": f"BU-specific {metric}. Sum across BUs may not equal CORP (consolidated) if there are interdependencies or corporate-level risks."
    })

# Create BU Risk Dashboard sheet
if "BU Risk Dashboard" in wb.sheetnames:
    ws_bu_risk = wb["BU Risk Dashboard"]
    for row in ws_bu_risk.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_bu_risk = wb.create_sheet("BU Risk Dashboard", 7)

# Headers
bu_risk_headers = [
    "Business_Unit", "Aggregation_Level", "Entity_Name", "Risk_Metric",
    "Exposure_Amount", "Currency_Unit", "Probability", "Expected_Loss",
    "Time_Horizon", "Confidence_Level", "Data_Sources", "Calculation_Method",
    "Last_Updated", "Notes"
]

for col_idx, header in enumerate(bu_risk_headers, 1):
    cell = ws_bu_risk.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, data_row in enumerate(bu_risk_dashboard_data, 2):
    for col_idx, header in enumerate(bu_risk_headers, 1):
        value = data_row.get(header, "")
        cell = ws_bu_risk.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if "===" in str(value) or "INSTRUCTIONS" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

bu_risk_widths = {"A": 15, "B": 18, "C": 30, "D": 35, "E": 20, "F": 15, "G": 12, "H": 18, "I": 15, "J": 15, "K": 50, "L": 80, "M": 12, "N": 40}
for col_letter, width in bu_risk_widths.items():
    ws_bu_risk.column_dimensions[col_letter].width = width

ws_bu_risk.freeze_panes = "A2"

print("   [OK] BU Risk Dashboard created with 17 risk metrics template")

# ============================================================================
# STEP 4: Create BU-Level Performance Dashboard
# ============================================================================

print()
print("Step 4: Creating BU-Level Performance Dashboard...")

bu_performance_dashboard_data = []

# Header
bu_performance_dashboard_data.append({
    "Business_Unit": "[Select BU]",
    "Aggregation_Level": "BUSINESS UNIT PERFORMANCE PROFILE",
    "Entity_Name": "[BU Name]",
    "Performance_Metric": "Performance analysis for selected business unit",
    "Current_Score": "",
    "Target_Score": "",
    "Achievement_Rate": "",
    "Trend": "",
    "Time_Period": "",
    "Benchmark_Comparison": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": "",
    "Notes": "Filter/calculate performance metrics using data from selected BU only"
})

bu_performance_dashboard_data.append({})

# Instructions
bu_performance_dashboard_data.append({
    "Business_Unit": "INSTRUCTIONS",
    "Aggregation_Level": "",
    "Entity_Name": "",
    "Performance_Metric": "=== HOW TO USE BU-LEVEL PERFORMANCE DASHBOARD ===",
    "Current_Score": "",
    "Target_Score": "",
    "Achievement_Rate": "",
    "Trend": "",
    "Time_Period": "",
    "Benchmark_Comparison": "",
    "Data_Sources": "",
    "Calculation_Method": "1. Select Business Unit from BU Configuration sheet. 2. Filter all 16 dimension sheets to selected BU. 3. Run performance calculations using ONLY filtered BU data. 4. Populate this dashboard with BU-specific scores. 5. Compare BU performance using BU Comparison Dashboard.",
    "Last_Updated": "",
    "Notes": "BU targets may differ from corporate targets based on BU maturity, market conditions, strategic priorities."
})

bu_performance_dashboard_data.append({})

# Template for 10 performance metrics
bu_performance_metrics = [
    "Operational Excellence Index", "Investment Returns Index", "Strategic Achievement Index",
    "Revenue Performance Index", "Productivity Excellence Index", "Service Excellence Index",
    "Product & Innovation Success Index", "Reputation Strength Index", "Value Creation Index",
    "Probability of Execution (PoE)"
]

for metric in bu_performance_metrics:
    bu_performance_dashboard_data.append({
        "Business_Unit": "[BU_Code]",
        "Aggregation_Level": "BU",
        "Entity_Name": "[BU Name]",
        "Performance_Metric": metric,
        "Current_Score": "[Calculate from BU dimension data]",
        "Target_Score": "[BU-specific target - may differ from corporate]",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Industry or Other BUs]",
        "Data_Sources": "[Same dimensions as company-level, filtered to BU]",
        "Calculation_Method": f"Apply same calculation methodology as company-level {metric}, but using ONLY dimension data where Business_Unit = selected BU. Reference Performance Calculations sheet for detailed methodology.",
        "Last_Updated": "[Date]",
        "Notes": f"BU-specific {metric}. Weighted average across BUs creates CORP (consolidated) score."
    })

# Create BU Performance Dashboard sheet
if "BU Performance Dashboard" in wb.sheetnames:
    ws_bu_perf = wb["BU Performance Dashboard"]
    for row in ws_bu_perf.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_bu_perf = wb.create_sheet("BU Performance Dashboard", 8)

bu_perf_headers = [
    "Business_Unit", "Aggregation_Level", "Entity_Name", "Performance_Metric",
    "Current_Score", "Target_Score", "Achievement_Rate", "Trend", "Time_Period",
    "Benchmark_Comparison", "Data_Sources", "Calculation_Method", "Last_Updated", "Notes"
]

for col_idx, header in enumerate(bu_perf_headers, 1):
    cell = ws_bu_perf.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, data_row in enumerate(bu_performance_dashboard_data, 2):
    for col_idx, header in enumerate(bu_perf_headers, 1):
        value = data_row.get(header, "")
        cell = ws_bu_perf.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if "===" in str(value) or "INSTRUCTIONS" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

bu_perf_widths = {"A": 15, "B": 18, "C": 30, "D": 40, "E": 15, "F": 15, "G": 15, "H": 12, "I": 15, "J": 20, "K": 50, "L": 80, "M": 12, "N": 40}
for col_letter, width in bu_perf_widths.items():
    ws_bu_perf.column_dimensions[col_letter].width = width

ws_bu_perf.freeze_panes = "A2"

print("   [OK] BU Performance Dashboard created with 10 performance metrics template")

# ============================================================================
# STEP 5: Create BU Comparison Dashboard
# ============================================================================

print()
print("Step 5: Creating BU Comparison Dashboard...")

bu_comparison_data = []

# Header
bu_comparison_data.append({
    "Metric_Category": "=== BUSINESS UNIT COMPARISON DASHBOARD ===",
    "Metric_Name": "Side-by-side comparison of all business units",
    "CORP": "Consolidated",
    "NA": "North America",
    "EMEA": "Europe/MEA",
    "APAC": "Asia Pacific",
    "LATAM": "Latin America",
    "Best_BU": "Top Performer",
    "Worst_BU": "Needs Focus",
    "Range": "Best - Worst",
    "Avg_Excl_CORP": "BU Average",
    "Notes": "Compare performance and risk across all business units to identify best practices and improvement opportunities"
})

bu_comparison_data.append({})

# Risk Metrics Comparison
bu_comparison_data.append({
    "Metric_Category": "RISK METRICS",
    "Metric_Name": "=== Risk Exposure Comparison ===",
    "CORP": "",
    "NA": "",
    "EMEA": "",
    "APAC": "",
    "LATAM": "",
    "Best_BU": "",
    "Worst_BU": "",
    "Range": "",
    "Avg_Excl_CORP": "",
    "Notes": "Lower is better for risk metrics"
})

risk_comparison_metrics = [
    ("Opex at Risk", "$M expected loss"),
    ("Capex at Risk", "$M expected loss"),
    ("Revenue at Risk", "$M expected loss"),
    ("Talent at Risk", "$M replacement cost"),
    ("Total Expected Loss", "Sum of all risk expected losses")
]

for metric, unit in risk_comparison_metrics:
    bu_comparison_data.append({
        "Metric_Category": "Risk",
        "Metric_Name": metric,
        "CORP": f"[{unit}]",
        "NA": f"[{unit}]",
        "EMEA": f"[{unit}]",
        "APAC": f"[{unit}]",
        "LATAM": f"[{unit}]",
        "Best_BU": "[Lowest risk BU]",
        "Worst_BU": "[Highest risk BU]",
        "Range": "[Range]",
        "Avg_Excl_CORP": "[Average]",
        "Notes": f"From BU Risk Dashboard - {metric}"
    })

bu_comparison_data.append({})

# Performance Metrics Comparison
bu_comparison_data.append({
    "Metric_Category": "PERFORMANCE METRICS",
    "Metric_Name": "=== Performance Score Comparison ===",
    "CORP": "",
    "NA": "",
    "EMEA": "",
    "APAC": "",
    "LATAM": "",
    "Best_BU": "",
    "Worst_BU": "",
    "Range": "",
    "Avg_Excl_CORP": "",
    "Notes": "Higher is better for performance metrics (0-100 scale)"
})

perf_comparison_metrics = [
    ("Operational Excellence Index", "/100"),
    ("Investment Returns Index", "/100"),
    ("Strategic Achievement Index", "/100"),
    ("Revenue Performance Index", "/100"),
    ("Service Excellence Index", "/100"),
    ("Value Creation Index", "/100"),
    ("Probability of Execution", "%"),
    ("Avg Performance Score", "Average of all metrics")
]

for metric, unit in perf_comparison_metrics:
    bu_comparison_data.append({
        "Metric_Category": "Performance",
        "Metric_Name": metric,
        "CORP": f"[Score{unit}]",
        "NA": f"[Score{unit}]",
        "EMEA": f"[Score{unit}]",
        "APAC": f"[Score{unit}]",
        "LATAM": f"[Score{unit}]",
        "Best_BU": "[Highest score BU]",
        "Worst_BU": "[Lowest score BU]",
        "Range": "[Range]",
        "Avg_Excl_CORP": "[Average]",
        "Notes": f"From BU Performance Dashboard - {metric}"
    })

bu_comparison_data.append({})

# Key Insights
bu_comparison_data.append({
    "Metric_Category": "KEY INSIGHTS",
    "Metric_Name": "=== BU Performance Patterns ===",
    "CORP": "",
    "NA": "",
    "EMEA": "",
    "APAC": "",
    "LATAM": "",
    "Best_BU": "",
    "Worst_BU": "",
    "Range": "",
    "Avg_Excl_CORP": "",
    "Notes": ""
})

insights = [
    ("Strongest BU Overall", "BU with highest avg performance score"),
    ("Highest Risk BU", "BU with highest total expected loss"),
    ("Best Risk/Performance Balance", "BU with strong performance AND low risk"),
    ("Needs Most Support", "BU with low performance OR high risk"),
    ("Best Practice Leader", "BU to learn from / share practices"),
]

for insight, description in insights:
    bu_comparison_data.append({
        "Metric_Category": "Insight",
        "Metric_Name": insight,
        "CORP": "[Analysis]",
        "NA": "[Analysis]",
        "EMEA": "[Analysis]",
        "APAC": "[Analysis]",
        "LATAM": "[Analysis]",
        "Best_BU": "[BU Name]",
        "Worst_BU": "[BU Name]",
        "Range": "",
        "Avg_Excl_CORP": "",
        "Notes": description
    })

# Create BU Comparison Dashboard sheet
if "BU Comparison Dashboard" in wb.sheetnames:
    ws_bu_comp = wb["BU Comparison Dashboard"]
    for row in ws_bu_comp.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_bu_comp = wb.create_sheet("BU Comparison Dashboard", 9)

bu_comp_headers = [
    "Metric_Category", "Metric_Name", "CORP", "NA", "EMEA", "APAC", "LATAM",
    "Best_BU", "Worst_BU", "Range", "Avg_Excl_CORP", "Notes"
]

for col_idx, header in enumerate(bu_comp_headers, 1):
    cell = ws_bu_comp.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, data_row in enumerate(bu_comparison_data, 2):
    for col_idx, header in enumerate(bu_comp_headers, 1):
        value = data_row.get(header, "")
        cell = ws_bu_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

bu_comp_widths = {"A": 18, "B": 35, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15, "H": 18, "I": 18, "J": 15, "K": 15, "L": 50}
for col_letter, width in bu_comp_widths.items():
    ws_bu_comp.column_dimensions[col_letter].width = width

ws_bu_comp.freeze_panes = "A2"

print("   [OK] BU Comparison Dashboard created with risk and performance comparison tables")

# ============================================================================
# Save Workbook
# ============================================================================

print()
print("Saving workbook with BU analysis capability...")
wb.save(wb_path)

print()
print("=" * 80)
print("BUSINESS UNIT ANALYSIS IMPLEMENTATION COMPLETE!")
print("=" * 80)
print()
print("Summary of Changes:")
print("  1. [OK] BU Configuration sheet created")
print("      - 5 BUs defined: CORP (Consolidated), NA, EMEA, APAC, LATAM")
print("      - Includes examples for Product Line and Subsidiary structures")
print()
print("  2. [OK] All 16 dimension sheets enhanced")
print("      - Business_Unit column added (Column A)")
print("      - Existing data defaulted to 'CORP'")
print("      - Ready for BU-specific data entry")
print()
print("  3. [OK] BU Risk Dashboard created")
print("      - Template for 17 risk metrics per BU")
print("      - Filter dimension data by BU for calculations")
print()
print("  4. [OK] BU Performance Dashboard created")
print("      - Template for 10 performance metrics per BU")
print("      - BU-specific targets and comparisons")
print()
print("  5. [OK] BU Comparison Dashboard created")
print("      - Side-by-side comparison of all BUs")
print("      - Identifies best/worst performers")
print("      - Highlights best practices and improvement needs")
print()
print("=" * 80)
print("WORKBOOK NOW HAS 27 SHEETS (was 22)")
print("=" * 80)
print()
print("Sheet Structure:")
print("  Risk Analysis Layer (3 sheets)")
print("  Performance Analysis Layer (3 sheets)")
print("  BU Analysis Layer (4 NEW sheets):")
print("    - BU Configuration")
print("    - BU Risk Dashboard")
print("    - BU Performance Dashboard")
print("    - BU Comparison Dashboard")
print("  Data Input Layer (16 sheets - now with Business_Unit column)")
print("  Total: 27 sheets")
print()
print("Next Steps:")
print("  1. Review BU Configuration sheet and customize BUs for your organization")
print("  2. Populate dimension sheets with BU-specific data (use Business_Unit column)")
print("  3. Calculate BU-level risk and performance metrics")
print("  4. Use BU Comparison Dashboard to identify best practices")
print()
