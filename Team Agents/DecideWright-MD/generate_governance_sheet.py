"""
Generate Complete Governance Sheet for VOC Predixtive Model

Expert Senior Business Analyst - Strategy Execution & Risk Management
3-Level hierarchical structure for multivariate Bayesian analysis
"""

import openpyxl
from openpyxl import load_workbook

# Load workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("=" * 80)
print("GENERATING GOVERNANCE SHEET")
print("=" * 80)
print()

ws_gov = wb["Governance"]

# Clear existing content
for row in ws_gov.iter_rows(min_row=2, max_row=ws_gov.max_row):
    for cell in row:
        cell.value = None

governance_data = [
    # LEVEL 1: DIMENSION
    {
        "Level": 1,
        "Hierarchy": "Governance",
        "Description": "Structures, processes, and accountability mechanisms that guide and control the organization",
        "Business Drivers": "Oversight & Accountability",
        "Business Drivers Description": "Board and management systems ensuring organization operates ethically, legally, and effectively",
        "Performance Factors": "Governance Excellence",
        "Performance Factors Description": "Strong oversight, clear accountability, ethical culture, effective controls",
        "Risk Factors": "Governance Failure",
        "Risk Factors Description": "Board ineffectiveness, management misconduct, compliance failures, control breakdowns",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of all governance elements - auto-calculated from Level 2"
    },

    # LEVEL 2: BOARD EFFECTIVENESS
    {
        "Level": 2,
        "Hierarchy": "Governance - Board Effectiveness",
        "Description": "Quality and effectiveness of board oversight and direction",
        "Business Drivers": "Board Composition & Engagement",
        "Business Drivers Description": "Right mix of skills, experience, and independence with active engagement",
        "Performance Factors": "Board Excellence",
        "Performance Factors Description": "Board provides effective oversight, strategic guidance, and accountability",
        "Risk Factors": "Board Failure",
        "Risk Factors Description": "Ineffective board failing to provide oversight or challenge management",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of board effectiveness sub-elements"
    },

    # LEVEL 3: BOARD EFFECTIVENESS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Governance - Board Effectiveness - Board Composition",
        "Description": "Mix of skills, experience, diversity, and independence",
        "Business Drivers": "Talent & Independence",
        "Business Drivers Description": "Directors bring relevant expertise and independence to provide effective oversight",
        "Performance Factors": "Optimal Board Mix",
        "Performance Factors Description": "Diverse skills, backgrounds, and perspectives; majority independent directors",
        "Risk Factors": "Board Composition Weakness",
        "Risk Factors Description": "Lack of relevant expertise, insufficient independence, or groupthink",
        "Metric": "Board Independence & Expertise Score",
        "Metric Description": "% independent directors + % with relevant industry/functional expertise",
        "Unit": "Percentage (independence) + Percentage (expertise coverage)",
        "Target": ">50% independent (private), >66% independent (public), >80% expertise match",
        "Instructions": "Calculate: Independent directors / total. Assess: Directors with industry/financial/risk expertise"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Board Effectiveness - Board Engagement",
        "Description": "Level of board involvement and time commitment",
        "Business Drivers": "Director Commitment",
        "Business Drivers Description": "Directors dedicate sufficient time and attention to fulfill duties",
        "Performance Factors": "Active Board",
        "Performance Factors Description": "High meeting attendance, robust discussion, thorough preparation",
        "Risk Factors": "Board Disengagement",
        "Risk Factors Description": "Directors treating role as ceremonial, insufficient time commitment",
        "Metric": "Board Meeting Attendance & Preparation",
        "Metric Description": "Average attendance % + preparation quality assessment",
        "Unit": "Percentage attendance + Score 1-10 preparation",
        "Target": ">90% attendance, >7/10 preparation quality",
        "Instructions": "Track attendance rates. Survey: Directors arrive prepared, materials reviewed in advance?"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Board Effectiveness - Board Oversight Quality",
        "Description": "Quality of board challenge, questioning, and oversight",
        "Business Drivers": "Oversight Rigor",
        "Business Drivers Description": "Board actively challenges management and exercises independent judgment",
        "Performance Factors": "Constructive Challenge",
        "Performance Factors Description": "Board asks tough questions, debates strategy, doesn't rubber-stamp",
        "Risk Factors": "Oversight Failure",
        "Risk Factors Description": "Board deferring to management without sufficient scrutiny",
        "Metric": "Board Oversight Effectiveness Score",
        "Metric Description": "Assessment: Board challenges management, robust debate, diverse views (1-10 scale)",
        "Unit": "Score 1-10",
        "Target": ">7/10 = effective challenge and oversight",
        "Instructions": "360 evaluation of board: Management, directors, observers rate quality of challenge/debate"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Board Effectiveness - Board Committee Function",
        "Description": "Effectiveness of audit, risk, compensation, nomination committees",
        "Business Drivers": "Specialized Oversight",
        "Business Drivers Description": "Committees provide focused oversight in key areas",
        "Performance Factors": "Committee Excellence",
        "Performance Factors Description": "Committees staffed with experts, meeting regularly, exercising oversight",
        "Risk Factors": "Committee Weakness",
        "Risk Factors Description": "Committees pro forma, lacking expertise, or failing to fulfill mandates",
        "Metric": "Committee Effectiveness Score",
        "Metric Description": "Assessment of Audit, Risk, Compensation committees (each scored 1-10)",
        "Unit": "Score 1-10 per committee (average or individual)",
        "Target": ">7/10 per committee",
        "Instructions": "Assess each committee: Expertise, meeting frequency, thoroughness, independence. Rate 1-10."
    },

    # LEVEL 2: RISK OVERSIGHT
    {
        "Level": 2,
        "Hierarchy": "Governance - Risk Oversight",
        "Description": "Board and management oversight of enterprise risks",
        "Business Drivers": "Risk Governance Structure",
        "Business Drivers Description": "Clear accountability and processes for identifying, assessing, and managing risks",
        "Performance Factors": "Risk Oversight Excellence",
        "Performance Factors Description": "Comprehensive risk identification, robust oversight, clear risk appetite",
        "Risk Factors": "Risk Oversight Failure",
        "Risk Factors Description": "Blind spots, inadequate oversight, or unclear risk accountability",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of risk oversight sub-elements"
    },

    # LEVEL 3: RISK OVERSIGHT SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Governance - Risk Oversight - Risk Governance Framework",
        "Description": "Formal structure for risk identification, assessment, and management",
        "Business Drivers": "Risk Management System",
        "Business Drivers Description": "Defined roles, processes, and tools for enterprise risk management",
        "Performance Factors": "ERM Maturity",
        "Performance Factors Description": "Mature, integrated enterprise risk management framework",
        "Risk Factors": "Risk Management Immaturity",
        "Risk Factors Description": "Ad hoc or siloed risk management without enterprise view",
        "Metric": "ERM Maturity Assessment",
        "Metric Description": "RIMS or COSO ERM maturity: Ad-hoc, Developing, Defined, Managed, Optimized",
        "Unit": "Maturity level (1-5)",
        "Target": "Level 4 (Managed) or 5 (Optimized)",
        "Instructions": "Self-assess against RIMS Maturity Model or COSO ERM framework. Document evidence."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Risk Oversight - Risk Appetite Definition",
        "Description": "Clear articulation of acceptable risk levels",
        "Business Drivers": "Risk Tolerance Clarity",
        "Business Drivers Description": "Board defines how much risk organization willing to accept in pursuit of objectives",
        "Performance Factors": "Defined Risk Appetite",
        "Performance Factors Description": "Documented, quantified risk appetite statements aligned to strategy",
        "Risk Factors": "Risk Appetite Ambiguity",
        "Risk Factors Description": "Unclear risk tolerance leading to inconsistent risk decisions",
        "Metric": "Risk Appetite Statement Quality",
        "Metric Description": "Assessment: Documented, Quantified, Aligned to strategy, Communicated (Yes/No each)",
        "Unit": "Score 0-4 (count of Yes answers)",
        "Target": "4/4 = fully defined risk appetite",
        "Instructions": "Check: Is risk appetite documented? Quantified? Aligned to strategy? Communicated to org?"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Risk Oversight - Board Risk Reporting",
        "Description": "Quality and frequency of risk reporting to board",
        "Business Drivers": "Risk Transparency",
        "Business Drivers Description": "Board receives timely, comprehensive information on key risks",
        "Performance Factors": "Risk Reporting Excellence",
        "Performance Factors Description": "Clear, concise, forward-looking risk reports delivered regularly",
        "Risk Factors": "Risk Reporting Gaps",
        "Risk Factors Description": "Board unaware of key risks due to inadequate reporting",
        "Metric": "Board Risk Report Frequency & Quality",
        "Metric Description": "Reporting frequency + Board assessment of quality (1-10 scale)",
        "Unit": "Frequency (per year) + Score 1-10",
        "Target": "Quarterly minimum, >7/10 quality score",
        "Instructions": "Count risk reports to board annually. Survey board: Rate quality, clarity, usefulness 1-10"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Risk Oversight - Three Lines of Defense",
        "Description": "Separation of risk-taking, risk management, and independent assurance",
        "Business Drivers": "Defense in Depth",
        "Business Drivers Description": "Multiple layers of risk oversight and assurance",
        "Performance Factors": "Three Lines Effectiveness",
        "Performance Factors Description": "Clear separation and coordination among three lines",
        "Risk Factors": "Blurred Lines",
        "Risk Factors Description": "Conflated roles reducing independence and effectiveness",
        "Metric": "Three Lines Model Implementation",
        "Metric Description": "Assessment: 1st line (operations), 2nd line (risk/compliance), 3rd line (audit) functioning",
        "Unit": "Score 1-10 for each line (composite score 3-30)",
        "Target": ">24/30 = strong three lines model",
        "Instructions": "Assess each line: Clearly defined, adequately resourced, independent where required. Sum scores."
    },

    # LEVEL 2: COMPLIANCE MANAGEMENT
    {
        "Level": 2,
        "Hierarchy": "Governance - Compliance Management",
        "Description": "Systems ensuring adherence to laws, regulations, and policies",
        "Business Drivers": "Legal & Regulatory Adherence",
        "Business Drivers Description": "Obligation to comply with applicable laws and regulations",
        "Performance Factors": "Compliance Excellence",
        "Performance Factors Description": "Comprehensive compliance program with clean track record",
        "Risk Factors": "Compliance Failure",
        "Risk Factors Description": "Violations leading to fines, sanctions, or reputational damage",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of compliance management sub-elements"
    },

    # LEVEL 3: COMPLIANCE MANAGEMENT SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Governance - Compliance Management - Compliance Program Maturity",
        "Description": "Sophistication and effectiveness of compliance function",
        "Business Drivers": "Compliance Infrastructure",
        "Business Drivers Description": "Resources, systems, and processes dedicated to compliance",
        "Performance Factors": "Mature Compliance Program",
        "Performance Factors Description": "Well-resourced compliance function with comprehensive coverage",
        "Risk Factors": "Compliance Program Weakness",
        "Risk Factors Description": "Under-resourced or ineffective compliance function",
        "Metric": "Compliance Program Maturity",
        "Metric Description": "Assessment: Ad-hoc, Reactive, Responsive, Proactive, Optimized (1-5 scale)",
        "Unit": "Maturity level 1-5",
        "Target": "Level 4 (Proactive) or 5 (Optimized)",
        "Instructions": "Self-assess against compliance maturity model: resourcing, coverage, integration, culture"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Compliance Management - Compliance Training & Awareness",
        "Description": "Employee understanding of compliance obligations",
        "Business Drivers": "Compliance Culture",
        "Business Drivers Description": "All employees know and understand relevant compliance requirements",
        "Performance Factors": "Compliance Competence",
        "Performance Factors Description": "High completion rates on training, good comprehension scores",
        "Risk Factors": "Compliance Ignorance",
        "Risk Factors Description": "Employees unaware of requirements, leading to violations",
        "Metric": "Compliance Training Completion & Effectiveness",
        "Metric Description": "% employees completing required training + Average assessment score",
        "Unit": "Percentage completion + Score (average %)",
        "Target": ">95% completion, >80% average score on assessments",
        "Instructions": "Track: Training completion rates, assessment scores, time-to-complete for new hires"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Compliance Management - Compliance Violations & Remediation",
        "Description": "Track record of compliance and response to issues",
        "Business Drivers": "Compliance Performance",
        "Business Drivers Description": "Actual compliance with laws, regulations, and internal policies",
        "Performance Factors": "Clean Compliance Record",
        "Performance Factors Description": "Zero or minimal violations, rapid remediation when issues arise",
        "Risk Factors": "Compliance Violations",
        "Risk Factors Description": "Pattern of violations or slow remediation",
        "Metric": "Compliance Violation Rate & Remediation Speed",
        "Metric Description": "# material violations past year + Average days to remediate",
        "Unit": "Count + Days to remediate",
        "Target": "Zero material violations, <30 days to remediate any minor issues",
        "Instructions": "Count violations past 12 months. Track remediation time from identification to closure."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Compliance Management - Whistleblower & Reporting Mechanisms",
        "Description": "Channels for reporting suspected violations",
        "Business Drivers": "Speak-Up Culture",
        "Business Drivers Description": "Safe, confidential channels encouraging reporting of concerns",
        "Performance Factors": "Effective Reporting",
        "Performance Factors Description": "Active hotline/reporting system with no retaliation",
        "Risk Factors": "Reporting Barriers",
        "Risk Factors Description": "Employees afraid to report, leading to hidden issues",
        "Metric": "Whistleblower System Utilization & Trust",
        "Metric Description": "# reports received per year + Employee trust score (survey 1-10)",
        "Unit": "Count of reports + Trust score 1-10",
        "Target": "Active reporting (proportionate to size), >7/10 trust in system",
        "Instructions": "Track reports received annually. Survey: Do you trust whistleblower system? Fear retaliation?"
    },

    # LEVEL 2: ETHICAL CULTURE
    {
        "Level": 2,
        "Hierarchy": "Governance - Ethical Culture",
        "Description": "Values, behaviors, and norms promoting ethical conduct",
        "Business Drivers": "Tone from the Top",
        "Business Drivers Description": "Leadership sets and models ethical standards",
        "Performance Factors": "Ethics Excellence",
        "Performance Factors Description": "Strong ethical culture with integrity at all levels",
        "Risk Factors": "Ethical Lapses",
        "Risk Factors Description": "Misconduct, conflicts of interest, or ethical failures",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of ethical culture sub-elements"
    },

    # LEVEL 3: ETHICAL CULTURE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Governance - Ethical Culture - Code of Conduct Quality",
        "Description": "Clarity and comprehensiveness of ethical standards",
        "Business Drivers": "Ethical Standards Definition",
        "Business Drivers Description": "Clear articulation of expected behaviors and ethical principles",
        "Performance Factors": "Comprehensive Code",
        "Performance Factors Description": "Code covers all key areas, written clearly, regularly updated",
        "Risk Factors": "Code Inadequacy",
        "Risk Factors Description": "Vague, outdated, or incomplete code of conduct",
        "Metric": "Code of Conduct Assessment",
        "Metric Description": "Evaluation: Comprehensive, Clear, Current (<3 yrs), Acknowledged by all (Yes/No each)",
        "Unit": "Score 0-4 (count of Yes)",
        "Target": "4/4 = strong code of conduct",
        "Instructions": "Check: Does code cover key areas? Written clearly? Updated recently? All employees signed?"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Ethical Culture - Leadership Tone & Behavior",
        "Description": "Extent to which leaders model ethical behavior",
        "Business Drivers": "Role Modeling",
        "Business Drivers Description": "Leaders demonstrate commitment to ethics through words and actions",
        "Performance Factors": "Ethical Leadership",
        "Performance Factors Description": "Leaders consistently demonstrate integrity and ethical decision-making",
        "Risk Factors": "Leadership Hypocrisy",
        "Risk Factors Description": "Leaders fail to model stated values, eroding trust",
        "Metric": "Leadership Ethics Score",
        "Metric Description": "Employee perception: Leaders demonstrate integrity and ethical behavior (1-10 scale)",
        "Unit": "Score 1-10",
        "Target": ">7/10 = strong ethical leadership",
        "Instructions": "Anonymous employee survey: Rate leadership integrity, consistency between words/actions"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Ethical Culture - Conflict of Interest Management",
        "Description": "Identification and management of conflicts of interest",
        "Business Drivers": "Independence & Objectivity",
        "Business Drivers Description": "Decisions made in organization's best interest, not personal interests",
        "Performance Factors": "COI Transparency",
        "Performance Factors Description": "Conflicts disclosed and appropriately managed or avoided",
        "Risk Factors": "Hidden Conflicts",
        "Risk Factors Description": "Undisclosed conflicts leading to biased decisions",
        "Metric": "Conflict of Interest Disclosure Rate",
        "Metric Description": "% employees/directors submitting annual COI disclosures + # material conflicts identified",
        "Unit": "Percentage + Count",
        "Target": "100% disclosure compliance, clear processes for material conflicts",
        "Instructions": "Track disclosure form completion. Count and manage material conflicts identified."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Ethical Culture - Ethical Decision-Making Framework",
        "Description": "Tools and processes guiding ethical choices",
        "Business Drivers": "Decision Support",
        "Business Drivers Description": "Providing framework for navigating ethical dilemmas",
        "Performance Factors": "Ethics Infrastructure",
        "Performance Factors Description": "Clear process for escalating and resolving ethical questions",
        "Risk Factors": "Ethics Confusion",
        "Risk Factors Description": "Employees unclear how to handle ethical dilemmas",
        "Metric": "Ethical Support System Utilization",
        "Metric Description": "# ethics questions escalated/resolved + Employee awareness score (survey 1-10)",
        "Unit": "Count + Awareness score 1-10",
        "Target": "Active utilization (proportionate to size), >7/10 awareness of process",
        "Instructions": "Track ethics hotline/consult usage. Survey: Do you know how to escalate ethical concerns?"
    },

    # LEVEL 2: STAKEHOLDER GOVERNANCE
    {
        "Level": 2,
        "Hierarchy": "Governance - Stakeholder Governance",
        "Description": "Consideration of and accountability to diverse stakeholder interests",
        "Business Drivers": "Stakeholder Capitalism",
        "Business Drivers Description": "Creating value for all stakeholders, not just shareholders",
        "Performance Factors": "Stakeholder Balance",
        "Performance Factors Description": "Fair consideration of employee, customer, community, investor interests",
        "Risk Factors": "Stakeholder Conflicts",
        "Risk Factors Description": "Neglecting key stakeholder groups creating backlash or opposition",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of stakeholder governance sub-elements"
    },

    # LEVEL 3: STAKEHOLDER GOVERNANCE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Governance - Stakeholder Governance - Stakeholder Identification & Engagement",
        "Description": "Mapping and engaging with key stakeholder groups",
        "Business Drivers": "Stakeholder Analysis",
        "Business Drivers Description": "Understanding who stakeholders are and what they care about",
        "Performance Factors": "Stakeholder Awareness",
        "Performance Factors Description": "Comprehensive stakeholder map with engagement plans",
        "Risk Factors": "Stakeholder Blindness",
        "Risk Factors Description": "Missing key stakeholders or failing to understand concerns",
        "Metric": "Stakeholder Engagement Score",
        "Metric Description": "# key stakeholder groups identified + Engagement frequency + Satisfaction score",
        "Unit": "Count + Frequency + Score 1-10",
        "Target": "All major groups identified, regular engagement, >7/10 satisfaction",
        "Instructions": "List stakeholder groups. Track engagement frequency. Survey satisfaction with engagement."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Stakeholder Governance - Shareholder Rights & Protections",
        "Description": "Protection of shareholder interests and rights",
        "Business Drivers": "Fiduciary Duty",
        "Business Drivers Description": "Legal obligation to act in shareholders' best interests",
        "Performance Factors": "Shareholder Stewardship",
        "Performance Factors Description": "Strong shareholder protections, fair treatment, transparency",
        "Risk Factors": "Shareholder Conflicts",
        "Risk Factors Description": "Majority/minority conflicts, poor communication, rights violations",
        "Metric": "Shareholder Rights Assessment",
        "Metric Description": "Evaluation: Voting rights, Info access, Fair treatment, Recourse (Yes/No each)",
        "Unit": "Score 0-4",
        "Target": "4/4 = strong shareholder protections",
        "Instructions": "Assess: Equal voting? Access to information? Fair treatment of minorities? Dispute mechanisms?"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Stakeholder Governance - Employee Voice & Representation",
        "Description": "Mechanisms for employee input into governance",
        "Business Drivers": "Employee Engagement",
        "Business Drivers Description": "Engaged employees more productive and aligned",
        "Performance Factors": "Employee Participation",
        "Performance Factors Description": "Forums for employee voice, consideration of employee perspectives",
        "Risk Factors": "Employee Disenfranchisement",
        "Risk Factors Description": "Employees feel unheard, leading to low morale or opposition",
        "Metric": "Employee Engagement Score",
        "Metric Description": "Annual engagement survey: Overall engagement + Trust in leadership (1-10 scale)",
        "Unit": "Score 1-10 (engagement) + Score 1-10 (trust)",
        "Target": ">7/10 engagement, >7/10 trust",
        "Instructions": "Conduct annual engagement survey. Key questions: Engagement level, trust in leadership"
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Stakeholder Governance - Community & Social Impact",
        "Description": "Consideration of community and societal impacts",
        "Business Drivers": "Social License",
        "Business Drivers Description": "Operating with community support and managing social impacts",
        "Performance Factors": "Positive Social Impact",
        "Performance Factors Description": "Net positive impact on communities and society",
        "Risk Factors": "Community Opposition",
        "Risk Factors Description": "Negative impacts generating opposition or social license loss",
        "Metric": "Community Impact Assessment",
        "Metric Description": "Community satisfaction score + Social impact metrics (jobs, investment, etc.)",
        "Unit": "Score 1-10 + Key impact metrics",
        "Target": ">7/10 community satisfaction, positive impact metrics",
        "Instructions": "Survey community stakeholders. Track: local employment, investment, environmental impact"
    },

    # LEVEL 2: INTERNAL CONTROLS
    {
        "Level": 2,
        "Hierarchy": "Governance - Internal Controls",
        "Description": "Processes ensuring reliable reporting, safeguarding assets, and compliance",
        "Business Drivers": "Control Environment",
        "Business Drivers Description": "Systems preventing and detecting errors, fraud, and non-compliance",
        "Performance Factors": "Control Effectiveness",
        "Performance Factors Description": "Strong controls providing reasonable assurance over operations",
        "Risk Factors": "Control Failures",
        "Risk Factors Description": "Weak or missing controls leading to losses, errors, or fraud",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of internal controls sub-elements"
    },

    # LEVEL 3: INTERNAL CONTROLS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Governance - Internal Controls - Financial Controls Effectiveness",
        "Description": "Controls over financial reporting and assets",
        "Business Drivers": "Financial Integrity",
        "Business Drivers Description": "Accurate financial reporting and safeguarding of assets",
        "Performance Factors": "Clean Financial Controls",
        "Performance Factors Description": "No material weaknesses in internal controls over financial reporting (ICFR)",
        "Risk Factors": "Financial Control Weaknesses",
        "Risk Factors Description": "Material weaknesses enabling errors, misstatements, or fraud",
        "Metric": "ICFR Assessment Results",
        "Metric Description": "SOX 404 or equivalent: # material weaknesses + # significant deficiencies",
        "Unit": "Count (MWs + SDs)",
        "Target": "Zero material weaknesses, <3 significant deficiencies",
        "Instructions": "Annual ICFR assessment. Count and classify control deficiencies. Track remediation."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Internal Controls - Operational Controls Maturity",
        "Description": "Controls ensuring efficient and effective operations",
        "Business Drivers": "Operational Discipline",
        "Business Drivers Description": "Standardized processes with controls preventing errors",
        "Performance Factors": "Process Control",
        "Performance Factors Description": "Key operational processes have effective controls",
        "Risk Factors": "Operational Control Gaps",
        "Risk Factors Description": "Weak controls leading to inefficiency, errors, or failures",
        "Metric": "Operational Process Control Assessment",
        "Metric Description": "% of critical processes with documented controls + Control testing pass rate",
        "Unit": "Percentage + Pass rate %",
        "Target": "100% critical processes controlled, >95% pass rate on testing",
        "Instructions": "Inventory critical processes. Document controls. Sample and test controls quarterly."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Internal Controls - IT General Controls (ITGC)",
        "Description": "Controls over IT systems and data",
        "Business Drivers": "IT Reliability & Security",
        "Business Drivers Description": "IT systems operate reliably and securely",
        "Performance Factors": "IT Control Strength",
        "Performance Factors Description": "Strong ITGCs supporting application controls and data integrity",
        "Risk Factors": "IT Control Weaknesses",
        "Risk Factors Description": "Weak ITGCs enabling unauthorized access, data loss, or system failures",
        "Metric": "ITGC Assessment Results",
        "Metric Description": "Assessment: Access controls, Change management, Backup/recovery, Operations (each 1-10)",
        "Unit": "Composite score 4-40",
        "Target": ">32/40 = strong IT controls",
        "Instructions": "Assess 4 ITGC areas: Access, Change, Backup, Operations. Rate 1-10 each. Sum scores."
    },
    {
        "Level": 3,
        "Hierarchy": "Governance - Internal Controls - Fraud Prevention & Detection",
        "Description": "Controls preventing and detecting fraudulent activity",
        "Business Drivers": "Asset Protection",
        "Business Drivers Description": "Safeguarding organization from internal and external fraud",
        "Performance Factors": "Fraud Resilience",
        "Performance Factors Description": "Strong fraud prevention with rapid detection if occurs",
        "Risk Factors": "Fraud Vulnerability",
        "Risk Factors Description": "Weak controls enabling material fraud",
        "Metric": "Fraud Controls & Incidents",
        "Metric Description": "Fraud control assessment score + # fraud incidents + $ losses past year",
        "Unit": "Score 1-10 + Count + Currency",
        "Target": ">7/10 control strength, zero material fraud, <0.1% revenue losses",
        "Instructions": "Assess fraud controls per fraud triangle. Track incidents and losses. Root cause analysis."
    },
]

# Write Governance data
row_idx = 2
for item in governance_data:
    ws_gov.cell(row=row_idx, column=1, value=item["Level"])
    ws_gov.cell(row=row_idx, column=2, value=item["Hierarchy"])
    ws_gov.cell(row=row_idx, column=3, value=item["Description"])
    ws_gov.cell(row=row_idx, column=4, value=item["Business Drivers"])
    ws_gov.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
    ws_gov.cell(row=row_idx, column=6, value=item["Performance Factors"])
    ws_gov.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
    ws_gov.cell(row=row_idx, column=8, value=item["Risk Factors"])
    ws_gov.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
    ws_gov.cell(row=row_idx, column=10, value=item["Metric"])
    ws_gov.cell(row=row_idx, column=11, value=item["Metric Description"])
    ws_gov.cell(row=row_idx, column=12, value=item["Unit"])
    ws_gov.cell(row=row_idx, column=13, value=item["Target"])
    ws_gov.cell(row=row_idx, column=14, value=item["Instructions"])
    row_idx += 1

# Auto-adjust column widths
for column in ws_gov.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws_gov.column_dimensions[column_letter].width = adjusted_width

# Save workbook
wb.save(wb_path)

print("Governance Sheet: {0} rows completed".format(len(governance_data)))
print("  - Level 1: 1 dimension")
print("  - Level 2: 6 elements")
print("  - Level 3: 24 sub-elements")
print()
print("=" * 80)
print("GOVERNANCE SHEET GENERATION COMPLETE")
print("=" * 80)
print()
print("Location: {0}".format(wb_path))
print("Sheet Name: Governance")
print()
print("ECONOMICS DOMAIN NOW COMPLETE:")
print("  1. Financials: 31 rows")
print("  2. Business Model: 31 rows")
print("  3. External Environment: 31 rows")
print("  4. Governance: 31 rows")
print()
print("Total: 124 rows across 4 dimensions")
print("=" * 80)
