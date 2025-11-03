"""
Detailed Analysis of the Financials Sheet
Focus on the 3-level data collection structure
"""

import openpyxl
from openpyxl import load_workbook
import json

# Load the workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path, data_only=True)
ws = wb["Financials"]

print("=" * 100)
print("FINANCIALS SHEET - 3-LEVEL DATA COLLECTION STRUCTURE ANALYSIS")
print("=" * 100)
print()

# 1. Extract complete header row
print("1. COLUMN HEADERS:")
print("-" * 100)
headers = []
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col_idx).value
    if header:
        headers.append((col_idx, header))
        print(f"   Col {col_idx:2d}: {header}")
print()

# 2. Extract all data rows
print("2. COMPLETE DATA STRUCTURE:")
print("-" * 100)
print()

data_rows = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx, header in headers:
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_data[header] = cell_value
    data_rows.append(row_data)

# 3. Analyze the 3-level structure
print("3. THREE-LEVEL HIERARCHY ANALYSIS:")
print("-" * 100)
print()

levels = {}
for idx, row in enumerate(data_rows, start=2):
    level = row.get('Level', 'Unknown')
    hierarchy = row.get('Hierarchy (Dimension - Element - Sub-Element)', 'N/A')
    description = row.get('Description', 'N/A')

    if level not in levels:
        levels[level] = []

    levels[level].append({
        'row': idx,
        'hierarchy': hierarchy,
        'description': description,
        'full_data': row
    })

# Display by level
for level in sorted(levels.keys(), key=lambda x: (x is None, x)):
    level_label = level if level else "[Empty]"
    print(f"LEVEL: {level_label}")
    print("=" * 100)

    for item in levels[level]:
        print(f"   Row {item['row']:2d}: {item['hierarchy']}")
        if item['description'] and item['description'] != 'N/A':
            print(f"           Description: {item['description']}")
        print()

print()

# 4. Detailed breakdown of each level
print("4. DETAILED LEVEL BREAKDOWN:")
print("-" * 100)
print()

for level in sorted(levels.keys(), key=lambda x: (x is None, x)):
    level_label = level if level else "[Empty]"
    count = len(levels[level])

    print(f"{'='*50}")
    print(f"LEVEL {level_label} ({count} items)")
    print(f"{'='*50}")
    print()

    for item in levels[level]:
        print(f"Row {item['row']} - {item['hierarchy']}")
        print("-" * 80)

        for key, value in item['full_data'].items():
            if value and value != 'N/A' and key not in ['Level', 'Hierarchy (Dimension - Element - Sub-Element)']:
                # Truncate long values
                value_str = str(value)
                if len(value_str) > 80:
                    value_str = value_str[:77] + "..."
                print(f"   {key:<40}: {value_str}")
        print()

print()

# 5. Structure Summary
print("5. STRUCTURE SUMMARY:")
print("-" * 100)
print()
print(f"Total Rows (excluding header): {len(data_rows)}")
print(f"Total Columns: {len(headers)}")
print()
print("Level Distribution:")
for level in sorted(levels.keys(), key=lambda x: (x is None, x)):
    level_label = level if level else "[Empty]"
    count = len(levels[level])
    print(f"   Level {level_label}: {count} items")
print()

# 6. Data Collection Fields Analysis
print("6. DATA COLLECTION FIELDS BY LEVEL:")
print("-" * 100)
print()

for level in sorted(levels.keys(), key=lambda x: (x is None, x)):
    level_label = level if level else "[Empty]"
    print(f"LEVEL {level_label}:")

    # Get all fields that have data in this level
    fields_with_data = set()
    for item in levels[level]:
        for key, value in item['full_data'].items():
            if value and value != 'N/A':
                fields_with_data.add(key)

    print(f"   Fields collected: {', '.join(sorted(fields_with_data))}")
    print()

print()
print("=" * 100)
print("ANALYSIS COMPLETE")
print("=" * 100)
print()

# 7. Export structure to text file for review
output_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Output\Financials_Structure_Analysis.txt"
with open(output_path, 'w', encoding='utf-8') as f:
    f.write("FINANCIALS SHEET - 3-LEVEL STRUCTURE\n")
    f.write("=" * 100 + "\n\n")

    for level in sorted(levels.keys(), key=lambda x: (x is None, x)):
        level_label = level if level else "[Empty]"
        f.write(f"\nLEVEL {level_label}\n")
        f.write("-" * 100 + "\n")

        for item in levels[level]:
            f.write(f"\n{item['hierarchy']}\n")
            f.write(f"Description: {item['description']}\n")

            for key, value in item['full_data'].items():
                if value and key not in ['Level', 'Hierarchy (Dimension - Element - Sub-Element)', 'Description']:
                    f.write(f"  {key}: {value}\n")
            f.write("\n")

print(f"Structure analysis exported to: {output_path}")
