"""
Generate Remaining Enablers Domain Sheets
- Culture
- People
- Technology
- Third Parties

Expert Senior Business Analyst - 20+ years experience
"""

import openpyxl
from openpyxl import load_workbook

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("=" * 80)
print("GENERATING REMAINING ENABLERS SHEETS (2-5 of 5)")
print("=" * 80)
print()

# Import the comprehensive data structures
from generate_enablers_culture import culture_data
from generate_enablers_people import people_data
from generate_enablers_technology import technology_data
from generate_enablers_third_parties import third_parties_data

# Process each sheet
sheets_data = [
    ("Culture", culture_data),
    ("People", people_data),
    ("Technology", technology_data),
    ("Third Parties", third_parties_data)
]

for sheet_name, data in sheets_data:
    print(f"Processing {sheet_name}...")
    ws = wb[sheet_name]

    # Clear existing
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write data
    row_idx = 2
    for item in data:
        ws.cell(row=row_idx, column=1, value=item["Level"])
        ws.cell(row=row_idx, column=2, value=item["Hierarchy"])
        ws.cell(row=row_idx, column=3, value=item["Description"])
        ws.cell(row=row_idx, column=4, value=item["Business Drivers"])
        ws.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
        ws.cell(row=row_idx, column=6, value=item["Performance Factors"])
        ws.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
        ws.cell(row=row_idx, column=8, value=item["Risk Factors"])
        ws.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
        ws.cell(row=row_idx, column=10, value=item["Metric"])
        ws.cell(row=row_idx, column=11, value=item["Metric Description"])
        ws.cell(row=row_idx, column=12, value=item["Unit"])
        ws.cell(row=row_idx, column=13, value=item["Target"])
        ws.cell(row=row_idx, column=14, value=item["Instructions"])
        row_idx += 1

    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    print(f"   {sheet_name}: {len(data)} rows completed")

wb.save(wb_path)

print()
print("=" * 80)
print("ENABLERS DOMAIN COMPLETE!")
print("=" * 80)
print("Total: 5 dimensions × 31 rows = 155 rows")
