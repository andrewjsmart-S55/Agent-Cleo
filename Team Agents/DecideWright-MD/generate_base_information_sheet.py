"""
Generate Base Information Sheet
Expert Senior Business Analyst - 20 years experience in Strategy Execution & Risk Management

This sheet collects fundamental organizational data that feeds into the multivariate Bayesian model
for both quantitative risk analysis and performance/strategy execution analysis.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Load workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

# Check if sheet exists, create or clear it
if "Input - Base Data" in wb.sheetnames:
    ws = wb["Input - Base Data"]
    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws = wb.create_sheet("Input - Base Data", 0)  # Insert at beginning

# Styling
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_font = Font(color="FFFFFF", bold=True, size=10)
subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subsection_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Base Information Data Structure
base_data = [
    # HEADER ROW
    ["DATA FIELD", "CLIENT INPUT", "UNIT/TYPE", "VALIDATION RULES", "PURPOSE IN MODEL"],

    # SECTION 1: ORGANIZATION PROFILE
    ["SECTION 1: ORGANIZATION PROFILE", "", "", "", ""],
    ["Company Legal Name", "", "Text", "Required", "Identification & reporting"],
    ["Trading Name(s)", "", "Text", "Optional", "Brand recognition analysis"],
    ["Industry (GICS Sector)", "", "Dropdown: Energy, Materials, Industrials, Consumer Discretionary, Consumer Staples, Health Care, Financials, Information Technology, Communication Services, Utilities, Real Estate", "Required", "Benchmark comparisons & risk profiles"],
    ["Industry Sub-Sector", "", "Text", "Optional", "Refined risk & performance baselines"],
    ["Incorporation Date", "", "Date (YYYY-MM-DD)", "Must be < today", "Organizational maturity factor"],
    ["Headquarters Country", "", "Text", "Required", "Regulatory & macro risk context"],
    ["Primary Operating Region", "", "Dropdown: North America, Europe, Asia-Pacific, Latin America, Middle East & Africa, Global", "Required", "Geographic risk & opportunity factors"],
    ["Number of Operating Countries", "", "Integer", "≥1", "Complexity & geographic diversification"],

    # SECTION 2: SIZE & SCALE
    ["SECTION 2: SIZE & SCALE METRICS", "", "", "", ""],
    ["Total Full-Time Equivalents (FTEs)", "", "Integer", "≥1", "Labor cost base & productivity analysis"],
    ["Total Headcount (All Employment Types)", "", "Integer", "≥1", "Total workforce capacity"],
    ["Annual Revenue (Most Recent FY)", "", "Currency", ">0", "Primary scaling factor for all calculations"],
    ["Currency of Reporting", "", "Dropdown: USD, EUR, GBP, JPY, CNY, AUD, CAD, CHF, Other", "Required", "Financial normalization"],
    ["Total Assets", "", "Currency", "≥0", "Balance sheet strength & capital requirements"],
    ["Market Capitalization (if public)", "", "Currency", "≥0 or N/A", "Enterprise value proxy"],
    ["Enterprise Value Estimate", "", "Currency", ">0", "Risk exposure denominator"],

    # SECTION 3: BUSINESS MODEL
    ["SECTION 3: BUSINESS MODEL ARCHITECTURE", "", "", "", ""],
    ["Primary Business Model", "", "Dropdown: Transaction/Sales, Fee-for-Service, Subscription/Recurring, B2B/Wholesale, Manufacturing/Production, Platform/Marketplace, Advertising/Media, Franchise, Freemium, Licensing/IP", "Required", "Links to business model risk/performance library"],
    ["Secondary Business Model (if applicable)", "", "Dropdown: [Same as above]", "Optional", "Hybrid model complexity factor"],
    ["Revenue Model Diversification", "", "Dropdown: Single stream (>80% from one), Dual stream (2 sources >40% each), Diversified (3+ significant streams)", "Required", "Revenue concentration risk"],
    ["Customer Concentration", "", "Dropdown: Highly concentrated (Top 3 customers >50%), Moderate (Top 10 >50%), Diversified (No customer >10%)", "Required", "Customer dependency risk"],
    ["Average Customer/Contract Tenure", "", "Dropdown: <1 year, 1-3 years, 3-5 years, 5-10 years, >10 years", "Required", "Revenue stability & switching cost proxy"],

    # SECTION 4: STRATEGIC CONTEXT
    ["SECTION 4: STRATEGIC CONTEXT & LIFECYCLE", "", "", "", ""],
    ["Lifecycle Stage", "", "Dropdown: Startup (0-3 yrs), Growth (3-10 yrs), Mature (>10 yrs, stable), Transformation (major change), Decline/Turnaround", "Required", "Risk appetite & volatility expectations"],
    ["Strategic Planning Horizon", "", "Dropdown: 1 year, 3 years, 5 years, 10+ years", "Required", "Time horizon for scenario modeling"],
    ["Growth Strategy", "", "Dropdown: Organic growth, Acquisition-led, Hybrid organic/inorganic, Market consolidation, Geographic expansion, Product innovation", "Required", "Growth risk & capital allocation"],
    ["Competitive Position", "", "Dropdown: Market leader (Top 3), Strong competitor (Top 10), Niche player, Challenger/Disruptor, Emerging/New entrant", "Required", "Competitive pressure & margin sustainability"],
    ["Strategic Priority 1", "", "Dropdown: Revenue growth, Margin expansion, Market share gain, Digital transformation, Operational excellence, Innovation leadership, Geographic expansion, M&A, Sustainability/ESG", "Required", "Alignment of risk/performance to strategy"],
    ["Strategic Priority 2", "", "Dropdown: [Same as above]", "Optional", "Secondary strategic focus"],
    ["Strategic Priority 3", "", "Dropdown: [Same as above]", "Optional", "Tertiary strategic focus"],

    # SECTION 5: RISK & PERFORMANCE CONTEXT
    ["SECTION 5: RISK & PERFORMANCE CONTEXT", "", "", "", ""],
    ["Risk Appetite Level", "", "Dropdown: Conservative (risk-averse), Moderate (balanced), Aggressive (growth-focused), Opportunistic (adaptive)", "Required", "Calibrates risk tolerance in model"],
    ["Risk Management Maturity", "", "Dropdown: Ad-hoc (reactive), Developing (some processes), Defined (documented framework), Managed (integrated), Optimized (continuous improvement)", "Required", "Control effectiveness assumptions"],
    ["Performance Management Approach", "", "Dropdown: Financial metrics only, Balanced Scorecard, OKRs, KPIs + leading indicators, Integrated (strategy/risk/performance)", "Required", "Performance measurement sophistication"],
    ["Data Quality Assessment", "", "Dropdown: Poor (estimates/guesses), Fair (some tracking), Good (regular reporting), Excellent (real-time dashboards)", "Required", "Confidence intervals in Bayesian model"],
    ["Previous Risk/Strategy Assessment Date", "", "Date (YYYY-MM-DD) or 'Never'", "Optional", "Trend analysis if repeat assessment"],

    # SECTION 6: TECHNOLOGY & OPERATIONS
    ["SECTION 6: TECHNOLOGY & OPERATIONS PROFILE", "", "", "", ""],
    ["IT Infrastructure Model", "", "Dropdown: Legacy/On-premises, Hybrid (cloud + on-prem), Cloud-first, Cloud-native", "Required", "Technology risk & agility factors"],
    ["Digital Maturity", "", "Dropdown: Traditional (minimal digital), Digitizing (transitioning), Digital (core systems modern), Digital-native", "Required", "Transformation risk & opportunity"],
    ["Supply Chain Complexity", "", "Dropdown: Simple (1-2 tiers), Moderate (3-5 tiers), Complex (5+ tiers, global), Not applicable", "Required", "Supply chain risk exposure"],
    ["Regulatory Intensity", "", "Dropdown: Light (minimal regulation), Moderate (some oversight), Heavy (highly regulated sector), Extreme (financial services, pharma)", "Required", "Compliance risk & cost"],

    # SECTION 7: FINANCIAL HEALTH INDICATORS
    ["SECTION 7: FINANCIAL HEALTH INDICATORS (Optional but Recommended)", "", "", "", ""],
    ["EBITDA Margin %", "", "Percentage", "0-100", "Operational efficiency"],
    ["Operating Cash Flow (OCF)", "", "Currency", "Any value", "Liquidity & financial flexibility"],
    ["Debt-to-Equity Ratio", "", "Ratio", "≥0", "Financial leverage & solvency risk"],
    ["Current Ratio", "", "Ratio", ">0", "Short-term liquidity"],
    ["Free Cash Flow", "", "Currency", "Any value", "Investment capacity"],
    ["Revenue Growth Rate (YoY %)", "", "Percentage", "-100 to 1000", "Growth trajectory"],

    # SECTION 8: ASSESSMENT PARAMETERS
    ["SECTION 8: ASSESSMENT SCOPE & PARAMETERS", "", "", "", ""],
    ["Assessment Type", "", "Dropdown: Risk Analysis, Performance/Strategy Execution, Combined Risk + Performance", "Required", "Determines model outputs"],
    ["Level of Detail Requested", "", "Dropdown: Base (high-level), Level 1 (department-level), Level 2 (process-level), Level 3 (detailed/control-level)", "Required", "Data collection scope"],
    ["Assessment Time Period", "", "Dropdown: Current state, Next 12 months, Next 3 years, Next 5 years", "Required", "Forecast horizon"],
    ["Include Scenario Analysis?", "", "Dropdown: Yes - Best/Base/Worst case, Yes - Custom scenarios, No", "Optional", "Monte Carlo scenario range"],
    ["Primary Stakeholder(s) for Results", "", "Text", "Optional", "Report customization"],
    ["Assessment Completion Date Target", "", "Date (YYYY-MM-DD)", "Must be > today", "Project timeline"],

    # FOOTER
    ["", "", "", "", ""],
    ["COMPLETION GUIDANCE", "", "", "", ""],
    ["1. Complete ALL fields marked 'Required' in the Validation Rules column", "", "", "", ""],
    ["2. Provide best estimates where exact data is unavailable - the model will account for uncertainty", "", "", "", ""],
    ["3. Use consistent currency across all financial fields", "", "", "", ""],
    ["4. Select 'Level of Detail' based on time available and data access:", "", "", "", ""],
    ["   • Base: 30-60 minutes, uses organizational averages", "", "", "", ""],
    ["   • Level 1: 2-4 hours, department-level breakdown", "", "", "", ""],
    ["   • Level 2: 1-2 days, process and function-level detail", "", "", "", ""],
    ["   • Level 3: 3-5 days, granular control and metric-level data", "", "", "", ""],
]

# Write data to sheet
for row_idx, row_data in enumerate(base_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border

        # Apply styling
        if row_idx == 1:  # Header row
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif "SECTION" in str(value):  # Section headers
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        elif row_data[0] and not row_data[1] and "SECTION" not in str(value):  # Subsections
            cell.fill = subsection_fill
            cell.font = subsection_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 40

# Freeze top row
ws.freeze_panes = 'A2'

# Save workbook
wb.save(wb_path)

print("=" * 80)
print("BASE INFORMATION SHEET GENERATED SUCCESSFULLY")
print("=" * 80)
print()
print(f"Location: {wb_path}")
print(f"Sheet Name: Input - Base Data")
print(f"Total Data Fields: {len([row for row in base_data if row[0] and not row[0].startswith('SECTION') and row[0] != 'DATA FIELD'])}")
print()
print("SECTIONS INCLUDED:")
print("  1. Organization Profile (9 fields)")
print("  2. Size & Scale Metrics (7 fields)")
print("  3. Business Model Architecture (5 fields)")
print("  4. Strategic Context & Lifecycle (7 fields)")
print("  5. Risk & Performance Context (5 fields)")
print("  6. Technology & Operations Profile (4 fields)")
print("  7. Financial Health Indicators (6 fields - optional)")
print("  8. Assessment Scope & Parameters (6 fields)")
print()
print("KEY FEATURES:")
print("  ✓ Feeds multivariate Bayesian model")
print("  ✓ Supports both risk and performance analysis")
print("  ✓ Scalable from Base to Level 3 detail")
print("  ✓ Industry-standard metrics and classifications")
print("  ✓ Built-in validation rules")
print("  ✓ Strategic alignment to VOC framework")
print("=" * 80)
