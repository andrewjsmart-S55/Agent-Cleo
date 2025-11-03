"""
Generate Execution Domain Sheets
- Innovation
- Change
- Processes
- Products & Services

Expert Senior Business Analyst - 20+ years experience
"""

import openpyxl
from openpyxl import load_workbook

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("=" * 80)
print("GENERATING EXECUTION DOMAIN SHEETS (4 of 4)")
print("=" * 80)
print()

# Import the comprehensive data structures
from generate_execution_innovation import innovation_data
from generate_execution_change import change_data
from generate_execution_processes import processes_data
from generate_execution_products_services import products_services_data

# Process each sheet
sheets_data = [
    ("Innovation", innovation_data),
    ("Change", change_data),
    ("Processes", processes_data),
    ("Product & Services", products_services_data)
]

for sheet_name, data in sheets_data:
    print(f"Processing {sheet_name}...")
    ws = wb[sheet_name]

    # Clear existing
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write data
    row_idx = 2
    for item in data:
        ws.cell(row=row_idx, column=1, value=item["Level"])
        ws.cell(row=row_idx, column=2, value=item["Hierarchy"])
        ws.cell(row=row_idx, column=3, value=item["Description"])
        ws.cell(row=row_idx, column=4, value=item["Business Drivers"])
        ws.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
        ws.cell(row=row_idx, column=6, value=item["Performance Factors"])
        ws.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
        ws.cell(row=row_idx, column=8, value=item["Risk Factors"])
        ws.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
        ws.cell(row=row_idx, column=10, value=item["Metric"])
        ws.cell(row=row_idx, column=11, value=item["Metric Description"])
        ws.cell(row=row_idx, column=12, value=item["Unit"])
        ws.cell(row=row_idx, column=13, value=item["Target"])
        ws.cell(row=row_idx, column=14, value=item["Instructions"])
        row_idx += 1

    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    print(f"   {sheet_name}: {len(data)} rows completed")

wb.save(wb_path)

print()
print("=" * 80)
print("EXECUTION DOMAIN COMPLETE!")
print("=" * 80)
print("Total: 4 dimensions × 31 rows = 124 rows")
print()
print("Overall Progress:")
print("  Economics Domain:  4 dimensions (124 rows) - Complete")
print("  Enablers Domain:   5 dimensions (155 rows) - Complete")
print("  Execution Domain:  4 dimensions (124 rows) - Complete")
print("  ------------------------------------------------")
print("  TOTAL COMPLETED:  13 of 16 dimensions (81%)")
print("  TOTAL ROWS:       403 rows")
print()
print("Remaining: VALUE Domain (3 dimensions)")
print("  - Annual Results")
print("  - Strategic Goals")
print("  - Reputation")
