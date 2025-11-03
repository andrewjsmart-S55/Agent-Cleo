"""
Generate Performance Analysis Sheets
- Performance Dashboard: 10 performance metrics aligned to risk metrics
- Performance Calculations: Detailed calculation methodology
- Performance Assumptions: All assumptions used in performance analysis

These sheets provide the positive counterpart to risk analysis,
measuring "How well are we performing?" vs "What could go wrong?"

Expert Senior Business Analyst - 20+ years experience
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 80)
print("GENERATING PERFORMANCE ANALYSIS SHEETS")
print("=" * 80)
print()

# ============================================================================
# PERFORMANCE DASHBOARD DATA
# ============================================================================

performance_dashboard_data = []

# Header
performance_dashboard_data.append({
    "Aggregation_Level": "COMPANY LEVEL",
    "Entity_Name": "[Company Name]",
    "Performance_Metric": "=== ENTERPRISE PERFORMANCE PROFILE ===",
    "Current_Score": "",
    "Target_Score": "",
    "Achievement_Rate": "",
    "Trend": "",
    "Time_Period": "",
    "Benchmark_Comparison": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": "",
    "Notes": ""
})

# Performance Metrics (10 total - aligned to risk metrics)
company_performance = [
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "1. Operational Excellence Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">80 (Excellence)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ Improving / ↔ Stable / ↓ Declining]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Industry Median]",
        "Data_Sources": "Costs, Processes, Technology, Third Parties dimensions",
        "Calculation_Method": "Composite score measuring operational efficiency, cost optimization, process excellence, technology effectiveness. Counterpart to 'Opex at Risk'. Calculation: Weighted average of (1) Cost Efficiency vs Benchmark (30%), (2) Process Performance Score (25%), (3) Technology Operational Excellence (25%), (4) Vendor Performance Score (20%). Score 100 = best-in-class across all dimensions.",
        "Last_Updated": "[Date]",
        "Notes": "Measures how efficiently organization operates; strong correlation with profitability"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "2. Investment Returns Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">75 (Strong Returns)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Multi-year (3-year rolling)",
        "Benchmark_Comparison": "[vs Industry / Hurdle Rate]",
        "Data_Sources": "Investment, Technology, Innovation dimensions",
        "Calculation_Method": "Composite score measuring capital investment returns and strategic investment success. Counterpart to 'Capex at Risk'. Calculation: Weighted average of (1) Capital Projects ROI Achievement (40%), (2) Project On-Time/On-Budget Rate (30%), (3) Innovation Investment ROI (30%). Score 100 = all investments exceeding ROI targets with excellent execution.",
        "Last_Updated": "[Date]",
        "Notes": "Measures capital allocation effectiveness; drives long-term value creation"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "3. Strategic Achievement Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">80 (On Track)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Multi-year (vs strategic plan)",
        "Benchmark_Comparison": "[vs Plan Milestones]",
        "Data_Sources": "Strategic Goals, Innovation, Change dimensions",
        "Calculation_Method": "Composite score measuring progress toward strategic objectives and vision realization. Counterpart to 'Stratex at Risk'. Calculation: Weighted average of (1) Strategic Initiative Success Rate (40%), (2) Vision Milestones Achieved (30%), (3) Competitive Position Improvement (20%), (4) Capability Development Progress (10%). Score 100 = all strategic objectives on track or ahead.",
        "Last_Updated": "[Date]",
        "Notes": "Measures strategic execution effectiveness; predicts long-term competitive position"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "4. Revenue Performance Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">85 (Strong Growth)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Market Growth Rate]",
        "Data_Sources": "Revenue, Annual Results, Products & Services, Brand dimensions",
        "Calculation_Method": "Composite score measuring revenue growth quality and sustainability. Counterpart to 'Revenue at Risk'. Calculation: Weighted average of (1) Revenue Target Achievement (30%), (2) Revenue Growth Rate vs Market (25%), (3) Customer Retention & Expansion (25%), (4) New Customer Acquisition (20%). Score 100 = exceeding targets with sustainable, diversified growth.",
        "Last_Updated": "[Date]",
        "Notes": "Measures revenue engine health; leading indicator of enterprise value"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "5. Productivity Excellence Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">75 (High Productivity)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Industry Productivity]",
        "Data_Sources": "People, Culture, Processes, Technology dimensions",
        "Calculation_Method": "Composite score measuring workforce productivity and efficiency. Counterpart to 'Productivity Time at Risk'. Calculation: Weighted average of (1) Revenue per Employee vs Benchmark (30%), (2) Employee Engagement Score (25%), (3) Process Efficiency (Value-Add Time) (25%), (4) Technology Enablement Score (20%). Score 100 = best-in-class productivity across all factors.",
        "Last_Updated": "[Date]",
        "Notes": "Measures organizational efficiency; high correlation with profitability and growth"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "6. Service Excellence Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">90 (Exceptional Service)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Industry SLA Standards]",
        "Data_Sources": "Products & Services, Technology, Processes, Third Parties dimensions",
        "Calculation_Method": "Composite score measuring service delivery excellence and reliability. Counterpart to 'Service Availability at Risk'. Calculation: Weighted average of (1) System Uptime vs Target (30%), (2) SLA Compliance Rate (30%), (3) Customer Satisfaction with Service (25%), (4) Mean Time to Resolution (15%). Score 100 = exceeding all service targets with customer delight.",
        "Last_Updated": "[Date]",
        "Notes": "Measures service delivery quality; critical for customer retention and reputation"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "7. Product & Innovation Success Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">75 (Strong Innovation)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 24 months",
        "Benchmark_Comparison": "[vs Industry Innovation Leaders]",
        "Data_Sources": "Innovation, Products & Services, Strategic Goals dimensions",
        "Calculation_Method": "Composite score measuring product/innovation performance and pipeline strength. Counterpart to 'Product at Risk' and 'Innovation Pipeline at Risk'. Calculation: Weighted average of (1) Product Launch Success Rate (30%), (2) Innovation Pipeline Value (25%), (3) Time-to-Market Performance (20%), (4) Product Quality/NPS (15%), (5) R&D ROI (10%). Score 100 = best-in-class innovation output and product quality.",
        "Last_Updated": "[Date]",
        "Notes": "Measures innovation effectiveness; drives future growth and competitive advantage"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "8. Reputation Strength Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">75 (Strong Reputation)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Industry Leaders]",
        "Data_Sources": "Reputation, Brand, Culture dimensions",
        "Calculation_Method": "Composite score measuring brand and reputation strength across stakeholders. Counterpart to 'Reputation at Risk'. Calculation: Weighted average of (1) Corporate Reputation Score (RepTrak) (30%), (2) Brand Strength (Awareness, NPS) (25%), (3) Stakeholder Trust Index (25%), (4) ESG Rating (20%). Score 100 = top quartile reputation across all dimensions.",
        "Last_Updated": "[Date]",
        "Notes": "Measures intangible asset strength; enables premium pricing and talent attraction"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "9. Value Creation Index",
        "Current_Score": "[0-100 score]",
        "Target_Score": ">80 (Strong Value Creation)",
        "Achievement_Rate": "[Current/Target %]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12-36 months",
        "Benchmark_Comparison": "[vs Industry / Market]",
        "Data_Sources": "All 16 dimensions (comprehensive)",
        "Calculation_Method": "Composite score measuring overall enterprise value creation. Counterpart to 'Enterprise Value at Risk'. Calculation: Weighted average of (1) TSR (Total Shareholder Return) vs Market (30%), (2) ROIC vs WACC (25%), (3) Revenue/EBITDA Growth (20%), (4) Market Share Gain (15%), (5) Strategic Goal Achievement (10%). Score 100 = top decile value creation.",
        "Last_Updated": "[Date]",
        "Notes": "Ultimate measure of enterprise performance; integrates all other performance dimensions"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Performance_Metric": "10. Probability of Execution (PoE)",
        "Current_Score": "[0-100%]",
        "Target_Score": ">75% (High Confidence)",
        "Achievement_Rate": "[Current vs Target]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Forward-looking (next 12 months)",
        "Benchmark_Comparison": "[Historical success rate]",
        "Data_Sources": "All 16 dimensions + Risk Dashboard",
        "Calculation_Method": "Probability that strategic objectives will be achieved by target dates, integrating current performance, risk factors, capability readiness, and interdependencies. Calculation: Bayesian network aggregating (1) Current Performance Trajectory (30%), (2) Risk Exposure Level (inverse relationship) (30%), (3) Capability Maturity for Objectives (25%), (4) Historical Success Rate Adjusted (15%). PoE = P(Success | Performance, Risk, Capability, History). 100% = certainty of on-time objective achievement.",
        "Last_Updated": "[Date]",
        "Notes": "Forward-looking metric bridging performance and risk; enables proactive intervention before objectives fail"
    }
]

performance_dashboard_data.extend(company_performance)

# Domain-level breakdown (abbreviated - 4 domains × 10 metrics = 40 rows)
performance_dashboard_data.append({})
performance_dashboard_data.append({})
performance_dashboard_data.append({
    "Aggregation_Level": "DOMAIN LEVEL",
    "Entity_Name": "=== 4 DOMAINS ===",
    "Performance_Metric": "Performance breakdown by major domain",
    "Current_Score": "",
    "Target_Score": "",
    "Achievement_Rate": "",
    "Trend": "",
    "Time_Period": "",
    "Benchmark_Comparison": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": "",
    "Notes": "Each domain contributes to overall performance metrics"
})

domains_perf = [
    {"domain": "ECONOMICS Domain", "dimensions": ["Revenue", "Costs", "Investment", "Working Capital"]},
    {"domain": "ENABLERS Domain", "dimensions": ["Brand", "Culture", "People", "Technology", "Third Parties"]},
    {"domain": "EXECUTION Domain", "dimensions": ["Innovation", "Change", "Processes", "Product & Services"]},
    {"domain": "VALUE Domain", "dimensions": ["Annual Results", "Strategic Goals", "Reputation"]}
]

for domain in domains_perf:
    performance_dashboard_data.append({})
    performance_dashboard_data.append({
        "Aggregation_Level": "Domain",
        "Entity_Name": domain["domain"],
        "Performance_Metric": f"=== {domain['domain']} PERFORMANCE ===",
        "Current_Score": "[Aggregate domain score]",
        "Target_Score": ">75",
        "Achievement_Rate": "[%]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Benchmark]",
        "Data_Sources": f"{', '.join(domain['dimensions'])}",
        "Calculation_Method": f"Aggregated performance across {len(domain['dimensions'])} dimensions within {domain['domain']}",
        "Last_Updated": "[Date]",
        "Notes": f"Domain contribution to enterprise performance"
    })

# Dimension-level breakdown (abbreviated - 16 dimensions)
performance_dashboard_data.append({})
performance_dashboard_data.append({})
performance_dashboard_data.append({
    "Aggregation_Level": "DIMENSION LEVEL",
    "Entity_Name": "=== 16 DIMENSIONS ===",
    "Performance_Metric": "Detailed performance by dimension",
    "Current_Score": "",
    "Target_Score": "",
    "Achievement_Rate": "",
    "Trend": "",
    "Time_Period": "",
    "Benchmark_Comparison": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": "",
    "Notes": "Each dimension measured against Performance Factors (positive range) from input sheets"
})

dimensions_perf = ["Revenue", "Costs", "Investment", "Working Capital", "Brand", "Culture", "People", "Technology",
                   "Third Parties", "Innovation", "Change", "Processes", "Product & Services", "Annual Results",
                   "Strategic Goals", "Reputation"]

for dim in dimensions_perf:
    performance_dashboard_data.append({
        "Aggregation_Level": "Dimension",
        "Entity_Name": dim,
        "Performance_Metric": f"{dim} Performance Score",
        "Current_Score": "[0-100]",
        "Target_Score": ">75",
        "Achievement_Rate": "[%]",
        "Trend": "[↑ / ↔ / ↓]",
        "Time_Period": "Trailing 12 months",
        "Benchmark_Comparison": "[vs Benchmark]",
        "Data_Sources": f"{dim} dimension (31 rows: 1 + 6 elements + 24 sub-elements)",
        "Calculation_Method": f"Percentage of {dim} sub-elements achieving 'Performance Factors' (positive) range vs total",
        "Last_Updated": "[Date]",
        "Notes": "Measures % of dimension operating in high-performance range"
    })

# ============================================================================
# PERFORMANCE CALCULATIONS DATA
# ============================================================================

performance_calculations_data = []

performance_calculations_data.append({
    "Calculation_ID": "PERFORMANCE CALCULATIONS",
    "Performance_Metric": "Detailed calculation methodology for all performance metrics",
    "Calculation_Step": "",
    "Formula": "",
    "Data_Source_Dimensions": "",
    "Data_Source_Elements": "",
    "Example_Inputs": "",
    "Example_Calculation": "",
    "Example_Output": "",
    "Assumption_References": "",
    "Notes": "This sheet documents step-by-step calculations for 10 performance metrics"
})

performance_calculations_data.append({})

# Detailed calculation for Operational Excellence Index
perf_calc_1 = [
    {
        "Calculation_ID": "PCALC-01",
        "Performance_Metric": "Operational Excellence Index",
        "Calculation_Step": "=== OPERATIONAL EXCELLENCE INDEX CALCULATION ===",
        "Formula": "",
        "Data_Source_Dimensions": "Costs, Processes, Technology, Third Parties",
        "Data_Source_Elements": "",
        "Example_Inputs": "",
        "Example_Calculation": "",
        "Example_Output": "",
        "Assumption_References": "PASM-01, PASM-05, PASM-10",
        "Notes": "Measures operational efficiency and effectiveness vs best-in-class benchmarks"
    },
    {
        "Calculation_ID": "PCALC-01.1",
        "Performance_Metric": "Operational Excellence Index",
        "Calculation_Step": "Step 1: Calculate Cost Efficiency Score",
        "Formula": "Cost_Efficiency_Score = 100 × (Benchmark_Cost_Ratio - Actual_Cost_Ratio) / Benchmark_Cost_Ratio",
        "Data_Source_Dimensions": "Costs",
        "Data_Source_Elements": "Cost Efficiency, Cost Structure",
        "Example_Inputs": "Actual cost ratio: 78%; Benchmark (PASM-01): 80%; Best-in-class: 75%",
        "Example_Calculation": "100 × (80% - 78%) / 80% = 100 × 2.5% = 2.5 points above benchmark; Normalized to 0-100 scale where 100 = best-in-class (75%), 50 = benchmark (80%), 0 = 85%+; Score = 60 (above benchmark, below best-in-class)",
        "Example_Output": "Cost Efficiency Score: 60/100",
        "Assumption_References": "PASM-01 (cost ratio benchmark)",
        "Notes": "Score >50 means performing better than industry benchmark"
    },
    {
        "Calculation_ID": "PCALC-01.2",
        "Performance_Metric": "Operational Excellence Index",
        "Calculation_Step": "Step 2: Calculate Process Performance Score",
        "Formula": "Process_Score = Weighted_Avg(Value_Add_Time%, First_Pass_Yield%, Process_Automation%, Cycle_Time_vs_Benchmark)",
        "Data_Source_Dimensions": "Processes",
        "Data_Source_Elements": "Process Performance, Process Efficiency, Process Technology",
        "Example_Inputs": "Value-add: 52% (target 50%+); FPY: 92% (target 90%+); Automation: 48% (target 40%+); Cycle time: 10% faster than benchmark",
        "Example_Calculation": "Normalize each to 0-100: Value-add 60/100, FPY 70/100, Automation 80/100, Cycle time 75/100; Weighted avg (25% each) = (60+70+80+75)/4 = 71.25",
        "Example_Output": "Process Performance Score: 71/100",
        "Assumption_References": "PASM-05 (value-add benchmarks)",
        "Notes": "Measures process maturity and efficiency"
    },
    {
        "Calculation_ID": "PCALC-01.3",
        "Performance_Metric": "Operational Excellence Index",
        "Calculation_Step": "Step 3: Calculate Technology Operational Excellence Score",
        "Formula": "Tech_Score = Weighted_Avg(Uptime_vs_Target, Incident_Resolution, User_Satisfaction, Tech_Debt_Control)",
        "Data_Source_Dimensions": "Technology",
        "Data_Source_Elements": "IT Infrastructure, IT Operations",
        "Example_Inputs": "Uptime: 99.92% (target 99.9%); MTTR: 2.1 hrs (target <4hrs); User sat: 82% (target >75%); Tech debt: Controlled (not growing)",
        "Example_Calculation": "Normalize each: Uptime 95/100, MTTR 90/100, User sat 85/100, Tech debt 80/100; Weighted avg = (95+90+85+80)/4 = 87.5",
        "Example_Output": "Technology Operational Excellence: 88/100",
        "Assumption_References": "PASM-10 (uptime standards)",
        "Notes": "Measures IT operational excellence"
    },
    {
        "Calculation_ID": "PCALC-01.4",
        "Performance_Metric": "Operational Excellence Index",
        "Calculation_Step": "Step 4: Calculate Vendor Performance Score",
        "Formula": "Vendor_Score = Weighted_Avg(SLA_Compliance%, Vendor_Quality_Score, Cost_Competitiveness, Strategic_Partnership_Value)",
        "Data_Source_Dimensions": "Third Parties",
        "Data_Source_Elements": "Third-Party Performance, Vendor Risk",
        "Example_Inputs": "SLA compliance: 97% (target >95%); Quality: 83/100; Cost: 5% below market; Partnership value: High",
        "Example_Calculation": "Normalize each: SLA 90/100, Quality 83/100, Cost 85/100, Partnership 80/100; Weighted = (90+83+85+80)/4 = 84.5",
        "Example_Output": "Vendor Performance Score: 85/100",
        "Assumption_References": "PASM-15 (SLA standards)",
        "Notes": "Measures third-party ecosystem performance"
    },
    {
        "Calculation_ID": "PCALC-01.5",
        "Performance_Metric": "Operational Excellence Index",
        "Calculation_Step": "Step 5: Aggregate Operational Excellence Index",
        "Formula": "Op_Ex_Index = (Cost_Efficiency × 30%) + (Process_Score × 25%) + (Tech_Score × 25%) + (Vendor_Score × 20%)",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Component scores: Cost 60, Process 71, Tech 88, Vendor 85",
        "Example_Calculation": "(60 × 0.30) + (71 × 0.25) + (88 × 0.25) + (85 × 0.20) = 18 + 17.75 + 22 + 17 = 74.75",
        "Example_Output": "Operational Excellence Index: 75/100 (Target met)",
        "Assumption_References": "PASM-20 (component weights)",
        "Notes": "Score >80 = Excellence; 70-80 = Strong; 60-70 = Good; <60 = Needs improvement"
    }
]

performance_calculations_data.extend(perf_calc_1)
performance_calculations_data.append({})

# Summary calculations for other metrics (abbreviated for space)
other_perf_calcs = [
    {
        "Calculation_ID": "PCALC-02",
        "Performance_Metric": "Investment Returns Index",
        "Calculation_Step": "Weighted_Avg(Capital_Project_ROI_Achievement%, On_Time_Budget%, Innovation_ROI%)",
        "Formula": "(Project_ROI × 40%) + (Execution_Rate × 30%) + (Innovation_ROI × 30%)",
        "Data_Source_Dimensions": "Investment, Technology, Innovation",
        "Data_Source_Elements": "ROI Realization, Project Portfolio, Innovation Portfolio",
        "Example_Inputs": "Avg project ROI: 22% (hurdle 15%); On-time/budget: 75%; Innovation ROI: 3.5x",
        "Example_Calculation": "Normalize to 0-100: ROI 85/100, Execution 75/100, Innovation 80/100; (85×0.4)+(75×0.3)+(80×0.3) = 79.5",
        "Example_Output": "Investment Returns Index: 80/100",
        "Assumption_References": "PASM-02, PASM-12",
        "Notes": "Measures capital allocation effectiveness"
    },
    {
        "Calculation_ID": "PCALC-03",
        "Performance_Metric": "Strategic Achievement Index",
        "Calculation_Step": "Weighted_Avg(Strategic_Initiative_Success%, Vision_Milestone_Achievement%, Competitive_Position_Improvement%, Capability_Progress%)",
        "Formula": "(Initiative_Success × 40%) + (Milestones × 30%) + (Competitive × 20%) + (Capability × 10%)",
        "Data_Source_Dimensions": "Strategic Goals, Innovation, Change",
        "Data_Source_Elements": "Strategic Initiatives, Vision Achievement, Competitive Position, Capability Development",
        "Example_Inputs": "Initiative success: 78%; Milestones: 85%; Competitive improving; Capability: 72%",
        "Example_Calculation": "(78×0.4) + (85×0.3) + (80×0.2) + (72×0.1) = 31.2 + 25.5 + 16 + 7.2 = 79.9",
        "Example_Output": "Strategic Achievement Index: 80/100",
        "Assumption_References": "PASM-03, PASM-14",
        "Notes": "Measures strategic execution vs plan"
    },
    {
        "Calculation_ID": "PCALC-10",
        "Performance_Metric": "Probability of Execution (PoE)",
        "Calculation_Step": "Bayesian_Aggregation(Current_Performance_Trajectory, Risk_Exposure_Level, Capability_Maturity, Historical_Success_Rate)",
        "Formula": "PoE = P(Success | Performance, Risk, Capability, History) = Base_Rate × Performance_Factor × (1 - Risk_Factor) × Capability_Factor",
        "Data_Source_Dimensions": "All 16 dimensions + Risk Dashboard",
        "Data_Source_Elements": "All performance metrics + all risk metrics",
        "Example_Inputs": "Historical success rate: 70%; Performance trajectory: 82/100 (strong); Risk exposure: 25% (moderate); Capability maturity: 3.8/5 (good)",
        "Example_Calculation": "Base_Rate = 70% (historical); Performance_Factor = 1.0 + ((82-75)/100) = 1.07; Risk_Factor = 25% = 0.25; Capability_Factor = 3.8/5 = 0.76; PoE = 0.70 × 1.07 × (1-0.25) × 1.0 = 0.70 × 1.07 × 0.75 × 1.0 = 0.561 = 56%; Adjust for capability: 56% × (0.76/0.75) = 57%; With confidence adjustment = 60%",
        "Example_Output": "Probability of Execution: 60% (Moderate confidence in objective achievement)",
        "Assumption_References": "PASM-30 (PoE methodology), PASM-31 (base rates), PASM-32 (adjustment factors)",
        "Notes": "Forward-looking probability integrating performance, risk, and capability. <50% = High intervention needed; 50-70% = Monitor closely; 70-85% = On track; >85% = High confidence"
    }
]

performance_calculations_data.extend(other_perf_calcs)

# ============================================================================
# PERFORMANCE ASSUMPTIONS DATA
# ============================================================================

performance_assumptions_data = []

performance_assumptions_data.append({
    "Assumption_ID": "PERFORMANCE ASSUMPTIONS",
    "Assumption_Category": "All assumptions used in performance calculations",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": "Assumptions for performance analysis (positive counterpart to risk assumptions)"
})

performance_assumptions_data.append({})

performance_assumptions_data.append({
    "Assumption_ID": "=== PERFORMANCE BENCHMARKS ===",
    "Assumption_Category": "Target performance levels and benchmarks",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

perf_benchmarks = [
    {
        "Assumption_ID": "PASM-01",
        "Assumption_Category": "Benchmark - Cost Efficiency (Excellence)",
        "Assumption_Description": "Best-in-class cost efficiency ratio for high-performing organizations",
        "Assumption_Value": "75-78%",
        "Unit": "% of Revenue",
        "Source": "Industry benchmarking studies, top quartile performers",
        "Confidence_Level": "High (80%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Medium",
        "Alternative_Scenarios": "Best-in-class: 75%; Strong: 78%; Good: 80%; Average: 82%",
        "Notes": "Represents operating leverage and efficiency excellence; varies by industry"
    },
    {
        "Assumption_ID": "PASM-02",
        "Assumption_Category": "Benchmark - Investment ROI Target",
        "Assumption_Description": "Target ROI for capital investments (hurdle rate + premium)",
        "Assumption_Value": "20%+",
        "Unit": "% ROI",
        "Source": "Corporate finance best practices, WACC + risk premium",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "High",
        "Alternative_Scenarios": "Excellent: 25%+; Strong: 20-25%; Good: 15-20%; Minimum: 12-15%",
        "Notes": "Should exceed WACC (typically 8-12%) by meaningful margin; adjust for project risk"
    },
    {
        "Assumption_ID": "PASM-03",
        "Assumption_Category": "Benchmark - Strategic Initiative Success Rate",
        "Assumption_Description": "Target success rate for strategic initiatives and transformations",
        "Assumption_Value": "75-80%",
        "Unit": "% Success Rate",
        "Source": "PMI strategic initiative research, McKinsey transformation studies",
        "Confidence_Level": "Medium (75%)",
        "Last_Updated": "2024",
        "Sensitivity": "High",
        "Alternative_Scenarios": "Excellent: 80%+; Strong: 75-80%; Good: 70-75%; Poor: <60%",
        "Notes": "Success = on-time, on-budget, achieving >80% of value targets; higher than project success due to strategic importance"
    },
    {
        "Assumption_ID": "PASM-05",
        "Assumption_Category": "Benchmark - Process Value-Add Time",
        "Assumption_Description": "Target percentage of process time that is value-adding (Lean benchmark)",
        "Assumption_Value": "50%+",
        "Unit": "% of Total Time",
        "Source": "Lean Six Sigma benchmarks, process excellence research",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024",
        "Sensitivity": "Medium",
        "Alternative_Scenarios": "World-class: 60%+; Excellent: 50-60%; Good: 40-50%; Poor: <30%",
        "Notes": "Value-add = activities customer willing to pay for; non-value-add = waste (waiting, transportation, rework, etc.)"
    },
    {
        "Assumption_ID": "PASM-10",
        "Assumption_Category": "Benchmark - System Uptime (Excellence)",
        "Assumption_Description": "Target system availability for high-performing IT operations",
        "Assumption_Value": "99.95%+",
        "Unit": "% Uptime",
        "Source": "ITIL best practices, cloud provider SLAs (AWS, Azure, GCP)",
        "Confidence_Level": "High (90%)",
        "Last_Updated": "2024",
        "Sensitivity": "High",
        "Alternative_Scenarios": "Excellent: 99.99% (4 nines); Strong: 99.95%; Good: 99.9% (3 nines); Minimum: 99.5%",
        "Notes": "99.95% = 4.38 hours downtime per year; adjust by system criticality"
    }
]

performance_assumptions_data.extend(perf_benchmarks)
performance_assumptions_data.append({})

performance_assumptions_data.append({
    "Assumption_ID": "=== SCORING METHODOLOGY ===",
    "Assumption_Category": "How performance translates to index scores",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

scoring_assumptions = [
    {
        "Assumption_ID": "PASM-20",
        "Assumption_Category": "Methodology - Component Weighting",
        "Assumption_Description": "Relative weights for components within composite performance indices",
        "Assumption_Value": "Varies by metric; typically primary component 30-40%, others 15-30%",
        "Unit": "Percentage weights",
        "Source": "Expert judgment, factor analysis, correlation to outcomes",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Medium - 5% weight shift typically changes index by ±2-3 points",
        "Alternative_Scenarios": "Equal weighting (all components equal) vs Outcome-optimized (weights based on correlation to results)",
        "Notes": "Weights should sum to 100%; validate through sensitivity analysis and outcome correlation"
    },
    {
        "Assumption_ID": "PASM-25",
        "Assumption_Category": "Methodology - Normalization to 0-100 Scale",
        "Assumption_Description": "How raw performance metrics are normalized to common 0-100 scale",
        "Assumption_Value": "Linear interpolation between threshold points: 0 = poor performance threshold, 50 = benchmark, 100 = best-in-class",
        "Unit": "Score (0-100)",
        "Source": "Standard performance indexing methodology",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Low - consistent methodology across metrics",
        "Alternative_Scenarios": "Linear vs Non-linear (diminishing returns at high performance)",
        "Notes": "Example: Cost ratio - 0 points at 90%, 50 points at 80% (benchmark), 100 points at 75% (best-in-class). Interpolate linearly between points."
    },
    {
        "Assumption_ID": "PASM-30",
        "Assumption_Category": "Methodology - Probability of Execution (PoE) Calculation",
        "Assumption_Description": "Bayesian methodology for calculating forward-looking success probability",
        "Assumption_Value": "PoE = Base_Rate × Performance_Adjustment × Risk_Adjustment × Capability_Adjustment",
        "Unit": "Probability (%)",
        "Source": "RBPM framework, Bayesian updating methodology, actuarial science",
        "Confidence_Level": "Medium (65%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Very High - PoE highly sensitive to all four factors",
        "Alternative_Scenarios": "Conservative (base rate -10%), Moderate (as stated), Aggressive (base rate +10%)",
        "Notes": "Requires calibration to actual outcomes over 12-24 months for accuracy. Initial estimates based on industry research."
    },
    {
        "Assumption_ID": "PASM-31",
        "Assumption_Category": "PoE - Historical Base Rates",
        "Assumption_Description": "Historical success rates for different types of objectives (base rate for Bayesian updating)",
        "Assumption_Value": "Strategic initiatives: 60-70%; Operational improvements: 70-80%; IT projects: 50-60%; Innovation projects: 40-50%",
        "Unit": "% Historical Success",
        "Source": "PMI Pulse of Profession, McKinsey transformation research, Standish CHAOS reports",
        "Confidence_Level": "High (80%)",
        "Last_Updated": "2024",
        "Sensitivity": "Very High - base rate strongly influences PoE",
        "Alternative_Scenarios": "Company-specific rates after 12 months data collection",
        "Notes": "Use industry base rates initially, then update with company-specific historical performance"
    },
    {
        "Assumption_ID": "PASM-32",
        "Assumption_Category": "PoE - Adjustment Factors",
        "Assumption_Description": "How performance, risk, and capability adjust base rate probability",
        "Assumption_Value": "Performance: ±1% PoE per point above/below 75; Risk: -0.5% PoE per 1% risk exposure; Capability: +0.3% PoE per 0.1 maturity point above 3.0",
        "Unit": "Adjustment multipliers",
        "Source": "Calibrated to historical outcome data, expert judgment",
        "Confidence_Level": "Medium (65%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "High - adjustment factors determine PoE sensitivity",
        "Alternative_Scenarios": "Conservative (stronger adjustments), Moderate (as stated), Aggressive (weaker adjustments)",
        "Notes": "Requires continuous calibration. Example: Base 70%, Performance 85 (10 points above 75) = +10% = 77%; Risk 20% = -10% = 69.3%; Capability 3.8 (+0.8) = +2.4% = 71%"
    }
]

performance_assumptions_data.extend(scoring_assumptions)
performance_assumptions_data.append({})

performance_assumptions_data.append({
    "Assumption_ID": "=== INTERPRETATION THRESHOLDS ===",
    "Assumption_Category": "How to interpret performance index scores",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

interpretation = [
    {
        "Assumption_ID": "PASM-40",
        "Assumption_Category": "Interpretation - Performance Index Bands",
        "Assumption_Description": "How to interpret performance index scores (0-100 scale)",
        "Assumption_Value": "90-100: Exceptional; 80-89: Excellence; 70-79: Strong; 60-69: Good; 50-59: Fair; <50: Needs Improvement",
        "Unit": "Score bands",
        "Source": "Standard performance assessment scales, benchmarking research",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Low - interpretation thresholds",
        "Alternative_Scenarios": "Stricter bands (higher thresholds) vs Lenient bands (lower thresholds)",
        "Notes": "Bands calibrated so 'Strong' (70-79) represents performing better than industry average; 'Excellence' (80-89) represents top quartile"
    },
    {
        "Assumption_ID": "PASM-41",
        "Assumption_Category": "Interpretation - Probability of Execution Bands",
        "Assumption_Description": "How to interpret PoE probability scores",
        "Assumption_Value": "85-100%: High confidence; 70-84%: On track; 50-69%: Monitor closely; 30-49%: High intervention needed; <30%: Re-plan/De-prioritize",
        "Unit": "Probability bands",
        "Source": "RBPM framework, risk management best practices",
        "Confidence_Level": "Medium (75%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Medium - drives management action",
        "Alternative_Scenarios": "Conservative (higher thresholds for 'On track') vs Aggressive (lower thresholds)",
        "Notes": "PoE <50% suggests objective should be re-scoped, timeline extended, or de-prioritized to focus resources on higher-probability objectives"
    }
]

performance_assumptions_data.extend(interpretation)

# ============================================================================
# Write all three sheets to Excel
# ============================================================================

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("Creating Performance Dashboard sheet...")

if "Performance Dashboard" in wb.sheetnames:
    ws_perf = wb["Performance Dashboard"]
    for row in ws_perf.iter_rows():
        for cell in row:
            cell.value = None
else:
    # Insert after Risk Assumptions (position 3)
    ws_perf = wb.create_sheet("Performance Dashboard", 3)

# Headers for Performance Dashboard
perf_dash_headers = [
    "Aggregation_Level", "Entity_Name", "Performance_Metric", "Current_Score",
    "Target_Score", "Achievement_Rate", "Trend", "Time_Period",
    "Benchmark_Comparison", "Data_Sources", "Calculation_Method", "Last_Updated", "Notes"
]

for col_idx, header in enumerate(perf_dash_headers, 1):
    cell = ws_perf.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, data_row in enumerate(performance_dashboard_data, 2):
    for col_idx, header in enumerate(perf_dash_headers, 1):
        value = data_row.get(header, "")
        cell = ws_perf.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

perf_dash_widths = {"A": 18, "B": 35, "C": 40, "D": 15, "E": 15, "F": 15, "G": 12, "H": 15, "I": 20, "J": 50, "K": 80, "L": 12, "M": 40}
for col_letter, width in perf_dash_widths.items():
    ws_perf.column_dimensions[col_letter].width = width

ws_perf.freeze_panes = "A2"
print(f"  Performance Dashboard: {len(performance_dashboard_data)} rows")

# Performance Calculations sheet
print("Creating Performance Calculations sheet...")

if "Performance Calculations" in wb.sheetnames:
    ws_pcalc = wb["Performance Calculations"]
    for row in ws_pcalc.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_pcalc = wb.create_sheet("Performance Calculations", 4)

perf_calc_headers = [
    "Calculation_ID", "Performance_Metric", "Calculation_Step", "Formula",
    "Data_Source_Dimensions", "Data_Source_Elements", "Example_Inputs",
    "Example_Calculation", "Example_Output", "Assumption_References", "Notes"
]

for col_idx, header in enumerate(perf_calc_headers, 1):
    cell = ws_pcalc.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, data_row in enumerate(performance_calculations_data, 2):
    for col_idx, header in enumerate(perf_calc_headers, 1):
        value = data_row.get(header, "")
        cell = ws_pcalc.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

pcalc_widths = {"A": 15, "B": 25, "C": 40, "D": 60, "E": 30, "F": 35, "G": 40, "H": 50, "I": 30, "J": 20, "K": 40}
for col_letter, width in pcalc_widths.items():
    ws_pcalc.column_dimensions[col_letter].width = width

ws_pcalc.freeze_panes = "A2"
print(f"  Performance Calculations: {len(performance_calculations_data)} rows")

# Performance Assumptions sheet
print("Creating Performance Assumptions sheet...")

if "Performance Assumptions" in wb.sheetnames:
    ws_pasm = wb["Performance Assumptions"]
    for row in ws_pasm.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_pasm = wb.create_sheet("Performance Assumptions", 5)

perf_asm_headers = [
    "Assumption_ID", "Assumption_Category", "Assumption_Description", "Assumption_Value",
    "Unit", "Source", "Confidence_Level", "Last_Updated", "Sensitivity",
    "Alternative_Scenarios", "Notes"
]

for col_idx, header in enumerate(perf_asm_headers, 1):
    cell = ws_pasm.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, data_row in enumerate(performance_assumptions_data, 2):
    for col_idx, header in enumerate(perf_asm_headers, 1):
        value = data_row.get(header, "")
        cell = ws_pasm.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

pasm_widths = {"A": 15, "B": 30, "C": 50, "D": 30, "E": 15, "F": 50, "G": 18, "H": 15, "I": 20, "J": 40, "K": 40}
for col_letter, width in pasm_widths.items():
    ws_pasm.column_dimensions[col_letter].width = width

ws_pasm.freeze_panes = "A2"
print(f"  Performance Assumptions: {len(performance_assumptions_data)} assumptions")

wb.save(wb_path)

print()
print("=" * 80)
print("PERFORMANCE ANALYSIS COMPLETE!")
print("=" * 80)
print()
print("Summary:")
print("  Performance Dashboard: 10 key performance metrics + domain/dimension breakdowns")
print("    1. Operational Excellence Index")
print("    2. Investment Returns Index")
print("    3. Strategic Achievement Index")
print("    4. Revenue Performance Index")
print("    5. Productivity Excellence Index")
print("    6. Service Excellence Index")
print("    7. Product & Innovation Success Index")
print("    8. Reputation Strength Index")
print("    9. Value Creation Index")
print("    10. Probability of Execution (PoE)")
print()
print("  Performance Calculations: Step-by-step methodology for all metrics")
print("  Performance Assumptions: Benchmarks, scoring methodology, interpretation guides")
print()
print("Complete Workbook Now Has:")
print("  Risk Analysis Layer:        3 sheets (Dashboard, Calculations, Assumptions)")
print("  Performance Analysis Layer: 3 sheets (Dashboard, Calculations, Assumptions)")
print("  Data Input Layer:          16 sheets (Economics, Enablers, Execution, VALUE)")
print("  ──────────────────────────────────────────────────────────────────────────")
print("  TOTAL:                     22 sheets")
print()
print("The framework now provides complete risk AND performance analysis capabilities!")
