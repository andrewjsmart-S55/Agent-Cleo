"""
Analyze the Brand sheet to see a completed example of the 3-level structure
"""

import openpyxl
from openpyxl import load_workbook

# Load the workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path, data_only=False)
ws = wb["Brand"]

print("=" * 120)
print("BRAND SHEET ANALYSIS - Example of Completed 3-Level Structure")
print("=" * 120)
print()

# Get headers
headers = []
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col_idx).value
    headers.append(str(header) if header else f"Col{col_idx}")

print("COLUMN HEADERS:")
for idx, header in enumerate(headers, 1):
    print(f"  {idx:2d}. {header}")
print()

print("=" * 120)
print("ROW-BY-ROW DATA:")
print("=" * 120)
print()

for row_idx in range(1, ws.max_row + 1):
    print(f"ROW {row_idx}:")
    print("-" * 120)

    has_data = False
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value:
            has_data = True
            header = headers[col_idx - 1]
            # Format the value
            value_str = str(cell_value)

            print(f"   [{col_idx:2d}] {header:<45}: {value_str}")

    if not has_data:
        print("   [Empty row]")

    print()

print("=" * 120)
print("ANALYSIS COMPLETE")
print("=" * 120)
