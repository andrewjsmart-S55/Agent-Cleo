"""
Detailed Analysis of Financials Sheet - Full Data Review
"""

import openpyxl
from openpyxl import load_workbook

# Load the workbook WITHOUT data_only to see formulas and all content
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path, data_only=False)
ws = wb["Financials"]

print("=" * 120)
print("FINANCIALS SHEET - COMPLETE DATA REVIEW")
print("=" * 120)
print()

# Create a formatted table view
print("FULL DATA TABLE VIEW:")
print("=" * 120)
print()

# Get headers
headers = []
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col_idx).value
    headers.append(str(header) if header else f"Col{col_idx}")

# Print headers with column numbers
print("Column Reference:")
for idx, header in enumerate(headers, 1):
    print(f"  {idx:2d}. {header}")
print()
print("-" * 120)
print()

# Print all data rows
for row_idx in range(1, ws.max_row + 1):
    if row_idx == 1:
        print(f"{'ROW':<4} | ", end="")
        for col_idx in range(1, min(6, ws.max_column + 1)):  # First 5 columns
            print(f"{headers[col_idx-1]:<25} | ", end="")
        print()
        print("-" * 120)
    else:
        print(f"{row_idx:<4} | ", end="")
        for col_idx in range(1, min(6, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            value_str = str(cell_value) if cell_value else ""
            if len(value_str) > 25:
                value_str = value_str[:22] + "..."
            print(f"{value_str:<25} | ", end="")
        print()

print()
print("-" * 120)
print()

# Detailed row-by-row breakdown
print("DETAILED ROW-BY-ROW ANALYSIS:")
print("=" * 120)
print()

for row_idx in range(1, ws.max_row + 1):
    print(f"ROW {row_idx}:")
    print("-" * 120)

    has_data = False
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value:
            has_data = True
            header = headers[col_idx - 1]
            # Format the value
            value_str = str(cell_value)
            if len(value_str) > 100:
                value_str = value_str[:97] + "..."

            print(f"   [{col_idx:2d}] {header:<40}: {value_str}")

    if not has_data:
        print("   [Empty row]")

    print()

# Summary of 3-level structure
print()
print("=" * 120)
print("3-LEVEL STRUCTURE SUMMARY:")
print("=" * 120)
print()

level_data = {1: [], 2: [], 3: []}

for row_idx in range(2, ws.max_row + 1):
    level = ws.cell(row=row_idx, column=1).value  # Level column
    hierarchy = ws.cell(row=row_idx, column=2).value  # Hierarchy column

    if level in [1, 2, 3]:
        level_data[level].append({
            'row': row_idx,
            'hierarchy': hierarchy if hierarchy else "[Not specified]"
        })

for level in [1, 2, 3]:
    print(f"LEVEL {level} ({len(level_data[level])} items):")
    print("-" * 80)
    for item in level_data[level]:
        print(f"   Row {item['row']:2d}: {item['hierarchy']}")
    print()

# Check for data in columns 4-14
print()
print("DATA AVAILABILITY CHECK (Columns 4-14):")
print("-" * 120)
print()

for col_idx in range(4, 15):
    header = headers[col_idx - 1]
    rows_with_data = []

    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        if cell_value:
            rows_with_data.append(row_idx)

    if rows_with_data:
        print(f"Column {col_idx:2d} - {header:<45}: Data in rows {rows_with_data}")
    else:
        print(f"Column {col_idx:2d} - {header:<45}: [No data]")

print()
print("=" * 120)
print("ANALYSIS COMPLETE")
print("=" * 120)
