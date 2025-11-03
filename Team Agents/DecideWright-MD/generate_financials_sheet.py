"""
Generate Complete Financials Sheet
Expert Senior Business Analyst - Strategy Execution & Risk Management

3-Level hierarchical data collection structure for multivariate Bayesian analysis
linking the Financials dimension of the Value Orchestration Canvas to both:
- Quantitative Risk Analysis
- Performance/Strategy Execution Analysis
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)
ws = wb["Financials"]

# Clear existing content (keep headers)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Financials Data Structure - 3 Levels
# Level 1: Financial (Dimension)
# Level 2: Elements (Costs, Revenue, Profit, Capital, Liquidity, Accuracy & Transparency)
# Level 3: Sub-elements (specific metrics and data points)

financials_data = [
    # LEVEL 1: DIMENSION
    {
        "Level": 1,
        "Hierarchy": "Financials",
        "Description": "Financial health, performance, and sustainability of the organization",
        "Business Drivers": "Value Creation & Preservation",
        "Business Drivers Description": "Financial capability to execute strategy while managing risk exposure",
        "Performance Factors": "Financial Strength & Efficiency",
        "Performance Factors Description": "Ability to fund operations, invest in growth, and deliver returns",
        "Risk Factors": "Financial Instability",
        "Risk Factors Description": "Exposure to financial distress, liquidity crisis, or value destruction",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of all financial elements - auto-calculated from Level 2"
    },

    # LEVEL 2: COSTS
    {
        "Level": 2,
        "Hierarchy": "Financials - Costs",
        "Description": "Total cost structure and cost management effectiveness",
        "Business Drivers": "Operational Efficiency",
        "Business Drivers Description": "Ability to deliver products/services at competitive cost levels",
        "Performance Factors": "Cost Optimization",
        "Performance Factors Description": "Efficient resource utilization and cost control",
        "Risk Factors": "Cost Overrun & Inflation",
        "Risk Factors Description": "Unplanned cost increases eroding margins and profitability",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of cost sub-elements - auto-calculated from Level 3"
    },

    # LEVEL 3: COSTS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Financials - Costs - Operating Expenses (Opex)",
        "Description": "Day-to-day operational costs required to run the business",
        "Business Drivers": "Business Activity Volume",
        "Business Drivers Description": "Scale of operations directly drives operating cost levels",
        "Performance Factors": "Opex Efficiency Ratio",
        "Performance Factors Description": "Operating expenses as % of revenue - lower indicates better efficiency",
        "Risk Factors": "Opex Overrun",
        "Risk Factors Description": "Operating costs exceeding budget due to inflation, inefficiency, or scope creep",
        "Metric": "Total Annual Operating Expenses",
        "Metric Description": "Sum of all recurring operational costs (excluding COGS, depreciation, interest)",
        "Unit": "Currency (aligned to Base Data)",
        "Target": "Industry benchmark ±10%",
        "Instructions": "Include: salaries, rent, utilities, marketing, admin, IT operations, travel, professional fees"
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Costs - Cost of Goods Sold (COGS)",
        "Description": "Direct costs attributable to production of goods/services sold",
        "Business Drivers": "Revenue Volume",
        "Business Drivers Description": "Direct correlation between units sold and production costs",
        "Performance Factors": "Gross Margin",
        "Performance Factors Description": "Revenue minus COGS as % of revenue - measures pricing power and efficiency",
        "Risk Factors": "Input Cost Inflation",
        "Risk Factors Description": "Supplier price increases, commodity volatility, or supply chain disruption",
        "Metric": "Total Annual COGS",
        "Metric Description": "Direct materials, direct labor, and manufacturing overhead",
        "Unit": "Currency",
        "Target": "Maintain or improve gross margin % YoY",
        "Instructions": "Include only variable costs directly tied to production/delivery. Exclude fixed overhead."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Costs - Fixed vs Variable Cost Ratio",
        "Description": "Proportion of costs that are fixed vs variable with volume",
        "Business Drivers": "Business Model Structure",
        "Business Drivers Description": "Asset intensity and operating leverage determine cost structure",
        "Performance Factors": "Operating Leverage",
        "Performance Factors Description": "Higher fixed costs = higher sensitivity to volume changes (risk + opportunity)",
        "Risk Factors": "Fixed Cost Burden",
        "Risk Factors Description": "High fixed costs create breakeven risk if volumes decline",
        "Metric": "Fixed Costs as % of Total Costs",
        "Metric Description": "Percentage of costs that don't vary with output (rent, salaries, depreciation)",
        "Unit": "Percentage",
        "Target": "Aligned to business model strategy",
        "Instructions": "Fixed costs / (Fixed costs + Variable costs) × 100. Track quarterly."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Costs - Cost Predictability",
        "Description": "Volatility and forecastability of cost base",
        "Business Drivers": "Cost Control Maturity",
        "Business Drivers Description": "Quality of budgeting, forecasting, and cost management processes",
        "Performance Factors": "Forecast Accuracy",
        "Performance Factors Description": "Actual costs vs budget variance - tighter variance = better control",
        "Risk Factors": "Cost Volatility",
        "Risk Factors Description": "Unpredictable cost swings create budgeting risk and margin uncertainty",
        "Metric": "Actual vs Budget Cost Variance",
        "Metric Description": "Percentage difference between actual and budgeted costs",
        "Unit": "Percentage",
        "Target": "±5% variance",
        "Instructions": "Calculate monthly: (Actual Costs - Budgeted Costs) / Budgeted Costs × 100"
    },

    # LEVEL 2: REVENUE
    {
        "Level": 2,
        "Hierarchy": "Financials - Revenue",
        "Description": "Top-line revenue generation and quality",
        "Business Drivers": "Market Demand & Sales Execution",
        "Business Drivers Description": "Ability to attract customers and convert demand into revenue",
        "Performance Factors": "Revenue Growth & Quality",
        "Performance Factors Description": "Sustainable, diversified revenue growth at attractive margins",
        "Risk Factors": "Revenue Shortfall & Concentration",
        "Risk Factors Description": "Failure to meet revenue targets or over-dependence on few sources",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of revenue sub-elements - auto-calculated from Level 3"
    },

    # LEVEL 3: REVENUE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Financials - Revenue - Total Revenue",
        "Description": "Gross revenue before any deductions",
        "Business Drivers": "Commercial Success",
        "Business Drivers Description": "Market acceptance of value proposition translates to revenue",
        "Performance Factors": "Revenue Growth Rate",
        "Performance Factors Description": "Year-over-year percentage increase in total revenue",
        "Risk Factors": "Revenue Stagnation",
        "Risk Factors Description": "Flat or declining revenue indicating market share loss or market maturity",
        "Metric": "Annual Gross Revenue",
        "Metric Description": "Total revenue from all sources before returns, allowances, or discounts",
        "Unit": "Currency",
        "Target": "Strategic plan target (typically 10-30% CAGR for growth companies)",
        "Instructions": "Sum all revenue streams. Use accrual accounting. Report in consistent currency."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Revenue - Revenue Diversification",
        "Description": "Spread of revenue across products, customers, and geographies",
        "Business Drivers": "Portfolio Strategy",
        "Business Drivers Description": "Deliberate choice to serve multiple segments reduces concentration risk",
        "Performance Factors": "Diversification Index",
        "Performance Factors Description": "Herfindahl-Hirschman Index (HHI) - lower = more diversified",
        "Risk Factors": "Revenue Concentration Risk",
        "Risk Factors Description": "Over-dependence on single customer, product, or geography creates volatility",
        "Metric": "Revenue Concentration (Top 3 Sources)",
        "Metric Description": "Percentage of total revenue from top 3 customers, products, or regions (highest)",
        "Unit": "Percentage",
        "Target": "<30% from any single source",
        "Instructions": "Calculate for customers, products, and geographies. Report highest concentration."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Revenue - Revenue Quality",
        "Description": "Predictability, recurrence, and margin quality of revenue",
        "Business Drivers": "Business Model Economics",
        "Business Drivers Description": "Recurring revenue models provide higher quality, more predictable revenue",
        "Performance Factors": "Recurring Revenue Ratio",
        "Performance Factors Description": "Percentage of revenue from subscriptions, contracts, or repeat customers",
        "Risk Factors": "Revenue Unpredictability",
        "Risk Factors Description": "High proportion of one-time or spot revenue creates forecasting uncertainty",
        "Metric": "Recurring Revenue as % of Total",
        "Metric Description": "Contracted, subscription, or repeat customer revenue / total revenue",
        "Unit": "Percentage",
        "Target": ">60% for SaaS/subscription models; >40% for B2B; varies by business model",
        "Instructions": "Include only revenue with contractual commitment or historical >80% repeat rate"
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Revenue - Revenue at Risk",
        "Description": "Portion of revenue vulnerable to identified threats",
        "Business Drivers": "External & Internal Risk Events",
        "Business Drivers Description": "Market changes, competition, operational failures can threaten revenue",
        "Performance Factors": "Revenue Defense Capability",
        "Performance Factors Description": "Strength of competitive moat, customer relationships, and contract terms",
        "Risk Factors": "Revenue Loss Exposure",
        "Risk Factors Description": "Identifiable threats that could materially reduce revenue",
        "Metric": "Estimated Revenue at Risk (12-month horizon)",
        "Metric Description": "Sum of revenue exposed to known risks weighted by probability",
        "Unit": "Currency",
        "Target": "<10% of annual revenue",
        "Instructions": "Assess: contract renewals at risk, competitive threats, market shifts, key customer churn"
    },

    # LEVEL 2: PROFIT
    {
        "Level": 2,
        "Hierarchy": "Financials - Profit",
        "Description": "Bottom-line profitability and margin sustainability",
        "Business Drivers": "Revenue - Costs Spread",
        "Business Drivers Description": "Ability to generate revenue in excess of all costs",
        "Performance Factors": "Margin Excellence",
        "Performance Factors Description": "Achieving target margins across gross, operating, and net levels",
        "Risk Factors": "Margin Compression",
        "Risk Factors Description": "Squeeze on profitability from cost inflation or price pressure",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of profit sub-elements - auto-calculated from Level 3"
    },

    # LEVEL 3: PROFIT SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Financials - Profit - Gross Profit Margin",
        "Description": "Profitability after direct costs of production",
        "Business Drivers": "Pricing Power & Production Efficiency",
        "Business Drivers Description": "Ability to price above cost of goods and maintain efficiency",
        "Performance Factors": "Gross Margin %",
        "Performance Factors Description": "(Revenue - COGS) / Revenue - measures unit economics",
        "Risk Factors": "Margin Erosion",
        "Risk Factors Description": "Price competition or cost inflation reducing gross profitability",
        "Metric": "Gross Profit Margin %",
        "Metric Description": "Percentage of revenue remaining after direct production costs",
        "Unit": "Percentage",
        "Target": "Industry-specific: SaaS 70-80%, Manufacturing 30-40%, Retail 20-30%",
        "Instructions": "(Revenue - COGS) / Revenue × 100. Calculate quarterly, track trend."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Profit - Operating Profit (EBIT)",
        "Description": "Profitability from core operations before interest and tax",
        "Business Drivers": "Operational Execution",
        "Business Drivers Description": "Efficiency of converting revenue to operating profit",
        "Performance Factors": "Operating Margin %",
        "Performance Factors Description": "EBIT as % of revenue - measures operational efficiency",
        "Risk Factors": "Operating Leverage Risk",
        "Risk Factors Description": "Small revenue changes create large profit swings if operating leverage is high",
        "Metric": "EBIT Margin %",
        "Metric Description": "(Revenue - COGS - Opex) / Revenue",
        "Unit": "Percentage",
        "Target": "Varies by industry: Tech 20-30%, Services 10-15%, Manufacturing 5-10%",
        "Instructions": "Exclude non-operating items. Report both $ amount and % margin."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Profit - Net Profit Margin",
        "Description": "Bottom-line profitability after all expenses",
        "Business Drivers": "Total Financial Management",
        "Business Drivers Description": "Comprehensive efficiency including financing and tax management",
        "Performance Factors": "Net Margin %",
        "Performance Factors Description": "Net income as % of revenue - ultimate profitability measure",
        "Risk Factors": "Profitability Failure",
        "Risk Factors Description": "Operating losses or negative net income",
        "Metric": "Net Profit Margin %",
        "Metric Description": "Net Income / Revenue after all expenses, interest, and taxes",
        "Unit": "Percentage",
        "Target": ">5% for mature businesses; breakeven to positive for growth stage",
        "Instructions": "Net Income / Total Revenue × 100. Compare to prior periods and peers."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Profit - Earnings Volatility",
        "Description": "Stability and predictability of profit generation",
        "Business Drivers": "Business Model Stability",
        "Business Drivers Description": "Recurring revenue and efficient operations create stable earnings",
        "Performance Factors": "Earnings Consistency",
        "Performance Factors Description": "Low variance in quarterly/annual earnings",
        "Risk Factors": "Earnings Unpredictability",
        "Risk Factors Description": "High earnings volatility creates uncertainty for investors and planning",
        "Metric": "Earnings Standard Deviation (3-year)",
        "Metric Description": "Standard deviation of quarterly EBIT over 12 quarters",
        "Unit": "Currency (absolute) or Percentage (CoV)",
        "Target": "Coefficient of Variation <20%",
        "Instructions": "Calculate StdDev of last 12 quarters EBIT. Lower = more stable."
    },

    # LEVEL 2: CAPITAL
    {
        "Level": 2,
        "Hierarchy": "Financials - Capital",
        "Description": "Capital structure, efficiency, and investment capacity",
        "Business Drivers": "Growth Funding & Capital Allocation",
        "Business Drivers Description": "Ability to fund operations, growth, and strategic initiatives",
        "Performance Factors": "Capital Efficiency",
        "Performance Factors Description": "Return on invested capital and optimal capital structure",
        "Risk Factors": "Capital Constraint & Leverage",
        "Risk Factors Description": "Insufficient capital for growth or excessive debt burden",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of capital sub-elements - auto-calculated from Level 3"
    },

    # LEVEL 3: CAPITAL SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Financials - Capital - Working Capital",
        "Description": "Short-term capital required for daily operations",
        "Business Drivers": "Operating Cycle Requirements",
        "Business Drivers Description": "Cash conversion cycle determines working capital needs",
        "Performance Factors": "Working Capital Efficiency",
        "Performance Factors Description": "Optimal level to support operations without excess tied up",
        "Risk Factors": "Working Capital Shortfall",
        "Risk Factors Description": "Insufficient working capital to meet operational obligations",
        "Metric": "Working Capital Ratio",
        "Metric Description": "Current Assets / Current Liabilities",
        "Unit": "Ratio",
        "Target": "1.5 to 2.0 (varies by industry)",
        "Instructions": "Calculate monthly. Ratio <1.0 = liquidity risk. Ratio >3.0 = inefficiency."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Capital - Capital Expenditure (Capex)",
        "Description": "Investment in long-term assets and infrastructure",
        "Business Drivers": "Growth & Maintenance Requirements",
        "Business Drivers Description": "Need to invest in capacity, technology, and asset replacement",
        "Performance Factors": "Capex Efficiency",
        "Performance Factors Description": "Return on capital invested - growth + margin improvement",
        "Risk Factors": "Capex Overrun",
        "Risk Factors Description": "Capital projects exceed budget or fail to deliver expected returns",
        "Metric": "Annual Capex as % of Revenue",
        "Metric Description": "Total capital expenditures / annual revenue",
        "Unit": "Percentage",
        "Target": "Maintenance capex: 2-5% of revenue; Growth capex: project-specific ROI >15%",
        "Instructions": "Track maintenance vs growth capex separately. Monitor project ROI."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Capital - Debt Levels",
        "Description": "Total debt obligations and leverage",
        "Business Drivers": "Financing Strategy",
        "Business Drivers Description": "Use of debt to fund growth vs equity financing",
        "Performance Factors": "Leverage Optimization",
        "Performance Factors Description": "Debt level that minimizes cost of capital without excessive risk",
        "Risk Factors": "Over-Leverage",
        "Risk Factors Description": "Excessive debt creating solvency risk and financial distress",
        "Metric": "Debt-to-EBITDA Ratio",
        "Metric Description": "Total Debt / EBITDA (last 12 months)",
        "Unit": "Ratio (x)",
        "Target": "<3x for investment grade; <5x acceptable for PE-backed; varies by industry",
        "Instructions": "Include all interest-bearing debt. Calculate quarterly. Monitor covenants."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Capital - Return on Invested Capital (ROIC)",
        "Description": "Profitability relative to total capital employed",
        "Business Drivers": "Capital Allocation Excellence",
        "Business Drivers Description": "Deploying capital to highest-return opportunities",
        "Performance Factors": "Value Creation",
        "Performance Factors Description": "ROIC > WACC = value creation; ROIC < WACC = value destruction",
        "Risk Factors": "Capital Misallocation",
        "Risk Factors Description": "Investing in low-return projects destroys shareholder value",
        "Metric": "ROIC %",
        "Metric Description": "NOPAT / (Debt + Equity)",
        "Unit": "Percentage",
        "Target": ">WACC + 5% premium",
        "Instructions": "NOPAT = Net Operating Profit After Tax. Calculate annually. Compare to WACC."
    },

    # LEVEL 2: LIQUIDITY
    {
        "Level": 2,
        "Hierarchy": "Financials - Liquidity",
        "Description": "Cash flow generation and financial flexibility",
        "Business Drivers": "Cash Conversion & Treasury Management",
        "Business Drivers Description": "Ability to generate cash and maintain liquidity buffers",
        "Performance Factors": "Cash Flow Strength",
        "Performance Factors Description": "Strong operating cash flow and adequate liquidity reserves",
        "Risk Factors": "Liquidity Crisis",
        "Risk Factors Description": "Inability to meet short-term obligations or fund operations",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of liquidity sub-elements - auto-calculated from Level 3"
    },

    # LEVEL 3: LIQUIDITY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Financials - Liquidity - Operating Cash Flow",
        "Description": "Cash generated from core business operations",
        "Business Drivers": "Profitability + Working Capital Management",
        "Business Drivers Description": "Converting earnings to cash through efficient operations",
        "Performance Factors": "Cash Generation",
        "Performance Factors Description": "Strong positive OCF funds growth without external financing",
        "Risk Factors": "Cash Burn",
        "Risk Factors Description": "Negative operating cash flow depleting reserves",
        "Metric": "Operating Cash Flow (OCF)",
        "Metric Description": "Cash from operations before investing and financing activities",
        "Unit": "Currency",
        "Target": ">0 for mature businesses; may be negative for high-growth startups",
        "Instructions": "From cash flow statement. Calculate OCF margin: OCF / Revenue. Track trend."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Liquidity - Cash Reserves",
        "Description": "Available cash and cash equivalents",
        "Business Drivers": "Liquidity Buffer Strategy",
        "Business Drivers Description": "Maintaining reserves for opportunities and contingencies",
        "Performance Factors": "Liquidity Adequacy",
        "Performance Factors Description": "Sufficient cash to weather disruptions and fund growth",
        "Risk Factors": "Cash Depletion",
        "Risk Factors Description": "Running out of cash before achieving profitability or next funding",
        "Metric": "Cash Runway (Months)",
        "Metric Description": "Current cash balance / average monthly cash burn",
        "Unit": "Months",
        "Target": ">12 months for early-stage; >6 months for mature businesses",
        "Instructions": "Monthly burn = Operating cash outflow / 12. Minimum acceptable: 6 months."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Liquidity - Free Cash Flow",
        "Description": "Cash available after capital expenditures",
        "Business Drivers": "Investment Discipline",
        "Business Drivers Description": "Generating cash in excess of maintenance capex requirements",
        "Performance Factors": "Cash Flow Conversion",
        "Performance Factors Description": "Percentage of EBITDA converted to free cash flow",
        "Risk Factors": "Negative Free Cash Flow",
        "Risk Factors Description": "Capex exceeding operating cash flow, requiring external funding",
        "Metric": "Free Cash Flow (FCF)",
        "Metric Description": "Operating Cash Flow - Capital Expenditures",
        "Unit": "Currency",
        "Target": "Positive and growing; FCF / Revenue >5%",
        "Instructions": "FCF = OCF - Capex. Exclude one-time items. Calculate FCF yield."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Liquidity - Cash Conversion Cycle",
        "Description": "Time to convert operations into cash",
        "Business Drivers": "Working Capital Efficiency",
        "Business Drivers Description": "Speed of collecting receivables and managing inventory/payables",
        "Performance Factors": "Cash Cycle Optimization",
        "Performance Factors Description": "Shorter cycle = faster cash generation and lower working capital needs",
        "Risk Factors": "Cash Cycle Extension",
        "Risk Factors Description": "Slow collections or excess inventory tying up cash",
        "Metric": "Cash Conversion Cycle (Days)",
        "Metric Description": "Days Inventory Outstanding + Days Sales Outstanding - Days Payable Outstanding",
        "Unit": "Days",
        "Target": "Industry-specific; minimize without harming relationships. <60 days ideal.",
        "Instructions": "DIO + DSO - DPO. Calculate quarterly. Lower is better (within reason)."
    },

    # LEVEL 2: ACCURACY & TRANSPARENCY
    {
        "Level": 2,
        "Hierarchy": "Financials - Accuracy & Transparency",
        "Description": "Quality, reliability, and transparency of financial reporting",
        "Business Drivers": "Financial Governance & Controls",
        "Business Drivers Description": "Robust financial systems, controls, and reporting discipline",
        "Performance Factors": "Financial Reporting Excellence",
        "Performance Factors Description": "Accurate, timely, and transparent financial information",
        "Risk Factors": "Financial Reporting Risk",
        "Risk Factors Description": "Inaccurate reporting, fraud, or loss of stakeholder confidence",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of accuracy & transparency sub-elements"
    },

    # LEVEL 3: ACCURACY & TRANSPARENCY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Financials - Accuracy & Transparency - Financial Close Cycle",
        "Description": "Time required to close books and produce financial statements",
        "Business Drivers": "Process Automation & Control Quality",
        "Business Drivers Description": "Efficient systems and controls enable fast, accurate closes",
        "Performance Factors": "Close Efficiency",
        "Performance Factors Description": "Fast close indicates strong processes and systems",
        "Risk Factors": "Delayed Financial Information",
        "Risk Factors Description": "Slow closes mean management operates with stale data",
        "Metric": "Month-End Close Cycle Time",
        "Metric Description": "Business days from month-end to final financial statements",
        "Unit": "Days",
        "Target": "Best-in-class: 3-5 days; Acceptable: 7-10 days; Poor: >15 days",
        "Instructions": "Measure each month. Target continuous improvement. Automate reconciliations."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Accuracy & Transparency - Financial Audit Results",
        "Description": "External audit opinion and findings",
        "Business Drivers": "Control Environment Strength",
        "Business Drivers Description": "Strong controls result in clean audits with no material weaknesses",
        "Performance Factors": "Audit Opinion Quality",
        "Performance Factors Description": "Unqualified opinion with minimal adjustments",
        "Risk Factors": "Audit Findings Risk",
        "Risk Factors Description": "Material weaknesses or qualified opinion damages credibility",
        "Metric": "Audit Opinion & Material Weaknesses",
        "Metric Description": "Clean/unqualified opinion with zero material weaknesses",
        "Unit": "Categorical: Clean / Qualified / Adverse + Count of material weaknesses",
        "Target": "Clean opinion, zero material weaknesses",
        "Instructions": "Report latest audit results. Track remediation of any findings."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Accuracy & Transparency - Forecast Accuracy",
        "Description": "Reliability of financial forecasts and budgets",
        "Business Drivers": "Planning & Budgeting Maturity",
        "Business Drivers Description": "Quality of assumptions, data, and planning processes",
        "Performance Factors": "Predictive Capability",
        "Performance Factors Description": "Actual results closely match forecasts/budgets",
        "Risk Factors": "Forecast Unreliability",
        "Risk Factors Description": "Large variances undermine planning and stakeholder confidence",
        "Metric": "Revenue & EBITDA Forecast Variance",
        "Metric Description": "Percentage difference between forecast and actual",
        "Unit": "Percentage",
        "Target": "±5% variance for mature businesses; ±10% for high-growth",
        "Instructions": "(Actual - Forecast) / Forecast × 100. Track for revenue, costs, EBITDA."
    },
    {
        "Level": 3,
        "Hierarchy": "Financials - Accuracy & Transparency - Financial Disclosure Quality",
        "Description": "Transparency and completeness of financial reporting",
        "Business Drivers": "Stakeholder Communication Strategy",
        "Business Drivers Description": "Commitment to transparent, comprehensive financial disclosure",
        "Performance Factors": "Transparency Excellence",
        "Performance Factors Description": "Clear, comprehensive disclosures build trust and reduce information asymmetry",
        "Risk Factors": "Opacity Risk",
        "Risk Factors Description": "Poor disclosure creates suspicion and increases cost of capital",
        "Metric": "Financial Disclosure Score",
        "Metric Description": "Assessment of disclosure quality against best practice framework",
        "Unit": "Score 1-10 or Percentage vs benchmark",
        "Target": "Score ≥8 or >80% of best practice standards",
        "Instructions": "Assess against framework: completeness, clarity, timeliness, accessibility, MD&A quality"
    },
]

# Write data to sheet
row_idx = 2  # Start after header
for item in financials_data:
    ws.cell(row=row_idx, column=1, value=item["Level"])
    ws.cell(row=row_idx, column=2, value=item["Hierarchy"])
    ws.cell(row=row_idx, column=3, value=item["Description"])
    ws.cell(row=row_idx, column=4, value=item["Business Drivers"])
    ws.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
    ws.cell(row=row_idx, column=6, value=item["Performance Factors"])
    ws.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
    ws.cell(row=row_idx, column=8, value=item["Risk Factors"])
    ws.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
    ws.cell(row=row_idx, column=10, value=item["Metric"])
    ws.cell(row=row_idx, column=11, value=item["Metric Description"])
    ws.cell(row=row_idx, column=12, value=item["Unit"])
    ws.cell(row=row_idx, column=13, value=item["Target"])
    ws.cell(row=row_idx, column=14, value=item["Instructions"])

    row_idx += 1

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save workbook
wb.save(wb_path)

print("=" * 80)
print("FINANCIALS SHEET GENERATED SUCCESSFULLY")
print("=" * 80)
print()
print(f"Location: {wb_path}")
print(f"Sheet Name: Financials")
print()
print("3-LEVEL HIERARCHY:")
print("  LEVEL 1: Financial (Dimension) - 1 item")
print("  LEVEL 2: Elements - 6 items")
print("    • Costs")
print("    • Revenue")
print("    • Profit")
print("    • Capital")
print("    • Liquidity")
print("    • Accuracy & Transparency")
print("  LEVEL 3: Sub-elements - 24 items (4 per Level 2 element)")
print()
print("TOTAL DATA POINTS: 31 rows")
print()
print("DATA STRUCTURE:")
print("  ✓ Hierarchical dimension - element - sub-element")
print("  ✓ Business Drivers (what drives this area)")
print("  ✓ Performance Factors (what good looks like)")
print("  ✓ Risk Factors (what could go wrong)")
print("  ✓ Specific Metrics with units and targets")
print("  ✓ Detailed data collection instructions")
print()
print("USE CASES:")
print("  • Base Assessment: Use aggregated Level 1 data")
print("  • Level 1 Detail: Collect data at Level 2 (6 elements)")
print("  • Level 2 Detail: Collect data at Level 3 (24 sub-elements)")
print("  • Level 3 Detail: Add supporting ratios and KPIs per sub-element")
print()
print("INTEGRATION:")
print("  • Links to Base Information sheet for org context")
print("  • Feeds multivariate Bayesian model")
print("  • Supports both risk and performance analysis")
print("  • Aligns to VOC Financials dimension")
print("=" * 80)
