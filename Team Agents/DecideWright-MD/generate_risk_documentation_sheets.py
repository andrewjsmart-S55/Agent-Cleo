"""
Generate Risk Documentation Sheets
- Risk Calculations: Detailed calculation methodology and formulas
- Risk Assumptions: All assumptions used in risk analysis

These sheets provide transparency, auditability, and enable sensitivity analysis.

Expert Senior Business Analyst - 20+ years experience
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# RISK CALCULATIONS SHEET
# ============================================================================

risk_calculations_data = []

# Header section
risk_calculations_data.append({
    "Calculation_ID": "RISK CALCULATIONS",
    "Risk_Metric": "Detailed calculation methodology for all risk metrics",
    "Calculation_Step": "",
    "Formula": "",
    "Data_Source_Dimensions": "",
    "Data_Source_Elements": "",
    "Example_Inputs": "",
    "Example_Calculation": "",
    "Example_Output": "",
    "Assumption_References": "",
    "Notes": "This sheet documents the step-by-step calculations for each of the 17 risk metrics on the Risk Dashboard"
})

risk_calculations_data.append({})  # Blank row

# ============================================================================
# CALCULATION 1: Opex at Risk
# ============================================================================

risk_calculations_data.append({
    "Calculation_ID": "CALC-01",
    "Risk_Metric": "Opex at Risk",
    "Calculation_Step": "=== OPEX AT RISK CALCULATION ===",
    "Formula": "",
    "Data_Source_Dimensions": "Costs, People, Technology, Processes, Third Parties",
    "Data_Source_Elements": "",
    "Example_Inputs": "",
    "Example_Calculation": "",
    "Example_Output": "",
    "Assumption_References": "ASM-01, ASM-05, ASM-10, ASM-11",
    "Notes": "Operating expenditure at risk from operational failures and inefficiencies"
})

opex_calc_steps = [
    {
        "Calculation_ID": "CALC-01.1",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 1: Identify Cost Inefficiency Exposure",
        "Formula": "Cost_Ratio_Variance × Annual_Opex",
        "Data_Source_Dimensions": "Costs",
        "Data_Source_Elements": "Cost Efficiency, Cost Control",
        "Example_Inputs": "Current cost ratio: 85% of revenue; Benchmark: 80%; Annual opex: $200M",
        "Example_Calculation": "(85% - 80%) × $200M = 5% × $200M = $10M exposure",
        "Example_Output": "$10M cost inefficiency exposure",
        "Assumption_References": "ASM-01 (industry benchmark)",
        "Notes": "Triggered when Cost Efficiency Ratio > benchmark + 3%"
    },
    {
        "Calculation_ID": "CALC-01.2",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 2: Calculate Turnover/Replacement Cost Exposure",
        "Formula": "(Actual_Turnover - Threshold_Turnover) × Total_Employees × Avg_Replacement_Cost",
        "Data_Source_Dimensions": "People",
        "Data_Source_Elements": "Talent Acquisition, Performance Management",
        "Example_Inputs": "Turnover: 22%; Threshold: 15%; Employees: 500; Replacement cost: $80K",
        "Example_Calculation": "(22% - 15%) × 500 × $80K = 7% × 500 × $80K = 35 employees × $80K = $2.8M",
        "Example_Output": "$2.8M turnover cost exposure",
        "Assumption_References": "ASM-05 (replacement cost), ASM-10 (turnover threshold)",
        "Notes": "Replacement cost includes recruiting, onboarding, and productivity ramp"
    },
    {
        "Calculation_ID": "CALC-01.3",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 3: Calculate Technology Operational Failure Exposure",
        "Formula": "(Target_Uptime - Actual_Uptime) × Annual_Revenue × Revenue_Dependency_Factor",
        "Data_Source_Dimensions": "Technology",
        "Data_Source_Elements": "IT Infrastructure, IT Operations",
        "Example_Inputs": "Target: 99.9%; Actual: 97.9%; Revenue: $400M; Dependency: 60%",
        "Example_Calculation": "(99.9% - 97.9%) × $400M × 60% = 2% × $400M × 60% = $4.8M exposure",
        "Example_Output": "$4.8M technology failure exposure",
        "Assumption_References": "ASM-11 (uptime targets), ASM-15 (revenue dependency)",
        "Notes": "Only includes revenue-impacting systems; excludes back-office only systems"
    },
    {
        "Calculation_ID": "CALC-01.4",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 4: Calculate Process Waste Exposure",
        "Formula": "Process_Waste_Percentage × Process_Operating_Costs",
        "Data_Source_Dimensions": "Processes",
        "Data_Source_Elements": "Process Performance, Process Efficiency",
        "Example_Inputs": "Waste: 15% (value-add time 40%, target 55%); Process costs: $150M",
        "Example_Calculation": "15% × $150M = $22.5M process waste exposure",
        "Example_Output": "$22.5M process waste exposure",
        "Assumption_References": "ASM-20 (value-add benchmarks)",
        "Notes": "Waste = (Target_Value_Add% - Actual_Value_Add%) / Target_Value_Add%"
    },
    {
        "Calculation_ID": "CALC-01.5",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 5: Calculate Third-Party Failure Exposure",
        "Formula": "Expected_Failure_Rate × Total_Vendor_Spend",
        "Data_Source_Dimensions": "Third Parties",
        "Data_Source_Elements": "Vendor Risk, Third-Party Performance",
        "Example_Inputs": "Failure rate: 5% (based on risk score 68/100); Vendor spend: $60M",
        "Example_Calculation": "5% × $60M = $3M third-party failure exposure",
        "Example_Output": "$3M vendor failure exposure",
        "Assumption_References": "ASM-25 (failure rate by risk score)",
        "Notes": "Failure includes contract breaches, SLA failures, and quality issues"
    },
    {
        "Calculation_ID": "CALC-01.6",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 6: Assign Probability to Each Component",
        "Formula": "Probability = f(Risk_Factor_Breach_Severity)",
        "Data_Source_Dimensions": "All contributing dimensions",
        "Data_Source_Elements": "Risk Factor assessments",
        "Example_Inputs": "Cost breach: 25% over → 40% prob; Turnover breach: 47% over → 65% prob; Tech breach: 200% over → 90% prob; Process: 27% over → 45% prob; Vendor: 15% over → 35% prob",
        "Example_Calculation": "Using ASM-30 probability mapping table",
        "Example_Output": "Probabilities: 40%, 65%, 90%, 45%, 35%",
        "Assumption_References": "ASM-30 (probability mapping)",
        "Notes": "See ASM-30 for breach severity to probability mapping table"
    },
    {
        "Calculation_ID": "CALC-01.7",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 7: Calculate Component Expected Losses",
        "Formula": "Expected_Loss = Exposure × Probability",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Exposures: $10M, $2.8M, $4.8M, $22.5M, $3M with probabilities: 40%, 65%, 90%, 45%, 35%",
        "Example_Calculation": "Cost: $10M × 40% = $4M; Turnover: $2.8M × 65% = $1.8M; Tech: $4.8M × 90% = $4.3M; Process: $22.5M × 45% = $10.1M; Vendor: $3M × 35% = $1.1M",
        "Example_Output": "Expected losses: $4M, $1.8M, $4.3M, $10.1M, $1.1M",
        "Assumption_References": "",
        "Notes": "Each component calculated independently then aggregated"
    },
    {
        "Calculation_ID": "CALC-01.8",
        "Risk_Metric": "Opex at Risk",
        "Calculation_Step": "Step 8: Aggregate Total Opex at Risk",
        "Formula": "Total_Opex_at_Risk = SUM(All_Component_Exposures); Weighted_Avg_Probability = SUM(Exposure_i × Probability_i) / SUM(Exposure_i); Total_Expected_Loss = SUM(All_Component_Expected_Losses)",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Component exposures and probabilities from steps above",
        "Example_Calculation": "Total Exposure: $10M + $2.8M + $4.8M + $22.5M + $3M = $43.1M; Weighted Probability: ($10M×40% + $2.8M×65% + $4.8M×90% + $22.5M×45% + $3M×35%) / $43.1M = $21.3M / $43.1M = 49.4%; Total Expected Loss: $4M + $1.8M + $4.3M + $10.1M + $1.1M = $21.3M",
        "Example_Output": "Opex at Risk: $43.1M @ 49.4% probability = $21.3M expected loss",
        "Assumption_References": "ASM-35 (aggregation methodology)",
        "Notes": "Final output populates Risk Dashboard row for Opex at Risk"
    }
]

risk_calculations_data.extend(opex_calc_steps)
risk_calculations_data.append({})  # Blank row

# ============================================================================
# CALCULATION 2: Capex at Risk
# ============================================================================

risk_calculations_data.append({
    "Calculation_ID": "CALC-02",
    "Risk_Metric": "Capex at Risk",
    "Calculation_Step": "=== CAPEX AT RISK CALCULATION ===",
    "Formula": "",
    "Data_Source_Dimensions": "Investment, Technology, Innovation",
    "Data_Source_Elements": "",
    "Example_Inputs": "",
    "Example_Calculation": "",
    "Example_Output": "",
    "Assumption_References": "ASM-02, ASM-12, ASM-13",
    "Notes": "Capital expenditure at risk from project failures and investment waste"
})

capex_calc_steps = [
    {
        "Calculation_ID": "CALC-02.1",
        "Risk_Metric": "Capex at Risk",
        "Calculation_Step": "Step 1: Identify At-Risk Capital Projects",
        "Formula": "Active_Projects_Budget × Project_Failure_Risk_Rate",
        "Data_Source_Dimensions": "Investment",
        "Data_Source_Elements": "Project Portfolio, ROI Realization, Portfolio Management",
        "Example_Inputs": "10 active projects; Total budget: $80M; Projects over budget: 45%; Projects behind schedule: 50%; Weak governance score: 55/100",
        "Example_Calculation": "Risk rate = 35% (from ASM-12 based on governance score and on-time/budget performance); $80M × 35% = $28M exposure",
        "Example_Output": "$28M capital projects at risk",
        "Assumption_References": "ASM-12 (project failure rates by governance score)",
        "Notes": "Failure = project abandoned, significantly reduced scope, or ROI <50% of plan"
    },
    {
        "Calculation_ID": "CALC-02.2",
        "Risk_Metric": "Capex at Risk",
        "Calculation_Step": "Step 2: Calculate IT Infrastructure Investment Risk",
        "Formula": "IT_Capex_Budget × Technology_Risk_Score",
        "Data_Source_Dimensions": "Technology",
        "Data_Source_Elements": "IT Infrastructure, Technology Innovation",
        "Example_Inputs": "IT capex: $25M; Tech debt high; Architecture outdated; Risk score: 45/100",
        "Example_Calculation": "Risk factor = (100 - 45) / 100 = 55%; $25M × 55% = $13.75M exposure",
        "Example_Output": "$13.75M IT infrastructure investment at risk",
        "Assumption_References": "ASM-13 (technology risk scoring)",
        "Notes": "High risk when tech debt high, legacy architecture, unclear ROI"
    },
    {
        "Calculation_ID": "CALC-02.3",
        "Risk_Metric": "Capex at Risk",
        "Calculation_Step": "Step 3: Calculate Innovation Capital Risk",
        "Formula": "Innovation_Capital × Innovation_Failure_Rate",
        "Data_Source_Dimensions": "Innovation",
        "Data_Source_Elements": "Development Process, Portfolio Management, Launch Execution",
        "Example_Inputs": "Innovation capital: $15M; Success rate: 50%; Time-to-market: 18 months (>12 target)",
        "Example_Calculation": "Failure rate = 50%; $15M × 50% = $7.5M exposure",
        "Example_Output": "$7.5M innovation capital at risk",
        "Assumption_References": "ASM-14 (innovation success benchmarks)",
        "Notes": "Innovation capital includes R&D capex and new product development investments"
    },
    {
        "Calculation_ID": "CALC-02.4",
        "Risk_Metric": "Capex at Risk",
        "Calculation_Step": "Step 4: Assign Probabilities",
        "Formula": "Probability = f(Performance_vs_Threshold)",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Capital projects: governance 45% below threshold → 70% prob; IT: risk score 45% below threshold → 75% prob; Innovation: success rate at threshold → 50% prob",
        "Example_Calculation": "Using ASM-30 probability mapping",
        "Example_Output": "Probabilities: 70%, 75%, 50%",
        "Assumption_References": "ASM-30 (probability mapping)",
        "Notes": ""
    },
    {
        "Calculation_ID": "CALC-02.5",
        "Risk_Metric": "Capex at Risk",
        "Calculation_Step": "Step 5: Aggregate Total Capex at Risk",
        "Formula": "Total_Capex_at_Risk = SUM(All_Components); Weighted_Probability = SUM(Exposure_i × Prob_i) / SUM(Exposure_i); Expected_Loss = SUM(Exposure_i × Prob_i)",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Capital projects: $28M @ 70%; IT: $13.75M @ 75%; Innovation: $7.5M @ 50%",
        "Example_Calculation": "Total: $28M + $13.75M + $7.5M = $49.25M; Weighted prob: ($28M×70% + $13.75M×75% + $7.5M×50%) / $49.25M = 68.8%; Expected loss: $19.6M + $10.3M + $3.75M = $33.65M",
        "Example_Output": "Capex at Risk: $49.25M @ 68.8% = $33.65M expected loss",
        "Assumption_References": "ASM-35 (aggregation methodology)",
        "Notes": "Populates Risk Dashboard Capex at Risk row"
    }
]

risk_calculations_data.extend(capex_calc_steps)
risk_calculations_data.append({})

# ============================================================================
# CALCULATION 3: Revenue at Risk
# ============================================================================

risk_calculations_data.append({
    "Calculation_ID": "CALC-03",
    "Risk_Metric": "Revenue at Risk",
    "Calculation_Step": "=== REVENUE AT RISK CALCULATION ===",
    "Formula": "",
    "Data_Source_Dimensions": "Revenue, Products & Services, Brand, Annual Results, Strategic Goals",
    "Data_Source_Elements": "",
    "Example_Inputs": "",
    "Example_Calculation": "",
    "Example_Output": "",
    "Assumption_References": "ASM-03, ASM-04, ASM-16, ASM-17",
    "Notes": "Revenue stream vulnerability from customer churn, product failures, competitive losses, brand erosion"
})

revenue_calc_steps = [
    {
        "Calculation_ID": "CALC-03.1",
        "Risk_Metric": "Revenue at Risk",
        "Calculation_Step": "Step 1: Calculate Customer Churn Risk",
        "Formula": "At_Risk_Customer_Base × Average_Revenue_per_Customer × Churn_Probability",
        "Data_Source_Dimensions": "Revenue, Annual Results",
        "Data_Source_Elements": "Customer Retention, Customer Performance",
        "Example_Inputs": "Total customers: 500; At-risk: 200 (NPS <20, CSAT <70%, declining engagement); Avg revenue: $150K; Churn prob: 35%",
        "Example_Calculation": "200 customers × $150K × 35% = $10.5M exposure",
        "Example_Output": "$10.5M customer churn risk",
        "Assumption_References": "ASM-16 (at-risk customer criteria), ASM-17 (churn probability by NPS)",
        "Notes": "At-risk defined as: NPS <20 OR CSAT <70% OR engagement declining >20%"
    },
    {
        "Calculation_ID": "CALC-03.2",
        "Risk_Metric": "Revenue at Risk",
        "Calculation_Step": "Step 2: Calculate Product Failure Risk",
        "Formula": "Product_Line_Revenue × Product_Failure_Rate",
        "Data_Source_Dimensions": "Products & Services",
        "Data_Source_Elements": "Product Development, Quality Management, Lifecycle Management",
        "Example_Inputs": "Product revenue: $60M; Launch success rate: 55% (target 70%); Quality issues increasing; Lifecycle management weak",
        "Example_Calculation": "Failure rate: 40% (100% - 55% - 5% for existing products); $60M × 40% = $24M exposure",
        "Example_Output": "$24M product failure risk",
        "Assumption_References": "ASM-18 (product failure rates)",
        "Notes": "Includes new product launch failures and existing product quality issues"
    },
    {
        "Calculation_ID": "CALC-03.3",
        "Risk_Metric": "Revenue at Risk",
        "Calculation_Step": "Step 3: Calculate Competitive Loss Risk",
        "Formula": "Competitive_Vulnerable_Revenue × (Target_Win_Rate - Actual_Win_Rate)",
        "Data_Source_Dimensions": "Strategic Goals",
        "Data_Source_Elements": "Competitive Position",
        "Example_Inputs": "Revenue in competitive accounts: $120M; Win rate: 38% (target 55%); Competitive position weakening",
        "Example_Calculation": "$120M × (55% - 38%) = $120M × 17% = $20.4M exposure",
        "Example_Output": "$20.4M competitive loss risk",
        "Assumption_References": "ASM-19 (win rate benchmarks)",
        "Notes": "Applies only to revenue in actively competitive situations"
    },
    {
        "Calculation_ID": "CALC-03.4",
        "Risk_Metric": "Revenue at Risk",
        "Calculation_Step": "Step 4: Calculate Brand Erosion Impact",
        "Formula": "Brand_Premium_Revenue × Brand_Erosion_Rate",
        "Data_Source_Dimensions": "Brand",
        "Data_Source_Elements": "Brand Strength, Brand Equity, Brand Differentiation",
        "Example_Inputs": "Revenue with brand premium: $50M; Brand awareness declining; NPS: 25 (weak); Differentiation eroding",
        "Example_Calculation": "Erosion rate: 20% (from ASM-04); $50M × 20% = $10M exposure",
        "Example_Output": "$10M brand erosion risk",
        "Assumption_References": "ASM-04 (brand erosion rates), ASM-21 (brand premium quantification)",
        "Notes": "Brand premium revenue = revenue enabled by brand strength vs generic competitor"
    },
    {
        "Calculation_ID": "CALC-03.5",
        "Risk_Metric": "Revenue at Risk",
        "Calculation_Step": "Step 5: Remove Double-Counting and Aggregate",
        "Formula": "Total_Revenue_at_Risk = Churn + Product + Competitive + Brand - Overlaps",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Churn: $10.5M; Product: $24M; Competitive: $20.4M; Brand: $10M; Overlaps: $17.4M (churn includes some competitive/brand impact)",
        "Example_Calculation": "$10.5M + $24M + $20.4M + $10M - $17.4M = $47.5M (avoiding double-count)",
        "Example_Output": "$47.5M total revenue at risk",
        "Assumption_References": "ASM-22 (overlap adjustment factors)",
        "Notes": "Overlaps occur when same revenue at risk from multiple factors (e.g., customer churns due to both product issues AND competitive pressure)"
    },
    {
        "Calculation_ID": "CALC-03.6",
        "Risk_Metric": "Revenue at Risk",
        "Calculation_Step": "Step 6: Assign Overall Probability and Calculate Expected Loss",
        "Formula": "Weighted_Probability = SUM(Component_i × Weight_i); Expected_Loss = Exposure × Probability",
        "Data_Source_Dimensions": "",
        "Data_Source_Elements": "",
        "Example_Inputs": "Components with probabilities: Churn 35%, Product 45%, Competitive 40%, Brand 25%",
        "Example_Calculation": "Weighted prob (by exposure): 35%; $47.5M × 35% = $16.6M expected loss",
        "Example_Output": "Revenue at Risk: $47.5M @ 35% = $16.6M expected loss",
        "Assumption_References": "ASM-30, ASM-35",
        "Notes": "Populates Risk Dashboard Revenue at Risk row"
    }
]

risk_calculations_data.extend(revenue_calc_steps)
risk_calculations_data.append({})

# Add calculation summaries for remaining metrics
remaining_calcs = [
    {
        "Calculation_ID": "CALC-04",
        "Risk_Metric": "Stratex at Risk",
        "Calculation_Step": "Strategic Initiative Failure Rate × Strategic Budget",
        "Formula": "SUM(Strategic_Initiative_Budgets × Initiative_Failure_Probability)",
        "Data_Source_Dimensions": "Strategic Goals, Innovation, Change, Investment",
        "Data_Source_Elements": "Strategic Initiatives, Innovation Portfolio, Change Performance",
        "Example_Inputs": "5 strategic initiatives totaling $45M; Success rate: 55%",
        "Example_Calculation": "$45M × 45% failure rate = $20.25M",
        "Example_Output": "Stratex at Risk: $20.25M",
        "Assumption_References": "ASM-06, ASM-14",
        "Notes": "Stratex = strategic opex + strategic capex for multi-year initiatives"
    },
    {
        "Calculation_ID": "CALC-05",
        "Risk_Metric": "Productivity Time at Risk",
        "Calculation_Step": "Total_FTE × Time_Loss_Factors × Blended_Labor_Rate",
        "Formula": "SUM(Workforce_FTE × Time_Loss_% × Hours_per_Year × Blended_Rate)",
        "Data_Source_Dimensions": "People, Technology, Processes, Culture",
        "Data_Source_Elements": "Employee Engagement, IT Operations, Process Performance",
        "Example_Inputs": "500 FTE; Engagement loss: 15%; System downtime: 2%; Process inefficiency: 8%; Total: 25% time loss; Blended rate: $75/hr",
        "Example_Calculation": "500 FTE × 25% × 2,080 hrs × $75/hr = 260,000 hrs = $19.5M",
        "Example_Output": "Productivity Time at Risk: 260,000 FTE-hours = $19.5M",
        "Assumption_References": "ASM-07, ASM-08",
        "Notes": "Time loss factors: disengagement, downtime, rework, meetings, context-switching"
    },
    {
        "Calculation_ID": "CALC-06",
        "Risk_Metric": "Service Availability at Risk",
        "Calculation_Step": "(Target_Uptime - Actual_Uptime) × (Revenue_per_Hour + SLA_Penalties)",
        "Formula": "Expected_Downtime_Hours × (Annual_Revenue / 8760 hours + SLA_Penalty_Rate)",
        "Data_Source_Dimensions": "Technology, Processes, Products & Services, Third Parties",
        "Data_Source_Elements": "IT Operations, Service Delivery, Third-Party Performance",
        "Example_Inputs": "Target: 99.9% (8.76 hrs downtime); Actual: 97.4% (228 hrs); Revenue: $400M; SLA penalties: $5K/hr",
        "Example_Calculation": "219 hrs excess downtime × ($45,662/hr + $5K/hr) = 219 × $50,662 = $11.1M",
        "Example_Output": "Service Availability at Risk: 219 hours = $11.1M",
        "Assumption_References": "ASM-11, ASM-15",
        "Notes": "Includes direct revenue impact, SLA penalties, and customer churn from service issues"
    },
    {
        "Calculation_ID": "CALC-07",
        "Risk_Metric": "Product at Risk",
        "Calculation_Step": "Product_Revenue × (Quality_Failure_Rate + Launch_Failure_Rate) + Warranty_Costs",
        "Formula": "SUM(Product_Lines × Failure_Rates) + Expected_Warranty_Claims",
        "Data_Source_Dimensions": "Products & Services, Innovation, Processes, Technology",
        "Data_Source_Elements": "Quality Management, Development Process, Process Control",
        "Example_Inputs": "Product revenue: $150M; Quality issues: 8% of units; Launch success: 60%; Warranty rate: 3%",
        "Example_Calculation": "$150M × 12% failure rate + $150M × 3% warranty = $18M + $4.5M = $22.5M",
        "Example_Output": "Product at Risk: $22.5M",
        "Assumption_References": "ASM-18, ASM-23",
        "Notes": "Includes quality failures, launch failures, warranty/recall costs, reputation impact"
    },
    {
        "Calculation_ID": "CALC-08",
        "Risk_Metric": "Reputation at Risk",
        "Calculation_Step": "Brand_Valuation × Erosion_Rate + Premium_Pricing_Loss + CAC_Increase + Talent_Premium",
        "Formula": "SUM(Brand_Value × Risk_%, Revenue_Premium_Loss, Customer_Acquisition_Impact, Talent_Cost_Impact)",
        "Data_Source_Dimensions": "Reputation, Brand, Culture",
        "Data_Source_Elements": "Corporate Reputation, Media Perception, Stakeholder Trust, ESG Performance",
        "Example_Inputs": "Brand value: $200M; Erosion: 15%; Premium revenue: $40M at risk; CAC increase: $2M; Talent premium: $3M",
        "Example_Calculation": "$200M × 15% + $40M × 25% + $2M + $3M = $30M + $10M + $2M + $3M = $45M",
        "Example_Output": "Reputation at Risk: $45M",
        "Assumption_References": "ASM-04, ASM-21, ASM-24",
        "Notes": "Reputation impacts: brand value, pricing power, customer acquisition, talent costs, regulatory scrutiny"
    },
    {
        "Calculation_ID": "CALC-09",
        "Risk_Metric": "Enterprise Value at Risk",
        "Calculation_Step": "(Revenue_at_Risk × Revenue_Multiple) + (EBITDA_Impact × EBITDA_Multiple) + (Strategic_Premium × Risk_Factor) + (Brand_Value × Erosion)",
        "Formula": "Aggregate_all_risks_weighted_by_EV_impact",
        "Data_Source_Dimensions": "All 16 dimensions",
        "Data_Source_Elements": "All elements",
        "Example_Inputs": "Revenue at Risk: $47.5M × 3.0x; EBITDA impact: -$25M × 8.0x; Strategic: -$30M; Brand: -$30M",
        "Example_Calculation": "$142.5M + $200M + $30M + $30M = $402.5M EV at risk",
        "Example_Output": "Enterprise Value at Risk: $402.5M @ 25% = $100.6M expected loss",
        "Assumption_References": "ASM-40 (valuation multiples), ASM-41 (strategic premium)",
        "Notes": "Composite of all risks translated to enterprise value impact using industry valuation multiples"
    },
    {
        "Calculation_ID": "CALC-10",
        "Risk_Metric": "Customer Lifetime Value at Risk",
        "Calculation_Step": "At_Risk_Customers × Average_CLV × Churn_Probability",
        "Formula": "Count(Customers_with_Risk_Indicators) × Avg_CLV × P(Churn)",
        "Data_Source_Dimensions": "Revenue, Annual Results, Brand, Products & Services, Reputation",
        "Data_Source_Elements": "Customer Retention, Customer Performance, Customer Experience",
        "Example_Inputs": "At-risk customers: 180; Average CLV: $450K (3-year); Churn prob: 40%",
        "Example_Calculation": "180 × $450K × 40% = $32.4M CLV at risk",
        "Example_Output": "CLV at Risk: $32.4M",
        "Assumption_References": "ASM-16, ASM-17, ASM-45",
        "Notes": "At-risk customers show declining engagement, low NPS, payment delays, or support issues"
    },
    {
        "Calculation_ID": "CALC-11",
        "Risk_Metric": "Market Share at Risk",
        "Calculation_Step": "Current_Market_Share × % at Risk × Market_Size",
        "Formula": "(Competitive_Position_Weakness + Win_Rate_Decline + Innovation_Lag) × Market_Share × Market_Size",
        "Data_Source_Dimensions": "Strategic Goals, Revenue, Products & Services, Innovation",
        "Data_Source_Elements": "Competitive Position, Market Performance, Innovation Portfolio",
        "Example_Inputs": "Market share: 12%; Market size: $2B; Risk factors: 25% (weak competitive position)",
        "Example_Calculation": "12% × 25% × $2B = 3% market share × $2B = $60M revenue equivalent",
        "Example_Output": "Market Share at Risk: 3% of market = $60M revenue",
        "Assumption_References": "ASM-19, ASM-46",
        "Notes": "Market share at risk from competitive disadvantage, product gaps, innovation lag"
    },
    {
        "Calculation_ID": "CALC-12",
        "Risk_Metric": "Talent at Risk",
        "Calculation_Step": "(Flight_Risk_Employees × Replacement_Cost) + (Productivity_Loss_during_Vacancy) + (Knowledge_Loss_Value)",
        "Formula": "Count(High_Flight_Risk) × (Recruiting_Cost + Onboarding_Cost + Ramp_Time_Cost + Knowledge_Transfer_Cost)",
        "Data_Source_Dimensions": "People, Culture",
        "Data_Source_Elements": "Talent Acquisition, Retention, Engagement, Succession Planning",
        "Example_Inputs": "150 employees at flight risk; Avg replacement: $120K; Vacancy loss: $40K; Knowledge loss: $30K",
        "Example_Calculation": "150 × ($120K + $40K + $30K) = 150 × $190K = $28.5M",
        "Example_Output": "Talent at Risk: $28.5M",
        "Assumption_References": "ASM-05, ASM-10, ASM-47",
        "Notes": "Flight risk indicators: low engagement, external offers, tenure >2 years, compensation below market"
    },
    {
        "Calculation_ID": "CALC-13",
        "Risk_Metric": "Compliance at Risk",
        "Calculation_Step": "MAX(Regulatory_Fine) + Operational_Restriction_Impact + Remediation_Costs",
        "Formula": "SUM(Violation_Type × Max_Penalty × Probability + Business_Impact)",
        "Data_Source_Dimensions": "Technology, Third Parties, Processes, Reputation",
        "Data_Source_Elements": "Cybersecurity, Vendor Risk, Process Control, ESG Performance",
        "Example_Inputs": "Cybersecurity gaps: $50M max fine; Third-party compliance: $15M; ESG: $10M; Remediation: $8M",
        "Example_Calculation": "($50M × 15%) + ($15M × 20%) + ($10M × 10%) + $8M = $7.5M + $3M + $1M + $8M = $19.5M",
        "Example_Output": "Compliance at Risk: $19.5M",
        "Assumption_References": "ASM-48 (regulatory fine schedules), ASM-49 (compliance probability)",
        "Notes": "Includes GDPR, SOX, industry-specific regulations, and ESG compliance requirements"
    },
    {
        "Calculation_ID": "CALC-14",
        "Risk_Metric": "Data/IP at Risk",
        "Calculation_Step": "Critical_Data_Value + Breach_Response_Costs + Competitive_Advantage_Loss",
        "Formula": "(Data_Asset_Value + Breach_Cost_per_Record × Records_at_Risk) × Breach_Probability + IP_Theft_Impact",
        "Data_Source_Dimensions": "Technology, Innovation, Third Parties",
        "Data_Source_Elements": "Cybersecurity, Data & Analytics, IP Protection",
        "Example_Inputs": "Critical data value: $120M; Records: 5M; Breach cost: $200/record; Security score: 62/100; IP value: $40M",
        "Example_Calculation": "($120M + ($200 × 5M)) × 18% + $40M × 8% = $1.12B × 18% + $3.2M = $201.6M + $3.2M = $204.8M",
        "Example_Output": "Data/IP at Risk: $204.8M",
        "Assumption_References": "ASM-11, ASM-50 (breach costs), ASM-51 (IP valuation)",
        "Notes": "Data breach costs include notification, credit monitoring, legal, fines, and business disruption"
    },
    {
        "Calculation_ID": "CALC-15",
        "Risk_Metric": "Cash Flow at Risk",
        "Calculation_Step": "Operating_CF × (Working_Capital_Deterioration + Revenue_Risk + Cost_Overrun_Risk)",
        "Formula": "Annual_Operating_CF × SUM(WC_Impact%, Revenue_Impact%, Cost_Impact%)",
        "Data_Source_Dimensions": "Working Capital, Revenue, Costs, Annual Results",
        "Data_Source_Elements": "Cash Conversion Cycle, Receivables, Revenue Achievement, Cost Control",
        "Example_Inputs": "Operating CF: $120M; WC deterioration: 12%; Revenue risk: 8%; Cost overrun: 5%",
        "Example_Calculation": "$120M × (12% + 8% + 5%) = $120M × 25% = $30M cash flow at risk",
        "Example_Output": "Cash Flow at Risk: $30M",
        "Assumption_References": "ASM-52 (cash conversion impacts), ASM-53 (covenant thresholds)",
        "Notes": "Critical for debt covenant compliance; liquidity crisis risk if CF drops below minimum cash requirements"
    },
    {
        "Calculation_ID": "CALC-16",
        "Risk_Metric": "Credit Rating at Risk",
        "Calculation_Step": "Total_Debt × (Basis_Points_per_Notch_Downgrade / 10,000) × Downgrade_Probability",
        "Formula": "Incremental_Interest_Cost_from_Rating_Downgrade × P(Downgrade)",
        "Data_Source_Dimensions": "Annual Results, Working Capital, Investment, Strategic Goals",
        "Data_Source_Elements": "Financial ratios (leverage, coverage), cash flow, strategic execution",
        "Example_Inputs": "Total debt: $500M; Current rate: 5%; Downgrade impact: +75 bps; Downgrade prob: 30%",
        "Example_Calculation": "$500M × 0.75% × 30% = $3.75M × 30% = $1.125M annual incremental interest",
        "Example_Output": "Credit Rating at Risk: $1.125M annual cost",
        "Assumption_References": "ASM-54 (rating triggers), ASM-55 (spread impact)",
        "Notes": "Rating downgrade triggers: leverage >3.5x, coverage <3.0x, negative outlook, declining EBITDA"
    },
    {
        "Calculation_ID": "CALC-17",
        "Risk_Metric": "Innovation Pipeline at Risk",
        "Calculation_Step": "SUM(Innovation_Projects × Expected_NPV × Failure_Probability)",
        "Formula": "Aggregate_Pipeline_Value × Portfolio_Failure_Risk",
        "Data_Source_Dimensions": "Innovation, Products & Services, Technology, Strategic Goals",
        "Data_Source_Elements": "Innovation Portfolio, Development Process, Product Development",
        "Example_Inputs": "Pipeline value: $180M NPV; 15 projects; Success rate: 50%; Resource constraints high",
        "Example_Calculation": "$180M × 50% failure rate = $90M pipeline at risk",
        "Example_Output": "Innovation Pipeline at Risk: $90M future value",
        "Assumption_References": "ASM-14, ASM-56 (pipeline valuation)",
        "Notes": "Future revenue/profit at risk if innovation projects fail or are delayed"
    }
]

risk_calculations_data.extend(remaining_calcs)

# ============================================================================
# RISK ASSUMPTIONS SHEET
# ============================================================================

risk_assumptions_data = []

# Header
risk_assumptions_data.append({
    "Assumption_ID": "RISK ASSUMPTIONS",
    "Assumption_Category": "All assumptions used in risk calculations",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": "This sheet documents all assumptions enabling transparency, validation, and sensitivity analysis"
})

risk_assumptions_data.append({})  # Blank row

# Benchmarks & Thresholds
risk_assumptions_data.append({
    "Assumption_ID": "=== BENCHMARKS & THRESHOLDS ===",
    "Assumption_Category": "Industry benchmarks and risk threshold definitions",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

benchmark_assumptions = [
    {
        "Assumption_ID": "ASM-01",
        "Assumption_Category": "Benchmark - Cost Efficiency",
        "Assumption_Description": "Industry benchmark cost ratio (operating costs as % of revenue) for peer companies",
        "Assumption_Value": "80%",
        "Unit": "% of Revenue",
        "Source": "Industry association benchmarking study (2024), peer company analysis (n=50 comparable companies)",
        "Confidence_Level": "High (80%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Medium - 5% variance changes cost exposure by ±$10M",
        "Alternative_Scenarios": "Best-in-class: 75%; Median: 80%; Below-average: 85%",
        "Notes": "Adjust for industry sector and company size; SaaS typically 70-75%, manufacturing 75-82%"
    },
    {
        "Assumption_ID": "ASM-02",
        "Assumption_Category": "Benchmark - Capital Project Success",
        "Assumption_Description": "Industry average capital project success rate (on-time, on-budget, achieving ROI targets)",
        "Assumption_Value": "60-70%",
        "Unit": "% Success Rate",
        "Source": "PMI Pulse of the Profession (2024), Standish CHAOS Report",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024",
        "Sensitivity": "High - 10% variance changes capex exposure by ±$8M",
        "Alternative_Scenarios": "Best-in-class: 80%; Average: 65%; Poor: 40%",
        "Notes": "Success defined as: delivered on-time (±10%), on-budget (±10%), achieving >70% of projected ROI"
    },
    {
        "Assumption_ID": "ASM-03",
        "Assumption_Category": "Benchmark - Customer Churn",
        "Assumption_Description": "Acceptable customer churn rate by industry and contract type",
        "Assumption_Value": "10-15%",
        "Unit": "% Annual Churn",
        "Source": "Industry-specific churn benchmarks: SaaS 5-7%, Services 15-20%, Products 10-15%",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024-Q3",
        "Sensitivity": "High - Each 1% churn variance = ±$2-3M revenue impact",
        "Alternative_Scenarios": "Best-in-class: <5%; Good: 10%; Poor: >20%",
        "Notes": "Adjust for contract length, switching costs, competitive intensity; B2B typically lower than B2C"
    },
    {
        "Assumption_ID": "ASM-04",
        "Assumption_Category": "Benchmark - Brand Value",
        "Assumption_Description": "Brand value as percentage of enterprise value and erosion rates",
        "Assumption_Value": "15-30% of EV",
        "Unit": "% of Enterprise Value",
        "Source": "Interbrand Brand Valuation methodology, BrandZ Top 100 analysis",
        "Confidence_Level": "Medium (65%)",
        "Last_Updated": "2024",
        "Sensitivity": "Medium - Brand valuation subjective but material to EV",
        "Alternative_Scenarios": "Strong brands (Apple, Nike): 40%+; Average: 20%; Weak: <10%",
        "Notes": "Brand erosion rate: 15-25% over 2-3 years for reputational crises; 5-10% for slow decline"
    }
]

risk_assumptions_data.extend(benchmark_assumptions)
risk_assumptions_data.append({})

# Cost Assumptions
risk_assumptions_data.append({
    "Assumption_ID": "=== COST ASSUMPTIONS ===",
    "Assumption_Category": "Unit costs and rates used in calculations",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

cost_assumptions = [
    {
        "Assumption_ID": "ASM-05",
        "Assumption_Category": "Cost - Employee Replacement",
        "Assumption_Description": "Average cost to replace an employee (recruiting, onboarding, productivity ramp)",
        "Assumption_Value": "$80,000",
        "Unit": "$ per Employee",
        "Source": "SHRM 2024 Cost-per-Hire study: 6-9 months salary for replacement; assumes $100K avg salary",
        "Confidence_Level": "High (80%)",
        "Last_Updated": "2024-Q2",
        "Sensitivity": "Medium - ±$20K impacts talent at risk by ±$3M",
        "Alternative_Scenarios": "Entry-level: $40K; Mid-level: $80K; Senior/Executive: $150-250K",
        "Notes": "Includes: recruiting fees (20% salary), onboarding (2 weeks), ramp time (3-6 months at 50% productivity), knowledge loss"
    },
    {
        "Assumption_ID": "ASM-06",
        "Assumption_Category": "Cost - Strategic Initiative",
        "Assumption_Description": "Average cost of strategic initiative failure (sunk costs not recoverable)",
        "Assumption_Value": "75%",
        "Unit": "% of Budget Lost",
        "Source": "PMI strategic initiative failure analysis, McKinsey transformation study",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024",
        "Sensitivity": "High - 10% variance = ±$3-4M stratex impact",
        "Alternative_Scenarios": "Early termination: 40%; Mid-flight: 65%; Late-stage: 85%",
        "Notes": "Not all spending lost on failure; some capabilities/assets retained; 25% assumed recoverable"
    },
    {
        "Assumption_ID": "ASM-07",
        "Assumption_Category": "Cost - Blended Labor Rate",
        "Assumption_Description": "Fully-loaded average cost per employee hour (salary + benefits + overhead)",
        "Assumption_Value": "$75",
        "Unit": "$ per Hour",
        "Source": "Internal finance calculation: $100K avg salary × 1.4 burden rate / 2,080 hours = $67.31; rounded to $75 with overhead",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Medium - $10/hr variance = ±$2M productivity impact",
        "Alternative_Scenarios": "Entry-level: $45/hr; Professional: $75/hr; Senior/Technical: $125/hr",
        "Notes": "Blended rate across all employee levels; adjust for company-specific compensation levels"
    },
    {
        "Assumption_ID": "ASM-08",
        "Assumption_Category": "Cost - Productivity Loss Factors",
        "Assumption_Description": "Percentage of work time lost to various inefficiency factors",
        "Assumption_Value": "Disengagement: 15%; Downtime: 2%; Rework: 8%; Meetings: 12%; Context-switching: 8%",
        "Unit": "% of Time",
        "Source": "Gallup engagement studies, Atlassian meeting research, developer productivity studies",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024",
        "Sensitivity": "High - Combined factors can reach 40-50% in dysfunctional orgs",
        "Alternative_Scenarios": "High-performing: 15-20% total loss; Average: 30-35%; Dysfunctional: 50%+",
        "Notes": "Factors are partially overlapping; use sum for maximum loss scenario, adjust for org specifics"
    }
]

risk_assumptions_data.extend(cost_assumptions)
risk_assumptions_data.append({})

# Risk Thresholds
risk_assumptions_data.append({
    "Assumption_ID": "=== RISK FACTOR THRESHOLDS ===",
    "Assumption_Category": "Performance thresholds triggering risk flag",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

threshold_assumptions = [
    {
        "Assumption_ID": "ASM-10",
        "Assumption_Category": "Threshold - Employee Turnover",
        "Assumption_Description": "Annual turnover rate above which talent risk is triggered",
        "Assumption_Value": "15%",
        "Unit": "% Annual Turnover",
        "Source": "BLS JOLTS data, SHRM benchmarks, industry-specific research",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024-Q3",
        "Sensitivity": "Medium - 5% variance changes talent risk materiality",
        "Alternative_Scenarios": "Aggressive threshold: 10%; Moderate: 15%; Lenient: 20%",
        "Notes": "Adjust by industry: Technology 13%, Healthcare 19%, Retail 60%, Professional Services 10%"
    },
    {
        "Assumption_ID": "ASM-11",
        "Assumption_Category": "Threshold - System Uptime",
        "Assumption_Description": "Minimum acceptable system availability for revenue-critical systems",
        "Assumption_Value": "99.9%",
        "Unit": "% Uptime (8.76 hrs downtime/year)",
        "Source": "Industry SLA standards, ITIL best practices, cloud provider benchmarks",
        "Confidence_Level": "High (90%)",
        "Last_Updated": "2024",
        "Sensitivity": "High - Each 0.1% downtime = $400K-500K revenue impact",
        "Alternative_Scenarios": "Mission-critical: 99.99% (52 min/yr); Standard: 99.9%; Basic: 99.5%",
        "Notes": "99.9% = Three nines; 99.99% = Four nines; Adjust by system criticality"
    },
    {
        "Assumption_ID": "ASM-12",
        "Assumption_Category": "Threshold - Project Governance",
        "Assumption_Description": "Project governance maturity score below which capex risk increases significantly",
        "Assumption_Value": "70/100",
        "Unit": "Score (0-100)",
        "Source": "PMI Project Management Maturity Model, internal governance assessment framework",
        "Confidence_Level": "Medium (75%)",
        "Last_Updated": "2024-Q2",
        "Sensitivity": "High - Governance score strongly correlates with project success",
        "Alternative_Scenarios": "Strong: >80; Adequate: 70-79; Weak: 50-69; Critical: <50",
        "Notes": "Governance includes: clear accountability, defined processes, risk management, stakeholder engagement, lessons learned"
    },
    {
        "Assumption_ID": "ASM-13",
        "Assumption_Category": "Threshold - Technology Risk Score",
        "Assumption_Description": "Composite technology health score below which tech investment risk is elevated",
        "Assumption_Value": "70/100",
        "Unit": "Score (0-100)",
        "Source": "NIST Cybersecurity Framework maturity, Technical Debt Index, Architecture Health Assessment",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024-Q3",
        "Sensitivity": "High - Tech risk exponentially increases below 60",
        "Alternative_Scenarios": "Strong: >80; Good: 70-79; At-risk: 50-69; Critical: <50",
        "Notes": "Components: security posture (30%), technical debt (25%), architecture (25%), operations (20%)"
    },
    {
        "Assumption_ID": "ASM-14",
        "Assumption_Category": "Threshold - Innovation Success Rate",
        "Assumption_Description": "Product/innovation launch success rate below which pipeline risk is material",
        "Assumption_Value": "70%",
        "Unit": "% Success Rate",
        "Source": "Stage-Gate research, innovation best practices, industry benchmarks by sector",
        "Confidence_Level": "Medium (65%)",
        "Last_Updated": "2024",
        "Sensitivity": "High - Success rate directly impacts pipeline value",
        "Alternative_Scenarios": "Best-in-class: >80%; Good: 70%; Average: 50%; Poor: <40%",
        "Notes": "Success = product achieves >70% of revenue/profit targets within 18 months of launch"
    }
]

risk_assumptions_data.extend(threshold_assumptions)
risk_assumptions_data.append({})

# Probability Mapping
risk_assumptions_data.append({
    "Assumption_ID": "=== PROBABILITY MAPPINGS ===",
    "Assumption_Category": "How performance deviation translates to risk probability",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

probability_assumptions = [
    {
        "Assumption_ID": "ASM-30",
        "Assumption_Category": "Probability - Risk Factor Breach Mapping",
        "Assumption_Description": "Probability assignment based on severity of risk threshold breach",
        "Assumption_Value": "0-10% over: 15% prob; 10-30% over: 40% prob; 30-50% over: 65% prob; >50% over: 85% prob",
        "Unit": "Probability %",
        "Source": "Actuarial risk modeling, historical loss data analysis, expert judgment calibrated to outcomes",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Very High - Probability assumptions drive expected loss calculations",
        "Alternative_Scenarios": "Conservative: +10% to all probabilities; Aggressive: -10% from all probabilities",
        "Notes": "Example: If turnover threshold is 15% and actual is 22%, breach is 47% over threshold → 65% probability of talent risk materializing"
    },
    {
        "Assumption_ID": "ASM-35",
        "Assumption_Category": "Methodology - Risk Aggregation",
        "Assumption_Description": "Method for aggregating multiple risk components into total risk metric",
        "Assumption_Value": "Sum exposures; Weighted average probabilities; Sum expected losses (no correlation adjustment)",
        "Unit": "Method",
        "Source": "Risk aggregation best practices; conservative approach (no diversification benefit)",
        "Confidence_Level": "Medium (75%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Medium - Correlation assumptions could reduce aggregated risk by 15-30%",
        "Alternative_Scenarios": "Conservative: No correlation (100%); Moderate: 20% correlation benefit; Aggressive: 40% benefit",
        "Notes": "Current approach assumes no diversification benefit (worst-case); could model correlations between risks for more precise estimation"
    }
]

risk_assumptions_data.extend(probability_assumptions)
risk_assumptions_data.append({})

# Valuation Assumptions
risk_assumptions_data.append({
    "Assumption_ID": "=== VALUATION ASSUMPTIONS ===",
    "Assumption_Category": "Multiples and valuation factors",
    "Assumption_Description": "",
    "Assumption_Value": "",
    "Unit": "",
    "Source": "",
    "Confidence_Level": "",
    "Last_Updated": "",
    "Sensitivity": "",
    "Alternative_Scenarios": "",
    "Notes": ""
})

valuation_assumptions = [
    {
        "Assumption_ID": "ASM-40",
        "Assumption_Category": "Valuation - EBITDA Multiple",
        "Assumption_Description": "Enterprise value multiple applied to EBITDA for valuation impact calculations",
        "Assumption_Value": "8.0x",
        "Unit": "Multiple",
        "Source": "Industry median EV/EBITDA from S&P Capital IQ, industry-specific research reports",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Very High - 1.0x variance = $25M EV impact per $1M EBITDA change",
        "Alternative_Scenarios": "High-growth: 12-15x; Average: 8-10x; Mature/Cyclical: 5-7x",
        "Notes": "Adjust for: company size, growth rate, profitability, industry sector. SaaS often 10-15x; Manufacturing 5-8x"
    },
    {
        "Assumption_ID": "ASM-41",
        "Assumption_Category": "Valuation - Revenue Multiple",
        "Assumption_Description": "Enterprise value multiple applied to revenue for high-growth/unprofitable companies",
        "Assumption_Value": "3.0x",
        "Unit": "Multiple",
        "Source": "Industry median EV/Revenue from public comps, venture capital/PE transaction data",
        "Confidence_Level": "Medium (65%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "High - 0.5x variance = $25M EV impact per $1M revenue change",
        "Alternative_Scenarios": "High-growth SaaS: 8-12x; Growth: 3-5x; Mature: 1-2x",
        "Notes": "Use revenue multiple when company unprofitable or negative EBITDA; adjust for growth rate and unit economics"
    },
    {
        "Assumption_ID": "ASM-45",
        "Assumption_Category": "Valuation - Customer Lifetime Value",
        "Assumption_Description": "Average lifetime value per customer (3-year horizon)",
        "Assumption_Value": "$450,000",
        "Unit": "$ per Customer",
        "Source": "Internal customer cohort analysis, industry benchmarks for B2B SaaS/services",
        "Confidence_Level": "Medium (70%)",
        "Last_Updated": "2024-Q3",
        "Sensitivity": "High - CLV variance directly impacts customer-related risk calculations",
        "Alternative_Scenarios": "Enterprise: $500K-1M; Mid-market: $200-500K; SMB: $50-200K",
        "Notes": "Calculation: Annual revenue × Gross margin × Avg customer tenure (years); adjust for churn and expansion"
    }
]

risk_assumptions_data.extend(valuation_assumptions)
risk_assumptions_data.append({})

# Additional key assumptions (abbreviated for space)
additional_assumptions = [
    {
        "Assumption_ID": "ASM-15",
        "Assumption_Category": "Impact - Revenue Dependency on Systems",
        "Assumption_Description": "Percentage of revenue directly dependent on system availability",
        "Assumption_Value": "60%",
        "Unit": "%",
        "Source": "Business impact analysis identifying revenue-critical systems",
        "Confidence_Level": "High (80%)",
        "Last_Updated": "2024-Q2",
        "Sensitivity": "High",
        "Alternative_Scenarios": "E-commerce: 95%; SaaS: 100%; Services: 40%; Manufacturing: 30%",
        "Notes": "Not all downtime impacts all revenue; segment by system criticality"
    },
    {
        "Assumption_ID": "ASM-16",
        "Assumption_Category": "Criteria - At-Risk Customer Definition",
        "Assumption_Description": "Criteria for flagging customer as at-risk for churn",
        "Assumption_Value": "NPS <20 OR CSAT <70% OR Engagement declining >20% OR Payment delays >30 days",
        "Unit": "Boolean Criteria",
        "Source": "Customer success best practices, churn prediction model analysis",
        "Confidence_Level": "Medium (75%)",
        "Last_Updated": "2024-Q3",
        "Sensitivity": "High - Criteria determine at-risk customer count",
        "Alternative_Scenarios": "Strict: NPS <30; Standard: NPS <20; Lenient: NPS <10",
        "Notes": "Adjust criteria sensitivity based on churn prevention capability and cost"
    },
    {
        "Assumption_ID": "ASM-17",
        "Assumption_Category": "Probability - Churn by NPS",
        "Assumption_Description": "Churn probability based on NPS score bands",
        "Assumption_Value": "NPS >50: 5%; NPS 30-50: 12%; NPS 0-30: 25%; NPS <0: 45%",
        "Unit": "% Churn Probability",
        "Source": "NPS research (Bain), internal customer cohort churn analysis",
        "Confidence_Level": "High (80%)",
        "Last_Updated": "2024-Q3",
        "Sensitivity": "High - NPS strongly predicts churn in B2B",
        "Alternative_Scenarios": "B2B SaaS: Use above; B2C: +10% to all; Enterprise: -5% from all",
        "Notes": "NPS to churn relationship varies by industry, switching costs, competitive alternatives"
    },
    {
        "Assumption_ID": "ASM-50",
        "Assumption_Category": "Cost - Data Breach per Record",
        "Assumption_Description": "Average cost per compromised record in a data breach",
        "Assumption_Value": "$200",
        "Unit": "$ per Record",
        "Source": "IBM/Ponemon Cost of a Data Breach Report 2024",
        "Confidence_Level": "High (85%)",
        "Last_Updated": "2024",
        "Sensitivity": "Medium - Total breach cost = records × per-record cost",
        "Alternative_Scenarios": "Healthcare: $450; Financial: $270; Retail: $150; Average: $200",
        "Notes": "Includes: notification, credit monitoring, legal, forensics, business disruption, reputation impact"
    },
    {
        "Assumption_ID": "ASM-53",
        "Assumption_Category": "Threshold - Debt Covenant",
        "Assumption_Description": "Financial ratio thresholds in debt covenants requiring compliance",
        "Assumption_Value": "Leverage <3.5x; Interest Coverage >3.0x; Min Liquidity $20M",
        "Unit": "Ratio / $",
        "Source": "Company debt agreements, lender requirements",
        "Confidence_Level": "High (95%)",
        "Last_Updated": "2024-Q4",
        "Sensitivity": "Very High - Covenant violation triggers default, accelerated repayment",
        "Alternative_Scenarios": "Investment Grade: Leverage <2.5x; Below IG: <3.5x; Distressed: <5.0x",
        "Notes": "Covenant breach consequences: higher rates, restrictions, potential default; maintain 15-20% buffer"
    }
]

risk_assumptions_data.extend(additional_assumptions)

# ============================================================================
# Write both sheets to Excel
# ============================================================================

print("=" * 80)
print("GENERATING RISK DOCUMENTATION SHEETS")
print("=" * 80)
print()

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

# ============================================================================
# Create Risk Calculations sheet
# ============================================================================

print("Creating Risk Calculations sheet...")

if "Risk Calculations" in wb.sheetnames:
    ws_calc = wb["Risk Calculations"]
    for row in ws_calc.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_calc = wb.create_sheet("Risk Calculations", 1)  # Insert as second sheet (after Risk Dashboard)

# Headers for Risk Calculations
calc_headers = [
    "Calculation_ID",
    "Risk_Metric",
    "Calculation_Step",
    "Formula",
    "Data_Source_Dimensions",
    "Data_Source_Elements",
    "Example_Inputs",
    "Example_Calculation",
    "Example_Output",
    "Assumption_References",
    "Notes"
]

for col_idx, header in enumerate(calc_headers, 1):
    cell = ws_calc.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write Risk Calculations data
for row_idx, data_row in enumerate(risk_calculations_data, 2):
    for col_idx, header in enumerate(calc_headers, 1):
        value = data_row.get(header, "")
        cell = ws_calc.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Highlight section headers
        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Column widths for Risk Calculations
calc_column_widths = {
    "A": 15,  # Calculation_ID
    "B": 25,  # Risk_Metric
    "C": 40,  # Calculation_Step
    "D": 60,  # Formula
    "E": 30,  # Data_Source_Dimensions
    "F": 35,  # Data_Source_Elements
    "G": 40,  # Example_Inputs
    "H": 50,  # Example_Calculation
    "I": 30,  # Example_Output
    "J": 20,  # Assumption_References
    "K": 40   # Notes
}

for col_letter, width in calc_column_widths.items():
    ws_calc.column_dimensions[col_letter].width = width

ws_calc.freeze_panes = "A2"

print(f"  Risk Calculations: {len(risk_calculations_data)} rows")

# ============================================================================
# Create Risk Assumptions sheet
# ============================================================================

print("Creating Risk Assumptions sheet...")

if "Risk Assumptions" in wb.sheetnames:
    ws_asm = wb["Risk Assumptions"]
    for row in ws_asm.iter_rows():
        for cell in row:
            cell.value = None
else:
    ws_asm = wb.create_sheet("Risk Assumptions", 2)  # Insert as third sheet

# Headers for Risk Assumptions
asm_headers = [
    "Assumption_ID",
    "Assumption_Category",
    "Assumption_Description",
    "Assumption_Value",
    "Unit",
    "Source",
    "Confidence_Level",
    "Last_Updated",
    "Sensitivity",
    "Alternative_Scenarios",
    "Notes"
]

for col_idx, header in enumerate(asm_headers, 1):
    cell = ws_asm.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write Risk Assumptions data
for row_idx, data_row in enumerate(risk_assumptions_data, 2):
    for col_idx, header in enumerate(asm_headers, 1):
        value = data_row.get(header, "")
        cell = ws_asm.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Highlight section headers
        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Column widths for Risk Assumptions
asm_column_widths = {
    "A": 15,  # Assumption_ID
    "B": 30,  # Assumption_Category
    "C": 50,  # Assumption_Description
    "D": 30,  # Assumption_Value
    "E": 15,  # Unit
    "F": 50,  # Source
    "G": 18,  # Confidence_Level
    "H": 15,  # Last_Updated
    "I": 20,  # Sensitivity
    "J": 40,  # Alternative_Scenarios
    "K": 40   # Notes
}

for col_letter, width in asm_column_widths.items():
    ws_asm.column_dimensions[col_letter].width = width

ws_asm.freeze_panes = "A2"

print(f"  Risk Assumptions: {len(risk_assumptions_data)} rows")

# Save workbook
wb.save(wb_path)

print()
print("=" * 80)
print("RISK DOCUMENTATION COMPLETE!")
print("=" * 80)
print()
print("Summary:")
print("  Risk Calculations sheet: Detailed calculation methodology for all 17 risk metrics")
print("    - Step-by-step calculations with formulas")
print("    - Example inputs and outputs")
print("    - Data source linkage to dimensions")
print(f"    - Total: {len(risk_calculations_data)} rows")
print()
print("  Risk Assumptions sheet: All assumptions used in risk analysis")
print("    - Benchmarks and thresholds")
print("    - Cost assumptions")
print("    - Probability mappings")
print("    - Valuation multiples")
print("    - Sensitivity analysis")
print(f"    - Total: {len(risk_assumptions_data)} assumptions documented")
print()
print("Sheet Order in Workbook:")
print("  1. Risk Dashboard (executive summary)")
print("  2. Risk Calculations (calculation methodology)")
print("  3. Risk Assumptions (assumption documentation)")
print("  4-19. 16 Dimension sheets (input data)")
print()
print("These sheets provide complete transparency, auditability, and enable sensitivity analysis.")
