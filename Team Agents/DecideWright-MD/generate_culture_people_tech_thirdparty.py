"""
Generate Remaining 4 Enablers Domain Sheets
Culture, People, Technology, Third Parties
Expert Senior Business Analyst - 20+ years
"""

import openpyxl
from openpyxl import load_workbook

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("=" * 80)
print("GENERATING CULTURE, PEOPLE, TECHNOLOGY, THIRD PARTIES SHEETS")
print("=" * 80)
print()

# CULTURE DATA - This file is very long, so I'm including all data structures inline
# Due to token limits, I'll generate a focused script that directly populates each sheet

# Helper function to generate sheet data
def generate_dimension_data(dimension_name, elements_data):
    """Generate complete 31-row structure for a dimension"""
    data = []

    # Level 1 - Dimension
    data.append(elements_data["dimension"])

    # Level 2 - Elements (6 elements)
    for element in elements_data["elements"]:
        data.append(element)

    # Level 3 - Sub-elements (24 sub-elements, 4 per element)
    for sub_elements in elements_data["sub_elements"]:
        for sub in sub_elements:
            data.append(sub)

    return data

# CULTURE SHEET DATA
print("1. CULTURE SHEET")
print("-" * 80)

ws_culture = wb["Culture"]
for row in ws_culture.iter_rows(min_row=2, max_row=ws_culture.max_row):
    for cell in row:
        cell.value = None

culture_data = [
    # LEVEL 1
    {"Level": 1, "Hierarchy": "Culture", "Description": "Shared values, beliefs, and behaviors shaping how work gets done",
     "Business Drivers": "Values & Norms", "Business Drivers Description": "Culture determines how decisions are made, how people collaborate, and how the organization adapts",
     "Performance Factors": "Positive Culture", "Performance Factors Description": "Culture aligned to strategy, high engagement, psychological safety, continuous improvement",
     "Risk Factors": "Toxic Culture", "Risk Factors Description": "Culture misaligned to strategy, low engagement, fear-based, resistant to change",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of all culture elements"},

    # LEVEL 2 - Values Alignment
    {"Level": 2, "Hierarchy": "Culture - Values Alignment", "Description": "Degree to which stated values are lived daily",
     "Business Drivers": "Values Authenticity", "Business Drivers Description": "Values guide behavior when lived authentically, not just posted on walls",
     "Performance Factors": "Values-Driven Behavior", "Performance Factors Description": "Employees model values in decisions and actions",
     "Risk Factors": "Values-Reality Gap", "Risk Factors Description": "Stated values disconnected from actual behavior",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of values alignment sub-elements"},

    # LEVEL 3 - Values Alignment Sub-elements
    {"Level": 3, "Hierarchy": "Culture - Values Alignment - Values Clarity", "Description": "Clear definition and communication of core values",
     "Business Drivers": "Values Definition", "Business Drivers Description": "Well-defined values provide behavioral compass",
     "Performance Factors": "Values Understanding", "Performance Factors Description": "Employees can articulate core values and their meaning",
     "Risk Factors": "Values Confusion", "Risk Factors Description": "Employees don't know or understand stated values",
     "Metric": "Values Awareness %", "Metric Description": "% employees who can name all core values without prompting",
     "Unit": "Percentage", "Target": ">80% can name all values", "Instructions": "Survey: Name our core values (no prompts). Calculate % correct."},

    {"Level": 3, "Hierarchy": "Culture - Values Alignment - Values-Behavior Consistency", "Description": "Observable behaviors match stated values",
     "Business Drivers": "Behavioral Integrity", "Business Drivers Description": "Actions speak louder than words - values must be demonstrated",
     "Performance Factors": "Values Lived", "Performance Factors Description": "Employees observe values being practiced consistently",
     "Risk Factors": "Hypocrisy", "Risk Factors Description": "Behaviors contradict stated values, breeding cynicism",
     "Metric": "Values-Behavior Alignment Score", "Metric Description": "Employee rating: Leadership and peers demonstrate values (1-10 scale)",
     "Unit": "Score 1-10", "Target": ">7/10 = values lived", "Instructions": "Survey: Rate how well leaders/peers demonstrate each value. Average across values."},

    {"Level": 3, "Hierarchy": "Culture - Values Alignment - Values-Decision Integration", "Description": "Values explicitly considered in major decisions",
     "Business Drivers": "Values-Based Decision Making", "Business Drivers Description": "Values serve as decision criteria, not just aspirations",
     "Performance Factors": "Values Guide Choices", "Performance Factors Description": "Major decisions evaluated against values framework",
     "Risk Factors": "Values Ignored", "Risk Factors Description": "Values irrelevant to decision-making, purely financial/operational focus",
     "Metric": "Values in Decisions", "Metric Description": "% major decisions with documented values consideration",
     "Unit": "Percentage", "Target": ">70% of strategic/major decisions", "Instructions": "Review decision documentation: Were values explicitly considered? Calculate %."},

    {"Level": 3, "Hierarchy": "Culture - Values Alignment - Values Recognition & Reinforcement", "Description": "Values-aligned behavior is recognized and rewarded",
     "Business Drivers": "Values Incentivization", "Business Drivers Description": "What gets rewarded gets repeated - recognition shapes culture",
     "Performance Factors": "Values Reinforcement", "Performance Factors Description": "Clear rewards for values demonstration, consequences for violations",
     "Risk Factors": "Values Irrelevance", "Risk Factors Description": "No connection between values and recognition/advancement",
     "Metric": "Values-Based Recognition %", "Metric Description": "% recognition/rewards explicitly tied to values demonstration",
     "Unit": "Percentage", "Target": ">50% of recognition references values", "Instructions": "Audit recognition programs: What % explicitly cite values? Track quarterly."},

    # LEVEL 2 - Employee Engagement
    {"Level": 2, "Hierarchy": "Culture - Employee Engagement", "Description": "Commitment, discretionary effort, and emotional connection to organization",
     "Business Drivers": "Work Environment Quality", "Business Drivers Description": "Engaging work, supportive management, growth opportunities drive engagement",
     "Performance Factors": "High Engagement", "Performance Factors Description": "Employees committed, enthusiastic, willing to go extra mile",
     "Risk Factors": "Disengagement", "Risk Factors Description": "Employees doing minimum, looking to leave, or actively undermining",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of employee engagement sub-elements"},

    # LEVEL 3 - Engagement Sub-elements
    {"Level": 3, "Hierarchy": "Culture - Employee Engagement - Overall Engagement Score", "Description": "Composite measure of employee engagement",
     "Business Drivers": "Multiple Engagement Drivers", "Business Drivers Description": "Engagement influenced by purpose, autonomy, mastery, relationships, rewards",
     "Performance Factors": "Highly Engaged Workforce", "Performance Factors Description": "Majority of employees engaged or highly engaged",
     "Risk Factors": "Disengaged Workforce", "Risk Factors Description": "Significant proportion actively disengaged or checked out",
     "Metric": "Engagement Index Score", "Metric Description": "Composite score from annual survey across multiple dimensions",
     "Unit": "Score (typically 0-100 or 1-5 scale)", "Target": ">70/100 or >4.0/5.0 = high engagement",
     "Instructions": "Annual engagement survey (Gallup Q12, custom index). Benchmark vs industry. Track trends."},

    {"Level": 3, "Hierarchy": "Culture - Employee Engagement - Discretionary Effort", "Description": "Willingness to go beyond job requirements",
     "Business Drivers": "Commitment & Pride", "Business Drivers Description": "Engaged employees willingly contribute extra effort",
     "Performance Factors": "Extra Mile Behavior", "Performance Factors Description": "Employees regularly go above and beyond",
     "Risk Factors": "Minimum Effort", "Risk Factors Description": "Employees do only what's required, no more",
     "Metric": "Discretionary Effort Score", "Metric Description": "Survey: How often do you go beyond job requirements? (1-5 scale)",
     "Unit": "Score 1-5 average", "Target": ">4.0/5.0 = strong discretionary effort",
     "Instructions": "Employee survey: Frequency of: staying late, helping colleagues, suggesting improvements, etc."},

    {"Level": 3, "Hierarchy": "Culture - Employee Engagement - Intent to Stay", "Description": "Employee plans to remain with organization",
     "Business Drivers": "Job Satisfaction", "Business Drivers Description": "Satisfied, engaged employees want to stay",
     "Performance Factors": "Strong Retention Intent", "Performance Factors Description": "Most employees plan to stay 2+ years",
     "Risk Factors": "Flight Risk", "Risk Factors Description": "Many employees actively job searching or planning to leave",
     "Metric": "Intent to Stay %", "Metric Description": "% likely/very likely to be here in 12 months",
     "Unit": "Percentage", "Target": ">80% plan to stay 12+ months",
     "Instructions": "Survey: How likely are you to still be here in 12 months? (1-5, % answering 4-5)"},

    {"Level": 3, "Hierarchy": "Culture - Employee Engagement - Pride in Organization", "Description": "Employee pride in being part of the organization",
     "Business Drivers": "Purpose & Reputation", "Business Drivers Description": "Employees proud of mission, reputation, and work",
     "Performance Factors": "Strong Pride", "Performance Factors Description": "Employees enthusiastically identify with organization",
     "Risk Factors": "Embarrassment", "Risk Factors Description": "Employees reluctant to say where they work",
     "Metric": "Organizational Pride Score", "Metric Description": "Survey: Proud to work here, Recommend to friends (1-5 scale average)",
     "Unit": "Score 1-5", "Target": ">4.0/5.0 = strong pride",
     "Instructions": "Survey: Rate pride in organization, likelihood to recommend as employer (1-5). Average."},

    # LEVEL 2 - Psychological Safety
    {"Level": 2, "Hierarchy": "Culture - Psychological Safety", "Description": "Belief that one can speak up without negative consequences",
     "Business Drivers": "Trust & Openness", "Business Drivers Description": "Safety to take risks and admit mistakes enables learning and innovation",
     "Performance Factors": "High Psychological Safety", "Performance Factors Description": "People freely share ideas, concerns, mistakes without fear",
     "Risk Factors": "Fear Culture", "Risk Factors Description": "People afraid to speak up, hide mistakes, avoid risks",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of psychological safety sub-elements"},

    # LEVEL 3 - Psychological Safety Sub-elements
    {"Level": 3, "Hierarchy": "Culture - Psychological Safety - Speak-Up Culture", "Description": "Comfort level sharing ideas, concerns, bad news",
     "Business Drivers": "Leader Receptivity", "Business Drivers Description": "Leaders who welcome dissent and bad news create safety",
     "Performance Factors": "Open Communication", "Performance Factors Description": "People freely voice opinions, even if contrary",
     "Risk Factors": "Silence & Suppression", "Risk Factors Description": "People self-censor, withhold concerns, tell leaders what they want to hear",
     "Metric": "Speak-Up Safety Score", "Metric Description": "Survey: Feel safe voicing opinions, raising concerns (1-5 scale)",
     "Unit": "Score 1-5", "Target": ">4.0/5.0 = high psychological safety",
     "Instructions": "Survey: I feel safe speaking up with ideas/concerns/opposing views. Rate 1-5 agreement."},

    {"Level": 3, "Hierarchy": "Culture - Psychological Safety - Failure Tolerance", "Description": "Organization's response to mistakes and failures",
     "Business Drivers": "Learning Orientation", "Business Drivers Description": "Treating failures as learning opportunities vs punishment",
     "Performance Factors": "Productive Failure Response", "Performance Factors Description": "Mistakes analyzed for learning, not blame",
     "Risk Factors": "Blame Culture", "Risk Factors Description": "Failures punished, leading to cover-ups and risk aversion",
     "Metric": "Failure Response Score", "Metric Description": "Survey: OK to make mistakes, Failures used for learning (1-5 scale average)",
     "Unit": "Score 1-5", "Target": ">3.5/5.0 = learning culture", "Instructions": "Survey: Rate agreement with: OK to make mistakes, failures lead to learning not blame."},

    {"Level": 3, "Hierarchy": "Culture - Psychological Safety - Interpersonal Trust", "Description": "Trust among team members and with leadership",
     "Business Drivers": "Relationship Quality", "Business Drivers Description": "Trust built through consistency, competence, caring",
     "Performance Factors": "High Trust", "Performance Factors Description": "People trust colleagues and leaders to have their back",
     "Risk Factors": "Distrust", "Risk Factors Description": "Suspicion, CYA behavior, unwillingness to be vulnerable",
     "Metric": "Trust in Leadership & Peers", "Metric Description": "Survey: Trust in direct leader, Trust in senior leadership, Trust in peers (1-5 each)",
     "Unit": "Score 1-5 (average across 3)", "Target": ">4.0/5.0 = high trust", "Instructions": "Survey: Rate trust in: Your manager, Senior leaders, Peers. Average across three."},

    {"Level": 3, "Hierarchy": "Culture - Psychological Safety - Risk-Taking Encouragement", "Description": "Organizational support for intelligent risk-taking",
     "Business Drivers": "Innovation Mandate", "Business Drivers Description": "Growth requires trying new things, which involves risk",
     "Performance Factors": "Smart Risks Supported", "Performance Factors Description": "Well-reasoned risks encouraged, not penalized if fail",
     "Risk Factors": "Risk Aversion", "Risk Factors Description": "Only sure things attempted, stagnation, fear of innovation",
     "Metric": "Risk-Taking Climate Score", "Metric Description": "Survey: Encouraged to try new approaches, Failure doesn't hurt career (1-5 average)",
     "Unit": "Score 1-5", "Target": ">3.5/5.0 = risk-tolerant culture", "Instructions": "Survey: Encouraged to experiment? Intelligent failures acceptable? Average scores."},

    # LEVEL 2 - Collaboration
    {"Level": 2, "Hierarchy": "Culture - Collaboration", "Description": "Quality and extent of cross-functional teamwork",
     "Business Drivers": "Interdependence", "Business Drivers Description": "Complex work requires effective collaboration across boundaries",
     "Performance Factors": "High Collaboration", "Performance Factors Description": "Teams work well together, share knowledge, support each other",
     "Risk Factors": "Silos & Turf Wars", "Risk Factors Description": "Functions compete, hoard information, undermine each other",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of collaboration sub-elements"},

    # LEVEL 3 - Collaboration Sub-elements
    {"Level": 3, "Hierarchy": "Culture - Collaboration - Cross-Functional Cooperation", "Description": "Quality of teamwork across organizational boundaries",
     "Business Drivers": "Shared Goals", "Business Drivers Description": "Common objectives align functions toward collaboration",
     "Performance Factors": "Seamless Cooperation", "Performance Factors Description": "Functions work together smoothly toward shared outcomes",
     "Risk Factors": "Siloed Behavior", "Risk Factors Description": "Functions optimize locally, don't collaborate",
     "Metric": "Cross-Functional Collaboration Score", "Metric Description": "Survey: Other teams cooperative, Easy to work cross-functionally (1-5 average)",
     "Unit": "Score 1-5", "Target": ">3.5/5.0 = good collaboration", "Instructions": "Survey: Rate cooperation from other teams, ease of cross-functional work. Average."},

    {"Level": 3, "Hierarchy": "Culture - Collaboration - Information Sharing", "Description": "Transparency and knowledge sharing across organization",
     "Business Drivers": "Knowledge Management", "Business Drivers Description": "Shared knowledge improves decisions and prevents duplication",
     "Performance Factors": "Open Information Flow", "Performance Factors Description": "Information readily shared, not hoarded",
     "Risk Factors": "Information Hoarding", "Risk Factors Description": "Knowledge is power mentality, information silos",
     "Metric": "Information Access Score", "Metric Description": "Survey: Can access info needed, People share knowledge (1-5 average)",
     "Unit": "Score 1-5", "Target": ">4.0/5.0 = transparent culture", "Instructions": "Survey: Rate ease of accessing info, willingness of others to share knowledge."},

    {"Level": 3, "Hierarchy": "Culture - Collaboration - Meeting Effectiveness", "Description": "Productivity and quality of collaborative meetings",
     "Business Drivers": "Time Management", "Business Drivers Description": "Meetings represent significant time investment requiring effectiveness",
     "Performance Factors": "Productive Meetings", "Performance Factors Description": "Meetings achieve objectives, right people, clear outcomes",
     "Risk Factors": "Meeting Waste", "Risk Factors Description": "Excessive, unproductive meetings draining time and morale",
     "Metric": "Meeting Effectiveness Score", "Metric Description": "Survey: Meetings productive, Right people attend, Clear outcomes (1-5 average)",
     "Unit": "Score 1-5", "Target": ">3.5/5.0 = effective meetings", "Instructions": "Survey: Rate meeting productivity, attendance relevance, outcome clarity. Average."},

    {"Level": 3, "Hierarchy": "Culture - Collaboration - Conflict Resolution", "Description": "Constructive handling of disagreements and tensions",
     "Business Drivers": "Conflict Management Skills", "Business Drivers Description": "Healthy organizations surface and resolve conflicts productively",
     "Performance Factors": "Constructive Conflict", "Performance Factors Description": "Disagreements addressed openly, resolved collaboratively",
     "Risk Factors": "Conflict Avoidance or Toxicity", "Risk Factors Description": "Conflicts either suppressed or become personal attacks",
     "Metric": "Conflict Resolution Quality", "Metric Description": "Survey: Conflicts addressed constructively, Disagreements resolved well (1-5 average)",
     "Unit": "Score 1-5", "Target": ">3.5/5.0 = healthy conflict dynamics", "Instructions": "Survey: Rate how conflicts are handled, quality of resolution processes. Average."},

    # LEVEL 2 - Learning Culture
    {"Level": 2, "Hierarchy": "Culture - Learning Culture", "Description": "Organizational commitment to continuous learning and improvement",
     "Business Drivers": "Capability Development", "Business Drivers Description": "Learning cultures adapt faster and innovate better",
     "Performance Factors": "Strong Learning Orientation", "Performance Factors Description": "Continuous learning embedded in work, supported by resources",
     "Risk Factors": "Learning Deficit", "Risk Factors Description": "Stagnant skills, resistant to new ideas, falling behind",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of learning culture sub-elements"},

    # LEVEL 3 - Learning Culture Sub-elements
    {"Level": 3, "Hierarchy": "Culture - Learning Culture - Learning Investment", "Description": "Resources dedicated to employee development",
     "Business Drivers": "Human Capital Development", "Business Drivers Description": "Investing in people improves performance and engagement",
     "Performance Factors": "Generous Development Budget", "Performance Factors Description": "Significant time and money allocated to learning",
     "Risk Factors": "Development Underfunding", "Risk Factors Description": "Learning seen as cost to cut, not investment",
     "Metric": "Learning Investment per Employee", "Metric Description": "Annual learning & development spend per FTE",
     "Unit": "Currency per employee", "Target": ">$1000/employee (knowledge work), >$500/employee (other), 40+ hours/year",
     "Instructions": "Calculate: Total L&D spend / Total FTEs. Include training, conferences, platforms, time."},

    {"Level": 3, "Hierarchy": "Culture - Learning Culture - Growth Mindset Prevalence", "Description": "Belief that abilities can be developed through effort",
     "Business Drivers": "Mindset Orientation", "Business Drivers Description": "Growth mindset (vs fixed) enables continuous improvement",
     "Performance Factors": "Growth Mindset Culture", "Performance Factors Description": "Challenges seen as opportunities, effort valued, learning from failure",
     "Risk Factors": "Fixed Mindset Culture", "Risk Factors Description": "Talent seen as fixed, avoid challenges, threatened by others' success",
     "Metric": "Growth Mindset Score", "Metric Description": "Survey: Abilities can be developed, Effort leads to mastery, Learn from failures (1-5 avg)",
     "Unit": "Score 1-5", "Target": ">4.0/5.0 = strong growth mindset", "Instructions": "Survey: Rate agreement with growth mindset statements. Use validated scale if available."},

    {"Level": 3, "Hierarchy": "Culture - Learning Culture - Knowledge Sharing Practices", "Description": "Systems and norms for transferring knowledge",
     "Business Drivers": "Organizational Learning", "Business Drivers Description": "Knowledge shared becomes organizational capability, not individual asset",
     "Performance Factors": "Active Knowledge Transfer", "Performance Factors Description": "Communities of practice, mentoring, documentation, lessons learned",
     "Risk Factors": "Knowledge Loss", "Risk Factors Description": "Knowledge walks out door with departures, lessons not captured",
     "Metric": "Knowledge Management Score", "Metric Description": "Assessment: Documentation systems, Communities of practice, Mentoring programs (score 0-10)",
     "Unit": "Composite score 0-30", "Target": ">20/30 = strong knowledge management", "Instructions": "Assess presence/quality of: Documentation systems, CoPs, Mentoring, Lessons learned. Score each 0-10."},

    {"Level": 3, "Hierarchy": "Culture - Learning Culture - Experimentation Encouragement", "Description": "Support for trying new approaches and methods",
     "Business Drivers": "Innovation Through Experiment", "Business Drivers Description": "Experimentation accelerates learning and innovation",
     "Performance Factors": "Experimentation Norms", "Performance Factors Description": "Small experiments encouraged, test-and-learn mindset",
     "Risk Factors": "Experimentation Resistance", "Risk Factors Description": "Always done it this way, change averse",
     "Metric": "Experimentation Rate", "Metric Description": "# active experiments or pilots / Total projects (or qualitative score 1-10)",
     "Unit": "Percentage or Score 1-10", "Target": ">20% of projects experimental, or >6/10 score",
     "Instructions": "Track projects with experimental/pilot approach. OR Survey: Rate encouragement to experiment 1-10."},

    # LEVEL 2 - Diversity & Inclusion
    {"Level": 2, "Hierarchy": "Culture - Diversity & Inclusion", "Description": "Representation of diverse groups and sense of belonging for all",
     "Business Drivers": "Talent Access & Innovation", "Business Drivers Description": "Diverse perspectives improve decision-making and innovation",
     "Performance Factors": "Diverse & Inclusive", "Performance Factors Description": "Diverse workforce where everyone feels valued and can contribute",
     "Risk Factors": "Homogeneity & Exclusion", "Risk Factors Description": "Lack of diversity, some groups feel excluded or marginalized",
     "Metric": "", "Metric Description": "", "Unit": "", "Target": "", "Instructions": "Aggregate of diversity & inclusion sub-elements"},

    # LEVEL 3 - Diversity & Inclusion Sub-elements
    {"Level": 3, "Hierarchy": "Culture - Diversity & Inclusion - Workforce Diversity", "Description": "Representation of diverse groups in workforce",
     "Business Drivers": "Equitable Hiring", "Business Drivers Description": "Intentional diversity efforts expand talent pool",
     "Performance Factors": "Representative Workforce", "Performance Factors Description": "Workforce demographics reflect community/customer base",
     "Risk Factors": "Homogeneous Workforce", "Risk Factors Description": "Lack of diversity in gender, ethnicity, background, perspective",
     "Metric": "Diversity Representation %", "Metric Description": "% women, % ethnic minorities, % other underrepresented groups vs targets",
     "Unit": "Percentage by group", "Target": "Representative of labor market/customer base, improving trend",
     "Instructions": "Track workforce composition by: Gender, Ethnicity, Age, Disability, etc. Compare to benchmarks."},

    {"Level": 3, "Hierarchy": "Culture - Diversity & Inclusion - Leadership Diversity", "Description": "Diversity in management and senior leadership",
     "Business Drivers": "Inclusive Leadership", "Business Drivers Description": "Diverse leadership signals real commitment and provides role models",
     "Performance Factors": "Diverse Leadership", "Performance Factors Description": "Management and senior roles reflect workforce diversity or better",
     "Risk Factors": "Leadership Homogeneity", "Risk Factors Description": "Diverse workforce but homogeneous leadership",
     "Metric": "Leadership Diversity %", "Metric Description": "% women/minorities in management vs workforce baseline",
     "Unit": "Percentage", "Target": "≥Workforce baseline, aspire to exceed", "Instructions": "Calculate diversity metrics for: Managers, Directors, Executives. Compare to workforce."},

    {"Level": 3, "Hierarchy": "Culture - Diversity & Inclusion - Inclusion Climate", "Description": "Extent to which all employees feel valued and included",
     "Business Drivers": "Belonging", "Business Drivers Description": "Diversity without inclusion doesn't yield benefits",
     "Performance Factors": "Strong Inclusion", "Performance Factors Description": "All employees feel they belong, can be authentic, are valued",
     "Risk Factors": "Exclusion Experience", "Risk Factors Description": "Some groups feel excluded, marginalized, or have to hide identity",
     "Metric": "Inclusion Score", "Metric Description": "Survey: Feel valued, Can be authentic, Respected (1-5 avg), disaggregated by group",
     "Unit": "Score 1-5", "Target": ">4.0/5.0 overall, no group <3.5", "Instructions": "Survey inclusion items. Disaggregate by demographic groups. Identify gaps."},

    {"Level": 3, "Hierarchy": "Culture - Diversity & Inclusion - Equitable Advancement", "Description": "Fairness in promotion and development opportunities",
     "Business Drivers": "Meritocracy", "Business Drivers Description": "Equitable advancement retains diverse talent and maintains credibility",
     "Performance Factors": "Fair Advancement", "Performance Factors Description": "Promotion rates similar across groups when performance equal",
     "Risk Factors": "Advancement Disparities", "Risk Factors Description": "Some groups advance slower despite equal or better performance",
     "Metric": "Promotion Parity", "Metric Description": "Promotion rate by demographic group, controlled for performance/tenure",
     "Unit": "Rate ratio (group vs majority)", "Target": "Ratio 0.9-1.1 = parity (within 10%)", "Instructions": "Calculate promotion rates by group. Control for: Performance rating, Tenure, Level. Identify gaps."},
]

# Write Culture data
row_idx = 2
for item in culture_data:
    ws_culture.cell(row=row_idx, column=1, value=item["Level"])
    ws_culture.cell(row=row_idx, column=2, value=item["Hierarchy"])
    ws_culture.cell(row=row_idx, column=3, value=item["Description"])
    ws_culture.cell(row=row_idx, column=4, value=item["Business Drivers"])
    ws_culture.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
    ws_culture.cell(row=row_idx, column=6, value=item["Performance Factors"])
    ws_culture.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
    ws_culture.cell(row=row_idx, column=8, value=item["Risk Factors"])
    ws_culture.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
    ws_culture.cell(row=row_idx, column=10, value=item["Metric"])
    ws_culture.cell(row=row_idx, column=11, value=item["Metric Description"])
    ws_culture.cell(row=row_idx, column=12, value=item["Unit"])
    ws_culture.cell(row=row_idx, column=13, value=item["Target"])
    ws_culture.cell(row=row_idx, column=14, value=item["Instructions"])
    row_idx += 1

# Auto-adjust
for column in ws_culture.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws_culture.column_dimensions[column_letter].width = adjusted_width

print("   Culture: {0} rows completed".format(len(culture_data)))
wb.save(wb_path)

print()
print("=" * 80)
print("CULTURE COMPLETE - 2 of 5 Enablers sheets done")
print("=" * 80)
print()
print("Note: People, Technology, Third Parties sheets are very large.")
print("Generating them will require separate focused scripts.")
print("=" * 80)
