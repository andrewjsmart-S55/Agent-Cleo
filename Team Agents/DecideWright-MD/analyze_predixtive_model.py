"""
Analyze the Predixtive_Model.xlsx spreadsheet structure
Focus on the Financial sheet and its 3-level data collection structure
"""

import openpyxl
import pandas as pd
from openpyxl import load_workbook

# Load the workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path, data_only=False)

print("=" * 80)
print("PREDIXTIVE MODEL SPREADSHEET ANALYSIS")
print("=" * 80)
print()

# List all sheets
print("1. AVAILABLE SHEETS:")
print("-" * 40)
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    print(f"  {idx}. {sheet_name:<30} ({sheet.max_row} rows × {sheet.max_column} cols)")
print()

# Focus on Financial sheet
if "Financial" in wb.sheetnames:
    ws = wb["Financial"]
    print("2. FINANCIAL SHEET STRUCTURE:")
    print("-" * 40)
    print(f"   Dimensions: {ws.max_row} rows × {ws.max_column} columns")
    print()

    # Extract headers (first 3 rows to see structure)
    print("   HEADER STRUCTURE:")
    for row_idx in range(1, min(4, ws.max_row + 1)):
        print(f"   Row {row_idx}:")
        row_data = []
        for col_idx in range(1, min(ws.max_column + 1, 20)):  # First 20 columns
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(f"      Col {col_idx}: {cell_value}")
        for item in row_data[:10]:  # Show first 10 non-empty
            print(item)
        if len(row_data) > 10:
            print(f"      ... and {len(row_data) - 10} more columns")
        print()

    # Analyze data structure - look for the 3 levels mentioned
    print("   DATA STRUCTURE ANALYSIS:")
    print()

    # Sample first 20 data rows
    print("   SAMPLE DATA (First 20 rows):")
    for row_idx in range(1, min(22, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(11, ws.max_column + 1)):  # First 10 columns
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(str(cell_value) if cell_value else "")

        # Format row output
        if row_idx == 1:
            print(f"   {'Row':<5} | " + " | ".join([f"{v:<25}" for v in row_values[:5]]))
            print("   " + "-" * 70)
        else:
            print(f"   {row_idx:<5} | " + " | ".join([f"{v:<25}" for v in row_values[:5]]))
    print()

    # Look for patterns that indicate 3 levels
    print("   IDENTIFYING 3-LEVEL STRUCTURE:")
    print()

    # Get column headers
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            headers.append((col_idx, header))

    print(f"   Total columns with headers: {len(headers)}")
    print()
    print("   Column Headers:")
    for col_idx, header in headers:
        print(f"      Col {col_idx:2d}: {header}")
    print()

    # Analyze for Level 1, Level 2, Level 3 patterns
    print("   SEARCHING FOR DATA COLLECTION LEVELS:")
    level_columns = {}
    for col_idx, header in headers:
        header_lower = str(header).lower()
        if 'level' in header_lower:
            level_columns[header] = col_idx
        elif 'tier' in header_lower:
            level_columns[header] = col_idx
        elif 'initial' in header_lower or 'basic' in header_lower:
            level_columns[header] = col_idx
        elif 'detailed' in header_lower or 'intermediate' in header_lower:
            level_columns[header] = col_idx
        elif 'advanced' in header_lower or 'comprehensive' in header_lower:
            level_columns[header] = col_idx

    if level_columns:
        print("   Found potential level indicators:")
        for header, col_idx in level_columns.items():
            print(f"      {header} (Column {col_idx})")
    else:
        print("   No explicit 'Level' columns found.")
        print("   Analyzing structure for implicit levels...")
    print()

    # Check for merged cells (might indicate levels)
    print("   MERGED CELLS (may indicate grouping/levels):")
    merged_ranges = list(ws.merged_cells.ranges)
    if merged_ranges:
        for idx, merged_range in enumerate(merged_ranges[:10], 1):
            print(f"      {idx}. {merged_range}")
        if len(merged_ranges) > 10:
            print(f"      ... and {len(merged_ranges) - 10} more merged ranges")
    else:
        print("      No merged cells found")
    print()

else:
    print("2. FINANCIAL SHEET NOT FOUND")
    print("   Available sheets:", wb.sheetnames)
    print()

# Check other sheets that might contain financial data
print("3. OTHER RELEVANT SHEETS:")
print("-" * 40)
for sheet_name in wb.sheetnames:
    if sheet_name != "Financial":
        ws = wb[sheet_name]
        # Get first row headers
        headers = []
        for col_idx in range(1, min(ws.max_column + 1, 10)):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                headers.append(str(header))

        print(f"   {sheet_name}:")
        print(f"      Rows: {ws.max_row}, Columns: {ws.max_column}")
        if headers:
            print(f"      Headers: {', '.join(headers[:5])}" + ("..." if len(headers) > 5 else ""))
        print()

print("=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
