"""
Generate Complete Economics Domain Sheets for VOC Predixtive Model
- Business Model
- External Environment
- Governance

Expert Senior Business Analyst - Strategy Execution & Risk Management
3-Level hierarchical structure for multivariate Bayesian analysis
"""

import openpyxl
from openpyxl import load_workbook

# Load workbook
wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

print("=" * 80)
print("GENERATING ECONOMICS DOMAIN SHEETS")
print("=" * 80)
print()

# ============================================================================
# BUSINESS MODEL SHEET
# ============================================================================

print("1. BUSINESS MODEL SHEET")
print("-" * 80)

ws_bm = wb["Business Model"]

# Clear existing content (keep headers but replace all data)
for row in ws_bm.iter_rows(min_row=2, max_row=ws_bm.max_row):
    for cell in row:
        cell.value = None

business_model_data = [
    # LEVEL 1: DIMENSION
    {
        "Level": 1,
        "Hierarchy": "Business Model",
        "Description": "How the organization creates, delivers, and captures value",
        "Business Drivers": "Value Creation Logic",
        "Business Drivers Description": "The fundamental logic of how the business generates value for customers and captures value for itself",
        "Performance Factors": "Business Model Effectiveness",
        "Performance Factors Description": "Strength, sustainability, and scalability of the value creation-delivery-capture cycle",
        "Risk Factors": "Business Model Failure",
        "Risk Factors Description": "Disruption, obsolescence, or fundamental flaws in value proposition or operating model",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of all business model elements - auto-calculated from Level 2"
    },

    # LEVEL 2: VALUE CREATION
    {
        "Level": 2,
        "Hierarchy": "Business Model - Value Creation",
        "Description": "How the organization designs and develops valuable offerings",
        "Business Drivers": "Customer Needs & Innovation",
        "Business Drivers Description": "Understanding customer problems and creating solutions that address them better than alternatives",
        "Performance Factors": "Value Proposition Strength",
        "Performance Factors Description": "Degree to which offerings solve real customer problems with differentiated solutions",
        "Risk Factors": "Value Proposition Failure",
        "Risk Factors Description": "Offerings fail to resonate with customers or get displaced by superior alternatives",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of value creation sub-elements"
    },

    # LEVEL 3: VALUE CREATION SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Creation - Customer Problem Clarity",
        "Description": "Depth of understanding of customer pain points and jobs to be done",
        "Business Drivers": "Customer Insight",
        "Business Drivers Description": "Research, feedback loops, and market intelligence that reveal true customer needs",
        "Performance Factors": "Problem-Solution Fit",
        "Performance Factors Description": "Evidence that offerings solve real, valuable customer problems",
        "Risk Factors": "Problem Misunderstanding",
        "Risk Factors Description": "Building solutions for problems customers don't have or don't care about",
        "Metric": "Customer Problem Validation Score",
        "Metric Description": "% of target customers who confirm the problem exists and is important (1-10 importance)",
        "Unit": "Percentage confirming problem + Average importance score",
        "Target": ">80% confirm problem exists, >7/10 importance",
        "Instructions": "Survey target customers: Do you experience [problem]? How important is solving it (1-10)?"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Creation - Solution Differentiation",
        "Description": "Unique aspects of the offering that create competitive advantage",
        "Business Drivers": "Innovation & IP",
        "Business Drivers Description": "Proprietary technology, unique processes, or business model innovation",
        "Performance Factors": "Differentiation Strength",
        "Performance Factors Description": "Clear, defensible differences vs competitors that customers value",
        "Risk Factors": "Commoditization",
        "Risk Factors Description": "Offerings become undifferentiated, competing only on price",
        "Metric": "Differentiation Score",
        "Metric Description": "Customer perception of uniqueness: Rate our solution vs alternatives (1-10 unique)",
        "Unit": "Score 1-10 (1=identical to competitors, 10=completely unique)",
        "Target": ">7/10 uniqueness score",
        "Instructions": "Customer survey: How different is our solution from alternatives? What makes it different?"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Creation - Innovation Pipeline",
        "Description": "Ongoing development of new offerings and improvements",
        "Business Drivers": "R&D Investment & Culture",
        "Business Drivers Description": "Resources and organizational capability dedicated to innovation",
        "Performance Factors": "Innovation Velocity",
        "Performance Factors Description": "Speed and success rate of bringing new offerings to market",
        "Risk Factors": "Innovation Stagnation",
        "Risk Factors Description": "Failure to evolve offerings as customer needs and technology change",
        "Metric": "New Product Revenue %",
        "Metric Description": "Revenue from products/features launched in past 12 months / total revenue",
        "Unit": "Percentage",
        "Target": ">20% for tech/innovation-driven; >10% for mature industries",
        "Instructions": "Calculate revenue from offerings launched in past 12 months. Track time-to-market."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Creation - Value Quantification",
        "Description": "Measurable customer value delivered (ROI, time saved, revenue gained)",
        "Business Drivers": "Tangible Customer Outcomes",
        "Business Drivers Description": "Ability to demonstrate concrete value creation for customers",
        "Performance Factors": "Proven ROI",
        "Performance Factors Description": "Documented evidence of customer value exceeding cost",
        "Risk Factors": "Value Claim Unsubstantiated",
        "Risk Factors Description": "Unable to prove value claims, leading to customer skepticism",
        "Metric": "Customer ROI Ratio",
        "Metric Description": "Customer value gained / customer cost paid (averaged across customers)",
        "Unit": "Ratio (e.g., 3:1 = $3 value per $1 cost)",
        "Target": ">3:1 for B2B; >5:1 for enterprise; varies by industry",
        "Instructions": "Calculate or survey: For every $1 spent with us, how much value do you receive?"
    },

    # LEVEL 2: VALUE DELIVERY
    {
        "Level": 2,
        "Hierarchy": "Business Model - Value Delivery",
        "Description": "How the organization delivers offerings to customers at scale",
        "Business Drivers": "Operations & Distribution",
        "Business Drivers Description": "Capabilities to produce and deliver offerings efficiently and reliably",
        "Performance Factors": "Delivery Excellence",
        "Performance Factors Description": "Quality, speed, and reliability of getting offerings to customers",
        "Risk Factors": "Delivery Failure",
        "Risk Factors Description": "Inability to deliver offerings at required quality, speed, or scale",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of value delivery sub-elements"
    },

    # LEVEL 3: VALUE DELIVERY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Delivery - Channel Effectiveness",
        "Description": "Performance of sales and distribution channels",
        "Business Drivers": "Go-to-Market Strategy",
        "Business Drivers Description": "Choice and execution of how to reach and serve target customers",
        "Performance Factors": "Channel Efficiency",
        "Performance Factors Description": "Channels deliver customers at optimal cost and with good experience",
        "Risk Factors": "Channel Conflict or Failure",
        "Risk Factors Description": "Channel partners underperform, conflict, or disintermediate",
        "Metric": "Customer Acquisition Cost by Channel",
        "Metric Description": "Total channel cost / customers acquired through that channel",
        "Unit": "Currency per customer",
        "Target": "CAC < (Customer LTV / 3) - varies by channel",
        "Instructions": "Calculate for each channel: sales costs + marketing costs + channel fees / # customers acquired"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Delivery - Operational Scalability",
        "Description": "Ability to increase delivery volume without proportional cost increase",
        "Business Drivers": "Leverage & Automation",
        "Business Drivers Description": "Business model enables serving more customers with limited incremental cost",
        "Performance Factors": "Economies of Scale",
        "Performance Factors Description": "Unit economics improve as volume increases",
        "Risk Factors": "Scalability Constraints",
        "Risk Factors Description": "Bottlenecks prevent growth or force expensive capacity additions",
        "Metric": "Marginal Cost of Next Customer",
        "Metric Description": "Incremental cost to serve one additional customer at current scale",
        "Unit": "Currency",
        "Target": "Approaching zero for digital/SaaS; <30% of price for physical goods",
        "Instructions": "Calculate variable cost per additional unit. Should decrease with scale."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Delivery - Customer Experience Quality",
        "Description": "End-to-end experience customers have with the organization",
        "Business Drivers": "Experience Design",
        "Business Drivers Description": "Deliberate design and execution of customer touchpoints and interactions",
        "Performance Factors": "Experience Excellence",
        "Performance Factors Description": "Customers rate experience as easy, fast, and pleasant",
        "Risk Factors": "Experience Degradation",
        "Risk Factors Description": "Poor experience drives customer churn and negative word-of-mouth",
        "Metric": "Net Promoter Score (NPS)",
        "Metric Description": "% Promoters (9-10) minus % Detractors (0-6) on recommendation likelihood",
        "Unit": "Score -100 to +100",
        "Target": ">50 excellent, >30 good, >0 acceptable, <0 action required",
        "Instructions": "Survey: How likely are you to recommend us (0-10)? Calculate NPS quarterly."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Delivery - Service Level Consistency",
        "Description": "Reliability and predictability of delivery performance",
        "Business Drivers": "Process Discipline",
        "Business Drivers Description": "Standardized, controlled processes ensure consistent outcomes",
        "Performance Factors": "Reliability",
        "Performance Factors Description": "Low variance in delivery time, quality, and customer satisfaction",
        "Risk Factors": "Service Inconsistency",
        "Risk Factors Description": "Variable quality or delivery erodes trust and satisfaction",
        "Metric": "Service Level Achievement %",
        "Metric Description": "% of deliveries meeting committed SLA (time, quality, completeness)",
        "Unit": "Percentage",
        "Target": ">95% for critical SLAs, >98% for premium tiers",
        "Instructions": "Track: On-time delivery %, First-time-right %, Customer satisfaction. Calculate monthly."
    },

    # LEVEL 2: VALUE CAPTURE
    {
        "Level": 2,
        "Hierarchy": "Business Model - Value Capture",
        "Description": "How the organization captures value as revenue and profit",
        "Business Drivers": "Monetization Strategy",
        "Business Drivers Description": "Pricing model and mechanisms to convert customer value into revenue",
        "Performance Factors": "Value Capture Efficiency",
        "Performance Factors Description": "Capturing fair share of value created through optimal pricing",
        "Risk Factors": "Value Leakage",
        "Risk Factors Description": "Creating value for customers but failing to capture appropriate return",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of value capture sub-elements"
    },

    # LEVEL 3: VALUE CAPTURE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Capture - Pricing Power",
        "Description": "Ability to set and maintain prices without demand destruction",
        "Business Drivers": "Differentiation & Switching Costs",
        "Business Drivers Description": "Strong value proposition creates pricing power vs commodities",
        "Performance Factors": "Price Realization",
        "Performance Factors Description": "Achieving list prices without excessive discounting",
        "Risk Factors": "Price Pressure",
        "Risk Factors Description": "Competition or buyer power forcing prices down",
        "Metric": "Discount Rate / Price Realization %",
        "Metric Description": "Average realized price / list price OR average discount given",
        "Unit": "Percentage (realization or discount)",
        "Target": ">90% price realization (<10% average discount)",
        "Instructions": "Calculate: Average actual price / list price. Track by customer segment and product."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Capture - Revenue Model Sustainability",
        "Description": "Predictability and recurring nature of revenue streams",
        "Business Drivers": "Business Model Type",
        "Business Drivers Description": "Subscription/recurring models provide more predictable revenue",
        "Performance Factors": "Revenue Predictability",
        "Performance Factors Description": "High % of revenue contracted or highly predictable",
        "Risk Factors": "Revenue Volatility",
        "Risk Factors Description": "Transactional model creates unpredictable revenue",
        "Metric": "Annual Recurring Revenue (ARR) %",
        "Metric Description": "Contracted recurring revenue / total revenue",
        "Unit": "Percentage",
        "Target": ">70% for SaaS, >50% for services, varies by business model",
        "Instructions": "Calculate revenue from subscriptions, contracts, maintenance / total revenue"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Capture - Customer Lifetime Value",
        "Description": "Total profit expected from customer over entire relationship",
        "Business Drivers": "Retention & Expansion",
        "Business Drivers Description": "Keeping customers long-term and growing their spend",
        "Performance Factors": "LTV Maximization",
        "Performance Factors Description": "High customer lifetime value relative to acquisition cost",
        "Risk Factors": "LTV Deterioration",
        "Risk Factors Description": "Customers churning faster or spending less over time",
        "Metric": "Customer Lifetime Value (LTV)",
        "Metric Description": "Average revenue per customer × gross margin × average customer lifespan",
        "Unit": "Currency",
        "Target": "LTV > 3× CAC (minimum), >5× CAC (good), >10× CAC (excellent)",
        "Instructions": "LTV = (Avg Monthly Revenue × Gross Margin %) / Monthly Churn Rate"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Value Capture - Monetization Alignment",
        "Description": "Alignment between pricing model and value delivered",
        "Business Drivers": "Value-Based Pricing",
        "Business Drivers Description": "Charging based on value delivered rather than cost or competition",
        "Performance Factors": "Pricing Model Fit",
        "Performance Factors Description": "Customers perceive pricing as fair and aligned with value received",
        "Risk Factors": "Pricing Misalignment",
        "Risk Factors Description": "Pricing model doesn't match customer value perception",
        "Metric": "Customer Value-to-Price Perception",
        "Metric Description": "Customer rating: Value received vs price paid (1-10)",
        "Unit": "Score 1-10 (1=expensive, 10=great value)",
        "Target": ">7/10 value perception",
        "Instructions": "Survey customers: Rate the value you receive relative to the price you pay (1-10)"
    },

    # LEVEL 2: REVENUE STREAMS
    {
        "Level": 2,
        "Hierarchy": "Business Model - Revenue Streams",
        "Description": "Different sources and types of revenue generation",
        "Business Drivers": "Portfolio Strategy",
        "Business Drivers Description": "Deliberate design of multiple revenue streams for growth and diversification",
        "Performance Factors": "Revenue Portfolio Health",
        "Performance Factors Description": "Balanced, growing revenue streams with low concentration risk",
        "Risk Factors": "Revenue Stream Concentration",
        "Risk Factors Description": "Over-dependence on single revenue source or customer type",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of revenue stream sub-elements"
    },

    # LEVEL 3: REVENUE STREAMS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Business Model - Revenue Streams - Primary Revenue Stream",
        "Description": "Largest or core revenue generating offering",
        "Business Drivers": "Core Value Proposition",
        "Business Drivers Description": "Main product/service that defines the business",
        "Performance Factors": "Core Stream Growth",
        "Performance Factors Description": "Primary revenue stream growing at target rate",
        "Risk Factors": "Core Stream Decline",
        "Risk Factors Description": "Main revenue source stagnating or declining",
        "Metric": "Primary Stream Revenue & Growth Rate",
        "Metric Description": "Revenue from core offering + YoY growth %",
        "Unit": "Currency + Percentage growth",
        "Target": "Growth rate ≥ strategic plan target (typically 15-30% for growth stage)",
        "Instructions": "Identify largest revenue stream. Track quarterly. Calculate YoY growth rate."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Revenue Streams - Adjacent Revenue Streams",
        "Description": "Additional revenue sources that leverage core capabilities",
        "Business Drivers": "Portfolio Extension",
        "Business Drivers Description": "Expanding into adjacent offerings for existing customers",
        "Performance Factors": "Portfolio Breadth",
        "Performance Factors Description": "Multiple revenue streams reducing concentration risk",
        "Risk Factors": "Diversification Failure",
        "Risk Factors Description": "New streams fail to gain traction or distract from core",
        "Metric": "Revenue Stream Count & Distribution",
        "Metric Description": "Number of revenue streams generating >5% of total revenue + Herfindahl Index",
        "Unit": "Count + HHI (0-10000, lower = more diversified)",
        "Target": "≥3 streams >5% of revenue, HHI <2500",
        "Instructions": "List all revenue streams >5% of total. Calculate HHI: Σ(stream % × stream %)×10000"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Revenue Streams - Emerging Revenue Streams",
        "Description": "New, experimental revenue sources in early stages",
        "Business Drivers": "Future Growth Options",
        "Business Drivers Description": "Testing new business models or market opportunities",
        "Performance Factors": "Innovation Portfolio",
        "Performance Factors Description": "Healthy pipeline of new revenue opportunities",
        "Risk Factors": "Future Growth Deficit",
        "Risk Factors Description": "No new revenue streams ready when core matures",
        "Metric": "Emerging Stream Investment & Traction",
        "Metric Description": "% of revenue invested in new streams + early traction metrics",
        "Unit": "Percentage investment + Key metrics (users, revenue, etc.)",
        "Target": "3-10% of revenue invested in innovation, ≥2 experiments with positive signals",
        "Instructions": "Track investment in new offerings. Measure early indicators: users, engagement, revenue."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Revenue Streams - Revenue Stream Synergy",
        "Description": "Degree to which revenue streams reinforce each other",
        "Business Drivers": "Strategic Architecture",
        "Business Drivers Description": "Designing streams that share resources and create cross-selling opportunities",
        "Performance Factors": "Cross-Sell Success",
        "Performance Factors Description": "Customers buying multiple offerings, increasing LTV",
        "Risk Factors": "Portfolio Cannibalization",
        "Risk Factors Description": "New streams stealing from existing rather than adding net new revenue",
        "Metric": "Cross-Sell Penetration %",
        "Metric Description": "% of customers buying 2+ offerings / total customers",
        "Unit": "Percentage",
        "Target": ">30% for B2B with multiple products, >50% for platform/ecosystem models",
        "Instructions": "Calculate: Customers with 2+ active offerings / total customers. Track expansion revenue."
    },

    # LEVEL 2: COST STRUCTURE
    {
        "Level": 2,
        "Hierarchy": "Business Model - Cost Structure",
        "Description": "Major cost drivers and economics of the business model",
        "Business Drivers": "Operating Model Design",
        "Business Drivers Description": "Structural choices that determine cost base and scalability",
        "Performance Factors": "Cost Structure Efficiency",
        "Performance Factors Description": "Low cost-to-serve with favorable unit economics",
        "Risk Factors": "Unfavorable Unit Economics",
        "Risk Factors Description": "Cost structure makes profitability impossible at scale",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of cost structure sub-elements"
    },

    # LEVEL 3: COST STRUCTURE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Business Model - Cost Structure - Unit Economics",
        "Description": "Profitability per unit/customer at steady state",
        "Business Drivers": "Contribution Margin",
        "Business Drivers Description": "Revenue per unit minus variable costs per unit",
        "Performance Factors": "Positive Unit Economics",
        "Performance Factors Description": "Each unit sold generates positive contribution to fixed costs",
        "Risk Factors": "Negative Unit Economics",
        "Risk Factors Description": "Losing money on each unit, unable to make up with volume",
        "Metric": "Contribution Margin per Unit/Customer",
        "Metric Description": "(Price - Variable Costs) / Price as %",
        "Unit": "Percentage or Currency per unit",
        "Target": ">60% for SaaS, >40% for services, >30% for products (varies by model)",
        "Instructions": "Calculate: (Revenue per customer - Variable costs per customer) / Revenue per customer"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Cost Structure - Fixed vs Variable Balance",
        "Description": "Proportion of costs that are fixed vs variable",
        "Business Drivers": "Operating Leverage",
        "Business Drivers Description": "Asset intensity and staffing model determine cost structure",
        "Performance Factors": "Leverage Optimization",
        "Performance Factors Description": "Cost structure aligned to business model and growth stage",
        "Risk Factors": "Leverage Mismatch",
        "Risk Factors Description": "High fixed costs with uncertain volumes create breakeven risk",
        "Metric": "Operating Leverage Ratio",
        "Metric Description": "Fixed costs / Total costs OR breakeven revenue level",
        "Unit": "Percentage or Currency (breakeven)",
        "Target": "Varies by business model: SaaS high leverage OK, services lower leverage",
        "Instructions": "Calculate fixed costs % and breakeven: Fixed costs / Contribution margin %"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Cost Structure - Cost of Customer Acquisition",
        "Description": "Total cost to acquire a new customer across all channels",
        "Business Drivers": "Marketing & Sales Efficiency",
        "Business Drivers Description": "Effectiveness of demand generation and sales conversion",
        "Performance Factors": "CAC Efficiency",
        "Performance Factors Description": "Low customer acquisition cost relative to lifetime value",
        "Risk Factors": "CAC Inflation",
        "Risk Factors Description": "Customer acquisition costs rising faster than LTV",
        "Metric": "Customer Acquisition Cost (CAC)",
        "Metric Description": "(Sales + Marketing Costs) / New Customers Acquired",
        "Unit": "Currency per customer",
        "Target": "LTV:CAC ratio >3:1 minimum, >5:1 good, payback <12 months",
        "Instructions": "Sum all sales + marketing expenses / # new customers. Calculate monthly, track trend."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Cost Structure - Cost Scalability",
        "Description": "How costs behave as revenue/volume increases",
        "Business Drivers": "Automation & Leverage",
        "Business Drivers Description": "Technology and process design enabling cost-effective scaling",
        "Performance Factors": "Economies of Scale",
        "Performance Factors Description": "Costs growing slower than revenue as business scales",
        "Risk Factors": "Diseconomies of Scale",
        "Risk Factors Description": "Complexity and coordination costs growing faster than revenue",
        "Metric": "Cost Growth Rate vs Revenue Growth Rate",
        "Metric Description": "YoY % change in operating costs vs YoY % change in revenue",
        "Unit": "Percentage (cost growth / revenue growth)",
        "Target": "Cost growth <70% of revenue growth (e.g., 20% cost growth with 30% revenue growth)",
        "Instructions": "Calculate YoY growth rates. Ratio <1.0 = positive leverage, <0.7 = strong leverage"
    },

    # LEVEL 2: BUSINESS MODEL RESILIENCE
    {
        "Level": 2,
        "Hierarchy": "Business Model - Business Model Resilience",
        "Description": "Durability and adaptability of the business model",
        "Business Drivers": "Strategic Defensibility",
        "Business Drivers Description": "Barriers to competition and ability to evolve with market changes",
        "Performance Factors": "Competitive Moat",
        "Performance Factors Description": "Defensible advantages that sustain profitability over time",
        "Risk Factors": "Business Model Disruption",
        "Risk Factors Description": "New entrants or substitutes rendering current model obsolete",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of business model resilience sub-elements"
    },

    # LEVEL 3: BUSINESS MODEL RESILIENCE SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "Business Model - Business Model Resilience - Competitive Barriers",
        "Description": "Structural advantages that deter competition",
        "Business Drivers": "Moat Creation",
        "Business Drivers Description": "Network effects, IP, scale economies, switching costs, brand",
        "Performance Factors": "Barrier Strength",
        "Performance Factors Description": "Competitors face significant disadvantages entering the market",
        "Risk Factors": "Barrier Erosion",
        "Risk Factors Description": "Competitive advantages weaken over time",
        "Metric": "Competitive Barrier Assessment",
        "Metric Description": "Score strength of barriers: Network effects, IP, Scale, Switching costs, Brand (each 1-10)",
        "Unit": "Composite score 5-50 (sum of 5 barriers)",
        "Target": ">35/50 strong moat, >25/50 moderate moat, <20/50 weak moat",
        "Instructions": "Rate each barrier type 1-10. Sum for composite score. Validate with market share trends."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Business Model Resilience - Market Position Strength",
        "Description": "Competitive standing and market share trajectory",
        "Business Drivers": "Competitive Success",
        "Business Drivers Description": "Winning against competitors in target markets",
        "Performance Factors": "Market Leadership",
        "Performance Factors Description": "Top 3 position in served markets with growing share",
        "Risk Factors": "Market Share Loss",
        "Risk Factors Description": "Losing ground to competitors or new entrants",
        "Metric": "Market Share & Trend",
        "Metric Description": "% of served addressable market + YoY change",
        "Unit": "Percentage + % point change",
        "Target": "Top 3 position in core markets, growing ≥1% point/year",
        "Instructions": "Calculate: Company revenue / Total served addressable market revenue. Track quarterly."
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Business Model Resilience - Business Model Adaptability",
        "Description": "Ability to evolve business model as markets change",
        "Business Drivers": "Strategic Agility",
        "Business Drivers Description": "Organizational capacity to sense and respond to disruption",
        "Performance Factors": "Evolution Capability",
        "Performance Factors Description": "Successfully adapting value proposition, delivery, or capture mechanisms",
        "Risk Factors": "Strategic Rigidity",
        "Risk Factors Description": "Locked into current model, unable to adapt to change",
        "Metric": "Business Model Innovation Index",
        "Metric Description": "Count of significant business model experiments/changes in past 3 years",
        "Unit": "Count of material business model pivots or tests",
        "Target": "≥2 significant experiments in 3 years for dynamic markets, ≥1 for stable markets",
        "Instructions": "Count material changes to: value proposition, channels, pricing model, customer segments"
    },
    {
        "Level": 3,
        "Hierarchy": "Business Model - Business Model Resilience - Customer Switching Costs",
        "Description": "Difficulty and cost for customers to switch to alternatives",
        "Business Drivers": "Lock-in Design",
        "Business Drivers Description": "Integration, data, training, or contractual elements that make switching difficult",
        "Performance Factors": "Retention Through Switching Costs",
        "Performance Factors Description": "High retention driven by switching barriers, not just satisfaction",
        "Risk Factors": "Low Switching Barriers",
        "Risk Factors Description": "Customers can easily leave for competitors",
        "Metric": "Customer Switching Cost Assessment",
        "Metric Description": "Estimated customer cost to switch (money + time + risk) as % of annual spend with you",
        "Unit": "Percentage of annual contract value",
        "Target": ">50% of ACV for B2B, >20% for B2C (highly defensible)",
        "Instructions": "Estimate: Migration costs + learning curve + integration + data/history loss / Annual contract value"
    },
]

# Write Business Model data
row_idx = 2
for item in business_model_data:
    ws_bm.cell(row=row_idx, column=1, value=item["Level"])
    ws_bm.cell(row=row_idx, column=2, value=item["Hierarchy"])
    ws_bm.cell(row=row_idx, column=3, value=item["Description"])
    ws_bm.cell(row=row_idx, column=4, value=item["Business Drivers"])
    ws_bm.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
    ws_bm.cell(row=row_idx, column=6, value=item["Performance Factors"])
    ws_bm.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
    ws_bm.cell(row=row_idx, column=8, value=item["Risk Factors"])
    ws_bm.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
    ws_bm.cell(row=row_idx, column=10, value=item["Metric"])
    ws_bm.cell(row=row_idx, column=11, value=item["Metric Description"])
    ws_bm.cell(row=row_idx, column=12, value=item["Unit"])
    ws_bm.cell(row=row_idx, column=13, value=item["Target"])
    ws_bm.cell(row=row_idx, column=14, value=item["Instructions"])
    row_idx += 1

# Auto-adjust column widths for Business Model
for column in ws_bm.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws_bm.column_dimensions[column_letter].width = adjusted_width

print(f"   Business Model: {len(business_model_data)} rows completed")
print(f"   - Level 1: 1 dimension")
print(f"   - Level 2: 6 elements")
print(f"   - Level 3: 24 sub-elements")
print()

# Save after Business Model
wb.save(wb_path)

# ============================================================================
# EXTERNAL ENVIRONMENT SHEET
# ============================================================================

print("2. EXTERNAL ENVIRONMENT SHEET")
print("-" * 80)

ws_ext = wb["External Environment"]

# Clear and populate
for row in ws_ext.iter_rows(min_row=2, max_row=ws_ext.max_row):
    for cell in row:
        cell.value = None

external_environment_data = [
    # LEVEL 1: DIMENSION
    {
        "Level": 1,
        "Hierarchy": "External Environment",
        "Description": "External forces and conditions affecting the organization",
        "Business Drivers": "Market Context & Forces",
        "Business Drivers Description": "Macro trends, competitive dynamics, and regulatory factors shaping the business environment",
        "Performance Factors": "Environmental Favorability",
        "Performance Factors Description": "External conditions support strategy execution and business model success",
        "Risk Factors": "Environmental Threats",
        "Risk Factors Description": "External shocks, disruptions, or adverse trends impacting operations",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of all external environment elements"
    },

    # LEVEL 2: MARKET DYNAMICS
    {
        "Level": 2,
        "Hierarchy": "External Environment - Market Dynamics",
        "Description": "Size, growth, and evolution of target markets",
        "Business Drivers": "Market Opportunity",
        "Business Drivers Description": "Addressable market size and growth rate determine opportunity ceiling",
        "Performance Factors": "Market Attractiveness",
        "Performance Factors Description": "Large, growing, accessible markets with favorable economics",
        "Risk Factors": "Market Decline",
        "Risk Factors Description": "Shrinking markets or market saturation limiting growth",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of market dynamics sub-elements"
    },

    # LEVEL 3: MARKET DYNAMICS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "External Environment - Market Dynamics - Total Addressable Market (TAM)",
        "Description": "Total market opportunity if 100% market share achieved",
        "Business Drivers": "Market Scope",
        "Business Drivers Description": "How broadly or narrowly the market is defined",
        "Performance Factors": "Market Size",
        "Performance Factors Description": "Large enough market to support growth ambitions",
        "Risk Factors": "Market Size Constraint",
        "Risk Factors Description": "Market too small to achieve scale or growth targets",
        "Metric": "Total Addressable Market (TAM)",
        "Metric Description": "Total annual revenue opportunity if captured 100% of target market",
        "Unit": "Currency (annual)",
        "Target": "TAM >10× current revenue (room to grow), >$500M for VC-backed",
        "Instructions": "Calculate: # potential customers × average revenue per customer OR total industry revenue"
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Market Dynamics - Market Growth Rate",
        "Description": "Speed at which the market is expanding or contracting",
        "Business Drivers": "Demand Trends",
        "Business Drivers Description": "Changing customer needs, adoption rates, and market penetration",
        "Performance Factors": "Market Tailwinds",
        "Performance Factors Description": "Operating in fast-growing market provides growth momentum",
        "Risk Factors": "Market Stagnation",
        "Risk Factors Description": "Flat or declining market makes growth difficult",
        "Metric": "Market Growth Rate (CAGR)",
        "Metric Description": "Compound annual growth rate of TAM over past 3-5 years",
        "Unit": "Percentage (CAGR)",
        "Target": ">10% for high-growth markets, >5% for established markets",
        "Instructions": "Research industry reports, analyst estimates. Calculate: ((Ending TAM / Beginning TAM)^(1/years) - 1)"
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Market Dynamics - Market Maturity Stage",
        "Description": "Lifecycle stage of the market (emerging, growth, mature, declining)",
        "Business Drivers": "Adoption Curve",
        "Business Drivers Description": "Position on S-curve of market adoption",
        "Performance Factors": "Lifecycle Fit",
        "Performance Factors Description": "Strategy aligned to market maturity (land grab vs margin optimization)",
        "Risk Factors": "Lifecycle Mismatch",
        "Risk Factors Description": "Strategy inappropriate for market stage (e.g., focusing on share in declining market)",
        "Metric": "Market Maturity Assessment",
        "Metric Description": "Stage: Emerging (0-25% penetration), Growth (25-75%), Mature (75-90%), Decline (90%+)",
        "Unit": "Categorical + penetration %",
        "Target": "Aligned strategy: Growth stage = market share focus, Mature = profitability focus",
        "Instructions": "Assess: Current users / Potential users. Map to stage. Validate with growth rate trends."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Market Dynamics - Market Fragmentation",
        "Description": "Concentration vs dispersion of market participants",
        "Business Drivers": "Competitive Structure",
        "Business Drivers Description": "Number and relative size of players in the market",
        "Performance Factors": "Consolidation Opportunity",
        "Performance Factors Description": "Fragmented markets offer share gain opportunities",
        "Risk Factors": "Market Concentration",
        "Risk Factors Description": "Dominated by incumbents, difficult to gain share",
        "Metric": "Market Concentration Ratio (CR4)",
        "Metric Description": "Combined market share of top 4 players",
        "Unit": "Percentage",
        "Target": "<40% = fragmented (opportunity), 40-60% = moderate, >60% = concentrated (barrier)",
        "Instructions": "Calculate: (Top 4 competitors' revenue / Total market revenue) × 100"
    },

    # Continue with remaining External Environment sections...
    # (I'll include the rest in the next sections for length management)
]

# Continue with External Environment data structure
external_environment_data.extend([
    # LEVEL 2: COMPETITIVE INTENSITY
    {
        "Level": 2,
        "Hierarchy": "External Environment - Competitive Intensity",
        "Description": "Degree and nature of competitive pressure",
        "Business Drivers": "Industry Structure",
        "Business Drivers Description": "Number of competitors, differentiation, and rivalry levels",
        "Performance Factors": "Competitive Positioning",
        "Performance Factors Description": "Favorable competitive position with defensible advantages",
        "Risk Factors": "Competitive Pressure",
        "Risk Factors Description": "Intense competition eroding margins and growth",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of competitive intensity sub-elements"
    },

    # LEVEL 3: COMPETITIVE INTENSITY SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "External Environment - Competitive Intensity - Number of Competitors",
        "Description": "Quantity of direct competitors in served markets",
        "Business Drivers": "Barriers to Entry",
        "Business Drivers Description": "Low barriers allow more entrants; high barriers limit competition",
        "Performance Factors": "Rational Competition",
        "Performance Factors Description": "Limited number of rational competitors maintaining healthy margins",
        "Risk Factors": "Overcrowded Market",
        "Risk Factors Description": "Too many players chasing limited customers drives margin pressure",
        "Metric": "Direct Competitor Count",
        "Metric Description": "Number of companies offering comparable solutions to same customer segments",
        "Unit": "Count",
        "Target": "<10 direct competitors = healthy, >25 = highly competitive",
        "Instructions": "List competitors offering similar value propositions to similar customers. Count."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Competitive Intensity - Competitive Differentiation",
        "Description": "Degree of differentiation among market participants",
        "Business Drivers": "Innovation & Positioning",
        "Business Drivers Description": "Players creating unique value propositions vs commoditized offerings",
        "Performance Factors": "Differentiation Opportunity",
        "Performance Factors Description": "Room to differentiate and avoid price competition",
        "Risk Factors": "Commoditization",
        "Risk Factors Description": "Offerings becoming undifferentiated, forcing price competition",
        "Metric": "Market Differentiation Score",
        "Metric Description": "Customer perception: How different are offerings in this market? (1-10 scale)",
        "Unit": "Score 1-10 (1=all same, 10=highly differentiated)",
        "Target": ">6/10 = differentiated market, <4/10 = commoditized",
        "Instructions": "Customer/expert survey: Rate differentiation among competitors. Average scores."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Competitive Intensity - Pricing Pressure",
        "Description": "Downward pressure on prices from competition",
        "Business Drivers": "Supply-Demand Balance",
        "Business Drivers Description": "Excess capacity or aggressive competitors drive price competition",
        "Performance Factors": "Price Stability",
        "Performance Factors Description": "Stable or rising prices indicate rational competition",
        "Risk Factors": "Price War",
        "Risk Factors Description": "Race to bottom on pricing destroying profitability",
        "Metric": "Industry Price Trend",
        "Metric Description": "Average price change in market over past 3 years (inflation-adjusted)",
        "Unit": "Percentage change (real terms)",
        "Target": "Stable to rising (0% to +5% real growth)",
        "Instructions": "Research industry pricing trends. Calculate real (inflation-adjusted) price changes."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Competitive Intensity - Threat of New Entrants",
        "Description": "Likelihood of new competitors entering the market",
        "Business Drivers": "Entry Barriers",
        "Business Drivers Description": "Capital requirements, regulation, IP, economies of scale, network effects",
        "Performance Factors": "Protected Market Position",
        "Performance Factors Description": "High barriers deter new entrants, protecting incumbents",
        "Risk Factors": "New Entrant Disruption",
        "Risk Factors Description": "New competitors with innovative models or deep pockets",
        "Metric": "Entry Barrier Assessment",
        "Metric Description": "Score barriers: Capital, Regulation, IP, Scale, Network (each 1-10, 10=high barrier)",
        "Unit": "Composite score 5-50",
        "Target": ">30/50 = high barriers (favorable), <20/50 = low barriers (risk)",
        "Instructions": "Rate each barrier type 1-10. Sum. Validate with new entrant activity past 3 years."
    },

    # LEVEL 2: REGULATORY ENVIRONMENT
    {
        "Level": 2,
        "Hierarchy": "External Environment - Regulatory Environment",
        "Description": "Legal, regulatory, and compliance requirements",
        "Business Drivers": "Regulatory Regime",
        "Business Drivers Description": "Government oversight and rules governing the industry",
        "Performance Factors": "Regulatory Stability",
        "Performance Factors Description": "Clear, stable regulations that don't create excessive burden",
        "Risk Factors": "Regulatory Risk",
        "Risk Factors Description": "Adverse regulatory changes or non-compliance consequences",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of regulatory environment sub-elements"
    },

    # LEVEL 3: REGULATORY ENVIRONMENT SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "External Environment - Regulatory Environment - Regulatory Intensity",
        "Description": "Degree of regulatory oversight and requirements",
        "Business Drivers": "Industry Risk Profile",
        "Business Drivers Description": "High-risk industries (finance, healthcare, energy) face heavier regulation",
        "Performance Factors": "Compliance Capability",
        "Performance Factors Description": "Systems and processes to manage regulatory requirements efficiently",
        "Risk Factors": "Regulatory Burden",
        "Risk Factors Description": "Excessive compliance costs limiting competitiveness",
        "Metric": "Regulatory Intensity Score",
        "Metric Description": "Assessment: Light (minimal), Moderate (some oversight), Heavy (highly regulated), Extreme (financial services)",
        "Unit": "Categorical",
        "Target": "Compliance costs <5% of revenue for Heavy, <2% for Moderate",
        "Instructions": "Assess against framework. Measure compliance costs as % of revenue."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Regulatory Environment - Regulatory Trend",
        "Description": "Direction of regulatory change (tightening vs loosening)",
        "Business Drivers": "Political & Social Pressure",
        "Business Drivers Description": "Public sentiment and political priorities drive regulatory changes",
        "Performance Factors": "Regulatory Tailwinds",
        "Performance Factors Description": "Deregulation or favorable policy changes supporting growth",
        "Risk Factors": "Regulatory Tightening",
        "Risk Factors Description": "New regulations increasing costs or constraining operations",
        "Metric": "Regulatory Trend Assessment",
        "Metric Description": "Assessment: Liberalizing (easing), Stable (unchanged), Tightening (increasing)",
        "Unit": "Categorical + impact estimate",
        "Target": "Stable to liberalizing; if tightening, manageable cost impact",
        "Instructions": "Track pending legislation, regulatory proposals. Estimate cost impact if enacted."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Regulatory Environment - Compliance Risk Level",
        "Description": "Exposure to penalties or business impact from non-compliance",
        "Business Drivers": "Enforcement Environment",
        "Business Drivers Description": "Regulatory agencies' focus and enforcement rigor",
        "Performance Factors": "Compliance Excellence",
        "Performance Factors Description": "Clean compliance record with no material violations",
        "Risk Factors": "Non-Compliance Exposure",
        "Risk Factors Description": "Violations could result in fines, license loss, or reputational damage",
        "Metric": "Compliance Violation History",
        "Metric Description": "Count and severity of violations past 3 years + potential exposure",
        "Unit": "Count + Currency (max potential fine)",
        "Target": "Zero material violations, <0.1% revenue potential exposure",
        "Instructions": "List violations past 3 years. Assess max potential fine for worst-case scenario."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Regulatory Environment - Regulatory Advantage",
        "Description": "Competitive advantage derived from regulatory factors",
        "Business Drivers": "License or Approval",
        "Business Drivers Description": "Regulatory barriers creating protected market positions",
        "Performance Factors": "Regulatory Moat",
        "Performance Factors Description": "Licenses, approvals, or compliance capabilities others lack",
        "Risk Factors": "Regulatory Disadvantage",
        "Risk Factors Description": "Competitors have better regulatory positions or access",
        "Metric": "Regulatory Position Score",
        "Metric Description": "Assessment of regulatory advantages vs competitors (1-10 scale)",
        "Unit": "Score 1-10 (1=disadvantaged, 10=strong advantage)",
        "Target": ">6/10 = regulatory advantage or neutrality",
        "Instructions": "Assess: licenses held, approval speed, regulatory relationships vs competitors."
    },

    # LEVEL 2: ECONOMIC CONDITIONS
    {
        "Level": 2,
        "Hierarchy": "External Environment - Economic Conditions",
        "Description": "Macroeconomic factors affecting business performance",
        "Business Drivers": "Economic Cycle",
        "Business Drivers Description": "GDP growth, employment, inflation, interest rates shape demand",
        "Performance Factors": "Economic Tailwinds",
        "Performance Factors Description": "Favorable economic conditions supporting growth",
        "Risk Factors": "Economic Headwinds",
        "Risk Factors Description": "Recession, inflation, or adverse economic trends",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of economic conditions sub-elements"
    },

    # LEVEL 3: ECONOMIC CONDITIONS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "External Environment - Economic Conditions - GDP Growth Trend",
        "Description": "Economic growth in primary operating markets",
        "Business Drivers": "Overall Economic Activity",
        "Business Drivers Description": "GDP growth creates tailwind for most businesses",
        "Performance Factors": "Economic Growth Support",
        "Performance Factors Description": "Operating in growing economies with rising incomes",
        "Risk Factors": "Economic Contraction",
        "Risk Factors Description": "Recession or slow growth limiting demand",
        "Metric": "Weighted Average GDP Growth",
        "Metric Description": "GDP growth rate of operating markets weighted by revenue exposure",
        "Unit": "Percentage (weighted avg)",
        "Target": ">2% for developed markets, >4% for emerging markets",
        "Instructions": "Calculate: Σ(Market GDP growth × % revenue from that market)"
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Economic Conditions - Inflation Environment",
        "Description": "Rate of price increases affecting costs and purchasing power",
        "Business Drivers": "Monetary Policy & Supply-Demand",
        "Business Drivers Description": "Central bank policy and supply chain dynamics drive inflation",
        "Performance Factors": "Inflation Management",
        "Performance Factors Description": "Ability to pass through cost increases to customers",
        "Risk Factors": "Margin Squeeze",
        "Risk Factors Description": "Input cost inflation faster than ability to raise prices",
        "Metric": "Inflation Rate & Pricing Power",
        "Metric Description": "CPI/PPI in operating markets + own price increase rate",
        "Unit": "Percentage (inflation) + Percentage (price increases)",
        "Target": "Price increases ≥ Inflation rate (maintaining real prices)",
        "Instructions": "Track relevant inflation indices. Monitor own price realization vs inflation."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Economic Conditions - Interest Rate Environment",
        "Description": "Cost of capital and financing conditions",
        "Business Drivers": "Monetary Policy",
        "Business Drivers Description": "Central bank rates determine borrowing costs",
        "Performance Factors": "Favorable Financing",
        "Performance Factors Description": "Low cost of capital supporting investment and growth",
        "Risk Factors": "Financing Cost Inflation",
        "Risk Factors Description": "Rising rates increasing debt service and reducing valuations",
        "Metric": "Interest Rate Trend & Cost of Debt",
        "Metric Description": "Prime/benchmark rate + weighted average interest rate on debt",
        "Unit": "Percentage",
        "Target": "Debt service coverage ratio >2.5× (EBITDA / Interest expense)",
        "Instructions": "Monitor central bank rates. Calculate weighted average rate on all debt."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Economic Conditions - Economic Cycle Position",
        "Description": "Current stage in economic expansion/contraction cycle",
        "Business Drivers": "Business Cycle Dynamics",
        "Business Drivers Description": "Economies move through expansion, peak, contraction, trough cycles",
        "Performance Factors": "Cycle Awareness",
        "Performance Factors Description": "Strategy adapted to cycle position (invest in expansion, conserve in contraction)",
        "Risk Factors": "Cycle Misread",
        "Risk Factors Description": "Expanding at peak or contracting in expansion",
        "Metric": "Economic Cycle Assessment",
        "Metric Description": "Current position: Early expansion, Mid expansion, Peak, Early contraction, Recession, Trough",
        "Unit": "Categorical",
        "Target": "Strategy aligned: Expansion = invest/grow, Peak/Contraction = optimize/conserve",
        "Instructions": "Assess using: GDP trend, unemployment, capacity utilization, leading indicators"
    },

    # LEVEL 2: TECHNOLOGY TRENDS
    {
        "Level": 2,
        "Hierarchy": "External Environment - Technology Trends",
        "Description": "Technological changes affecting industry and business model",
        "Business Drivers": "Innovation Waves",
        "Business Drivers Description": "New technologies creating opportunities and threats",
        "Performance Factors": "Technology Adoption",
        "Performance Factors Description": "Leveraging new technologies for competitive advantage",
        "Risk Factors": "Technology Disruption",
        "Risk Factors Description": "New technologies making current model obsolete",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of technology trends sub-elements"
    },

    # LEVEL 3: TECHNOLOGY TRENDS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "External Environment - Technology Trends - Disruptive Technology Threats",
        "Description": "Emerging technologies that could disrupt current business model",
        "Business Drivers": "Innovation Cycles",
        "Business Drivers Description": "New technology paradigms periodically disrupt industries",
        "Performance Factors": "Technology Monitoring",
        "Performance Factors Description": "Early identification of disruptive threats and opportunities",
        "Risk Factors": "Technology Blindness",
        "Risk Factors Description": "Missing disruptive technologies until too late to adapt",
        "Metric": "Disruptive Technology Assessment",
        "Metric Description": "Identification of top 3 disruptive threats + maturity assessment + response plan",
        "Unit": "List of technologies + maturity stage + response status",
        "Target": "Top threats identified, maturity tracked, response plans active",
        "Instructions": "Identify emerging technologies impacting industry. Assess: Emerging / Developing / Mature"
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Technology Trends - Technology Adoption Rate",
        "Description": "Speed at which new technologies are adopted in industry",
        "Business Drivers": "Innovation Culture",
        "Business Drivers Description": "Industry willingness and ability to adopt new technologies",
        "Performance Factors": "Fast Follower Position",
        "Performance Factors Description": "Adopting proven technologies quickly without bleeding-edge risk",
        "Risk Factors": "Technology Lag",
        "Risk Factors Description": "Falling behind competitors in technology adoption",
        "Metric": "Technology Adoption Relative to Industry",
        "Metric Description": "Assessment: Laggard, Follower, Early majority, Early adopter, Innovator",
        "Unit": "Categorical position on adoption curve",
        "Target": "Early majority or Early adopter (fast follower sweet spot)",
        "Instructions": "Compare technology stack maturity to industry benchmarks. Position on curve."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Technology Trends - Digital Transformation Pressure",
        "Description": "Urgency and necessity of digital transformation in industry",
        "Business Drivers": "Customer & Competitive Expectations",
        "Business Drivers Description": "Stakeholders expecting digital-first experiences",
        "Performance Factors": "Digital Maturity",
        "Performance Factors Description": "Digital capabilities match or exceed industry standards",
        "Risk Factors": "Digital Disruption",
        "Risk Factors Description": "Digital-native competitors displacing traditional players",
        "Metric": "Digital Maturity Score",
        "Metric Description": "Assessment across: Customer experience, Operations, Business model, Data/analytics (1-10 each)",
        "Unit": "Composite score 4-40",
        "Target": ">30/40 for digital industries, >20/40 for traditional",
        "Instructions": "Rate 4 dimensions 1-10 each. Sum. Benchmark against digital leaders in industry."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Technology Trends - Technology Investment Trends",
        "Description": "Where competitors and industry are investing in technology",
        "Business Drivers": "Strategic Bets",
        "Business Drivers Description": "Industry consensus on important technology directions",
        "Performance Factors": "Technology Investment Parity",
        "Performance Factors Description": "Technology investments aligned with industry and strategic needs",
        "Risk Factors": "Technology Investment Gaps",
        "Risk Factors Description": "Underinvestment in critical technologies vs competitors",
        "Metric": "Technology Investment as % Revenue",
        "Metric Description": "IT + Digital + R&D investment / Revenue vs industry average",
        "Unit": "Percentage + comparison to industry",
        "Target": "Within 80-120% of industry average unless differentiation strategy",
        "Instructions": "Calculate: (IT + Digital + R&D) / Revenue. Compare to industry benchmarks."
    },

    # LEVEL 2: SOCIO-POLITICAL FACTORS
    {
        "Level": 2,
        "Hierarchy": "External Environment - Socio-Political Factors",
        "Description": "Social trends and political conditions affecting business",
        "Business Drivers": "Social & Political Context",
        "Business Drivers Description": "Demographics, values, political stability shape operating environment",
        "Performance Factors": "Social License",
        "Performance Factors Description": "Positive standing with stakeholders and society",
        "Risk Factors": "Social or Political Risk",
        "Risk Factors Description": "Social backlash, political instability, or policy changes",
        "Metric": "",
        "Metric Description": "",
        "Unit": "",
        "Target": "",
        "Instructions": "Aggregate of socio-political factors sub-elements"
    },

    # LEVEL 3: SOCIO-POLITICAL FACTORS SUB-ELEMENTS
    {
        "Level": 3,
        "Hierarchy": "External Environment - Socio-Political Factors - Political Stability",
        "Description": "Stability of governments in operating markets",
        "Business Drivers": "Governance Quality",
        "Business Drivers Description": "Stable political systems enable business predictability",
        "Performance Factors": "Political Risk Mitigation",
        "Performance Factors Description": "Operating primarily in stable democracies with rule of law",
        "Risk Factors": "Political Instability",
        "Risk Factors Description": "Coups, civil unrest, or erratic policy changes disrupting operations",
        "Metric": "Political Risk Score",
        "Metric Description": "Weighted average of country political risk scores (0-100, lower = higher risk)",
        "Unit": "Risk score 0-100 (weighted by revenue)",
        "Target": "<30 = low risk, 30-60 = moderate, >60 = high risk",
        "Instructions": "Use World Bank Governance Indicators or similar. Weight by revenue exposure per country."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Socio-Political Factors - Social Trends Alignment",
        "Description": "Alignment of business model with social trends and values",
        "Business Drivers": "Social Acceptance",
        "Business Drivers Description": "Business practices align with evolving social expectations",
        "Performance Factors": "Social License to Operate",
        "Performance Factors Description": "Strong public support and stakeholder acceptance",
        "Risk Factors": "Social Backlash",
        "Risk Factors Description": "Public opposition or boycotts due to practices or positions",
        "Metric": "Brand Reputation Score",
        "Metric Description": "Public perception rating: Trust, ethics, social responsibility (1-10 scale)",
        "Unit": "Score 1-10",
        "Target": ">7/10 = strong reputation, <5/10 = reputational risk",
        "Instructions": "Track: brand surveys, social media sentiment, ESG ratings, media coverage tone"
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Socio-Political Factors - Demographic Trends",
        "Description": "Population and demographic shifts affecting customer base",
        "Business Drivers": "Population Dynamics",
        "Business Drivers Description": "Age, income, education, urbanization trends shape markets",
        "Performance Factors": "Demographic Tailwinds",
        "Performance Factors Description": "Favorable demographic trends expanding target markets",
        "Risk Factors": "Adverse Demographics",
        "Risk Factors Description": "Shrinking, aging, or changing demographics reducing addressable market",
        "Metric": "Target Demographic Growth Rate",
        "Metric Description": "Annual growth rate of primary customer demographic segments",
        "Unit": "Percentage growth",
        "Target": ">2% growth in key demographics",
        "Instructions": "Identify key customer demographics. Track growth rates from census/projections."
    },
    {
        "Level": 3,
        "Hierarchy": "External Environment - Socio-Political Factors - ESG Expectations",
        "Description": "Environmental, social, governance expectations from stakeholders",
        "Business Drivers": "Stakeholder Values",
        "Business Drivers Description": "Investors, customers, employees increasingly prioritizing ESG",
        "Performance Factors": "ESG Leadership",
        "Performance Factors Description": "Meeting or exceeding ESG expectations across stakeholder groups",
        "Risk Factors": "ESG Gaps",
        "Risk Factors Description": "Failure to meet ESG standards damaging reputation and access to capital",
        "Metric": "ESG Performance Score",
        "Metric Description": "Rating from MSCI, Sustainalytics, or internal framework (0-100)",
        "Unit": "Score 0-100",
        "Target": ">60 = acceptable, >75 = good, >85 = leader",
        "Instructions": "Obtain third-party ESG rating or conduct self-assessment against standard frameworks"
    },
])

# Write External Environment data
row_idx = 2
for item in external_environment_data:
    ws_ext.cell(row=row_idx, column=1, value=item["Level"])
    ws_ext.cell(row=row_idx, column=2, value=item["Hierarchy"])
    ws_ext.cell(row=row_idx, column=3, value=item["Description"])
    ws_ext.cell(row=row_idx, column=4, value=item["Business Drivers"])
    ws_ext.cell(row=row_idx, column=5, value=item["Business Drivers Description"])
    ws_ext.cell(row=row_idx, column=6, value=item["Performance Factors"])
    ws_ext.cell(row=row_idx, column=7, value=item["Performance Factors Description"])
    ws_ext.cell(row=row_idx, column=8, value=item["Risk Factors"])
    ws_ext.cell(row=row_idx, column=9, value=item["Risk Factors Description"])
    ws_ext.cell(row=row_idx, column=10, value=item["Metric"])
    ws_ext.cell(row=row_idx, column=11, value=item["Metric Description"])
    ws_ext.cell(row=row_idx, column=12, value=item["Unit"])
    ws_ext.cell(row=row_idx, column=13, value=item["Target"])
    ws_ext.cell(row=row_idx, column=14, value=item["Instructions"])
    row_idx += 1

# Auto-adjust column widths
for column in ws_ext.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws_ext.column_dimensions[column_letter].width = adjusted_width

print(f"   External Environment: {len(external_environment_data)} rows completed")
print(f"   - Level 1: 1 dimension")
print(f"   - Level 2: 6 elements")
print(f"   - Level 3: 24 sub-elements")
print()

# Save after External Environment
wb.save(wb_path)

print("=" * 80)
print("SCRIPT EXECUTION COMPLETE")
print("=" * 80)
print()
print(f"Location: {wb_path}")
print()
print("COMPLETED SHEETS:")
print("  1. Business Model: 31 rows (1 dimension, 6 elements, 24 sub-elements)")
print("  2. External Environment: 31 rows (1 dimension, 6 elements, 24 sub-elements)")
print()
print("NOTE: Governance sheet will be generated in next script")
print("=" * 80)
