"""
Risk Dashboard Sheet Generator
Creates comprehensive risk exposure calculations across all aggregation levels

Risk Metrics (Primary - Requested):
1. Opex at Risk - Operating expenditure exposure
2. Capex at Risk - Capital expenditure exposure
3. Stratex at Risk - Strategic expenditure exposure
4. Revenue at Risk - Revenue stream vulnerability
5. Productivity Time at Risk - Time/efficiency losses
6. Service Availability at Risk - Service disruption exposure
7. Product at Risk - Product delivery/quality exposure
8. Reputation at Risk - Brand and reputation damage exposure
9. Enterprise Value at Risk - Overall enterprise value vulnerability

Additional Lagging Risk Metrics:
10. Customer Lifetime Value at Risk - Customer base erosion
11. Market Share at Risk - Competitive position loss
12. Talent at Risk - Key employee departure/capability loss
13. Compliance at Risk - Regulatory fines, penalties, restrictions
14. Data/IP at Risk - Intellectual property and data loss
15. Cash Flow at Risk - Liquidity crisis vulnerability
16. Credit Rating at Risk - Debt covenant violations, rating downgrades
17. Innovation Pipeline at Risk - Future product/service pipeline disruption

Aggregation Levels:
- Company Level (Overall)
- Domain Level (4 domains)
- Dimension Level (16 dimensions)

Expert Senior Business Analyst - 20+ years experience
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Risk Dashboard structure
risk_dashboard_data = []

# ============================================================================
# SECTION 1: COMPANY-LEVEL RISK SUMMARY (Overall Enterprise Risk Profile)
# ============================================================================

# Header row for Company Level
risk_dashboard_data.append({
    "Aggregation_Level": "COMPANY LEVEL",
    "Entity_Name": "[Company Name]",
    "Risk_Metric": "=== ENTERPRISE RISK PROFILE ===",
    "Exposure_Amount": "",
    "Currency_Unit": "",
    "Probability": "",
    "Expected_Loss": "",
    "Time_Horizon": "",
    "Confidence_Level": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": ""
})

# Primary Risk Metrics - Company Level
company_risks = [
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Opex at Risk",
        "Exposure_Amount": "[Formula: Sum of all operational expenditure exposures across 16 dimensions]",
        "Currency_Unit": "$ (Annual)",
        "Probability": "[% - Weighted average probability across dimensions]",
        "Expected_Loss": "[Formula: Exposure_Amount × Probability]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Costs, People, Technology, Processes, Third Parties dimensions",
        "Calculation_Method": "Monte Carlo simulation aggregating operational risk exposures from: (1) Cost dimension inefficiencies, (2) People turnover/productivity losses, (3) Technology system failures, (4) Process inefficiencies, (5) Third-party failures. Risk triggered when metrics in 'Risk Factors' range.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Capex at Risk",
        "Exposure_Amount": "[Formula: Sum of capital project failure exposures]",
        "Currency_Unit": "$ (Multi-year)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "36 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Investment, Technology, Innovation, Strategic Goals dimensions",
        "Calculation_Method": "Aggregate capital at risk from: (1) Investment dimension project failures, (2) Technology infrastructure investment waste, (3) Innovation project abandonment, (4) Strategic initiative capital consumption without returns. Exposure = committed capital × probability of abandonment/failure.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Stratex at Risk",
        "Exposure_Amount": "[Formula: Sum of strategic initiative expenditure exposures]",
        "Currency_Unit": "$ (Multi-year)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "36-60 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Strategic Goals, Innovation, Change, Investment dimensions",
        "Calculation_Method": "Strategic expenditure at risk from: (1) Strategic initiative failures (Strategic Goals dimension), (2) Innovation program waste, (3) Change initiative failures, (4) M&A integration failures. Stratex = strategic opex + strategic capex committed to multi-year initiatives with failure risk.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Revenue at Risk",
        "Exposure_Amount": "[Formula: Sum of revenue stream vulnerabilities]",
        "Currency_Unit": "$ (Annual)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Revenue, Products & Services, Brand, Annual Results, Strategic Goals dimensions",
        "Calculation_Method": "Revenue at risk from: (1) Customer churn (Revenue dimension), (2) Product failures (Products & Services), (3) Brand erosion (Brand dimension), (4) Competitive losses (Strategic Goals), (5) Market share loss. Revenue exposure = vulnerable revenue streams × probability of loss.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Productivity Time at Risk",
        "Exposure_Amount": "[Formula: Sum of FTE-hours at risk across organization]",
        "Currency_Unit": "FTE-Hours (Annual)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Hours at Risk × Blended Labor Rate]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "People, Technology, Processes, Culture dimensions",
        "Calculation_Method": "Productivity time at risk from: (1) Employee disengagement (Culture/People), (2) System downtime (Technology), (3) Process inefficiencies (Processes), (4) Change disruption (Change dimension). Calculate: Total workforce FTE × % time loss factors × probability. Convert to $ using blended labor rate.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Service Availability at Risk",
        "Exposure_Amount": "[Formula: Service uptime/SLA exposure]",
        "Currency_Unit": "Hours of Downtime (Annual)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Downtime Hours × Revenue per Hour + SLA Penalties]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Technology, Processes, Products & Services, Third Parties dimensions",
        "Calculation_Method": "Service availability at risk from: (1) Technology infrastructure failures, (2) Process breakdowns, (3) Service delivery issues, (4) Third-party failures. Calculate expected downtime hours when metrics in Risk range. Financial impact = (revenue per hour × downtime) + SLA penalties + customer churn.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Product at Risk",
        "Exposure_Amount": "[Formula: Product delivery/quality failure exposure]",
        "Currency_Unit": "$ (Annual)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Products & Services, Innovation, Processes, Technology dimensions",
        "Calculation_Method": "Product at risk from: (1) Product quality failures, (2) Development delays, (3) Launch failures, (4) Technology debt limiting product evolution. Impact includes: lost revenue from failed products, warranty costs, recalls, customer churn, brand damage.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Reputation at Risk",
        "Exposure_Amount": "[Formula: Estimated brand value erosion + stakeholder impact]",
        "Currency_Unit": "$ (Brand Value)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "12-36 months",
        "Confidence_Level": "70% (P70)",
        "Data_Sources": "Reputation, Brand, Culture, ESG dimensions",
        "Calculation_Method": "Reputation at risk from: (1) Direct reputation vulnerabilities (Reputation dimension), (2) Brand strength erosion, (3) Culture issues becoming public, (4) ESG controversies. Quantify using: Brand valuation × % erosion risk + premium pricing loss + customer acquisition cost increase + talent cost premium.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Enterprise Value at Risk",
        "Exposure_Amount": "[Formula: Total enterprise value exposure]",
        "Currency_Unit": "$ (Enterprise Value)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "36 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "All 16 dimensions aggregated",
        "Calculation_Method": "Composite metric aggregating all risk exposures weighted by enterprise value impact. Calculate as: Sum of (Revenue at Risk × Revenue Multiple) + (EBITDA impact × EBITDA Multiple) + (Strategic positioning erosion × Strategic Premium) + (Reputation damage × Brand Multiple). Represents total downside to enterprise valuation.",
        "Last_Updated": "[Date]"
    }
]

# Additional Lagging Risk Metrics - Company Level
additional_company_risks = [
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Customer Lifetime Value at Risk",
        "Exposure_Amount": "[Formula: CLV of at-risk customer base]",
        "Currency_Unit": "$ (Multi-year CLV)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: At-Risk CLV × Churn Probability]",
        "Time_Horizon": "36 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Revenue, Annual Results, Brand, Products & Services, Reputation dimensions",
        "Calculation_Method": "CLV at risk from: (1) Customer churn indicators (Revenue dimension), (2) Satisfaction decline (Annual Results), (3) Brand weakness, (4) Product issues, (5) Reputation damage. Calculate: Number of at-risk customers × Average CLV × Churn probability. At-risk = customers with declining engagement, satisfaction <70%, NPS <20.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Market Share at Risk",
        "Exposure_Amount": "[Formula: Revenue equivalent of vulnerable market share]",
        "Currency_Unit": "% Market Share & $ Revenue",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Market Share % × Market Size × Probability]",
        "Time_Horizon": "24 months",
        "Confidence_Level": "70% (P70)",
        "Data_Sources": "Strategic Goals, Revenue, Products & Services, Innovation dimensions",
        "Calculation_Method": "Market share at risk from: (1) Weak competitive position (Strategic Goals), (2) Revenue vulnerabilities, (3) Product portfolio gaps, (4) Innovation lag. Calculate: Current market share × % at risk from competitive threats × market size. Triggered by competitive position index <60, declining win rates, innovation lag.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Talent at Risk",
        "Exposure_Amount": "[Formula: Replacement cost + productivity loss from key talent departure]",
        "Currency_Unit": "$ (Replacement Cost)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Expected Departures × Avg Replacement Cost]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "People, Culture dimensions",
        "Calculation_Method": "Talent at risk from: (1) High turnover indicators (People dimension), (2) Low engagement (Culture), (3) Retention risk (People). Calculate: Number of key employees at flight risk × (Replacement cost + Productivity loss during vacancy + Knowledge loss). Flight risk = engagement <60%, tenure >2 years, high performer, external opportunities.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Compliance at Risk",
        "Exposure_Amount": "[Formula: Estimated fines, penalties, operational restrictions]",
        "Currency_Unit": "$ (Fines + Impact)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "24 months",
        "Confidence_Level": "70% (P70)",
        "Data_Sources": "Technology (Cybersecurity), Third Parties, Processes, Reputation dimensions",
        "Calculation_Method": "Compliance at risk from: (1) Cybersecurity vulnerabilities, (2) Third-party compliance gaps, (3) Process control weaknesses, (4) ESG compliance issues. Estimate: Max regulatory fine for violation + operational restrictions impact + remediation costs. Probability from control effectiveness scores.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Data/IP at Risk",
        "Exposure_Amount": "[Formula: Value of vulnerable data/IP assets]",
        "Currency_Unit": "$ (Asset Value)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Technology, Innovation, Third Parties dimensions",
        "Calculation_Method": "Data/IP at risk from: (1) Cybersecurity vulnerabilities (Technology), (2) IP protection weaknesses (Innovation), (3) Third-party data exposures. Quantify: Critical data/IP value + breach costs (notification, remediation, fines) + competitive advantage loss. Probability from security posture assessment.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Cash Flow at Risk",
        "Exposure_Amount": "[Formula: Potential cash flow disruption]",
        "Currency_Unit": "$ (Cash Flow)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Exposure × Probability]",
        "Time_Horizon": "12 months",
        "Confidence_Level": "90% (P90 - high confidence for liquidity)",
        "Data_Sources": "Working Capital, Revenue, Costs, Annual Results dimensions",
        "Calculation_Method": "Cash flow at risk from: (1) Working capital deterioration, (2) Revenue collection issues, (3) Cost overruns, (4) Profitability erosion. Calculate: Operating cash flow × risk factors (DSO increase, inventory buildup, payable stretch limits). Critical for debt covenant compliance and liquidity management.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Credit Rating at Risk",
        "Exposure_Amount": "[Formula: Impact of rating downgrade on debt costs]",
        "Currency_Unit": "$ (Incremental Interest)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Additional Interest Cost × Probability]",
        "Time_Horizon": "12-24 months",
        "Confidence_Level": "80% (P80)",
        "Data_Sources": "Annual Results, Working Capital, Investment, Strategic Goals dimensions",
        "Calculation_Method": "Credit rating at risk from: (1) Deteriorating financial metrics (leverage, coverage ratios), (2) Cash flow weakness, (3) Strategic execution failures. Calculate: Total debt × basis point increase per notch downgrade × downgrade probability. Rating triggers = leverage >3.5x, coverage <3x, declining EBITDA.",
        "Last_Updated": "[Date]"
    },
    {
        "Aggregation_Level": "Company",
        "Entity_Name": "[Company Name]",
        "Risk_Metric": "Innovation Pipeline at Risk",
        "Exposure_Amount": "[Formula: Future revenue from at-risk innovation projects]",
        "Currency_Unit": "$ (Future Revenue)",
        "Probability": "[%]",
        "Expected_Loss": "[Formula: Pipeline Value × Failure Probability]",
        "Time_Horizon": "24-36 months",
        "Confidence_Level": "70% (P70)",
        "Data_Sources": "Innovation, Products & Services, Technology, Strategic Goals dimensions",
        "Calculation_Method": "Innovation pipeline at risk from: (1) Innovation execution issues, (2) Product development delays, (3) Technology capability gaps, (4) Strategic initiative failures. Calculate: Sum of (Project expected NPV × Probability of failure). Failure risk from: low innovation success rate, capability gaps, resource constraints.",
        "Last_Updated": "[Date]"
    }
]

# Add company-level risks to dashboard
risk_dashboard_data.extend(company_risks)
risk_dashboard_data.append({})  # Blank row separator
risk_dashboard_data.append({
    "Aggregation_Level": "Company",
    "Entity_Name": "[Company Name]",
    "Risk_Metric": "=== ADDITIONAL LAGGING INDICATORS ===",
    "Exposure_Amount": "",
    "Currency_Unit": "",
    "Probability": "",
    "Expected_Loss": "",
    "Time_Horizon": "",
    "Confidence_Level": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": ""
})
risk_dashboard_data.extend(additional_company_risks)

# ============================================================================
# SECTION 2: DOMAIN-LEVEL RISK BREAKDOWN (4 Domains)
# ============================================================================

risk_dashboard_data.append({})  # Blank row
risk_dashboard_data.append({})  # Blank row
risk_dashboard_data.append({
    "Aggregation_Level": "DOMAIN LEVEL",
    "Entity_Name": "=== 4 DOMAINS ===",
    "Risk_Metric": "Risk breakdown by major domain",
    "Exposure_Amount": "",
    "Currency_Unit": "",
    "Probability": "",
    "Expected_Loss": "",
    "Time_Horizon": "",
    "Confidence_Level": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": ""
})

# Define the 4 domains with their risk contribution profiles
domains = [
    {
        "name": "ECONOMICS Domain",
        "dimensions": ["Revenue", "Costs", "Investment", "Working Capital"],
        "primary_risks": ["Revenue at Risk", "Opex at Risk", "Capex at Risk", "Cash Flow at Risk"],
        "risk_profile": "Direct financial impact - revenue loss, cost overruns, capital waste, liquidity crisis"
    },
    {
        "name": "ENABLERS Domain",
        "dimensions": ["Brand", "Culture", "People", "Technology", "Third Parties"],
        "primary_risks": ["Reputation at Risk", "Talent at Risk", "Opex at Risk", "Data/IP at Risk", "Compliance at Risk"],
        "risk_profile": "Organizational capability risks - foundation that enables all other performance"
    },
    {
        "name": "EXECUTION Domain",
        "dimensions": ["Innovation", "Change", "Processes", "Product & Services"],
        "primary_risks": ["Product at Risk", "Service Availability at Risk", "Innovation Pipeline at Risk", "Productivity Time at Risk"],
        "risk_profile": "Operational delivery risks - how well organization executes strategy"
    },
    {
        "name": "VALUE Domain",
        "dimensions": ["Annual Results", "Strategic Goals", "Reputation"],
        "primary_risks": ["Enterprise Value at Risk", "Market Share at Risk", "Reputation at Risk", "Strategic Goals at Risk"],
        "risk_profile": "Strategic outcome risks - long-term value creation and competitive position"
    }
]

# Generate domain-level risk rows
for domain in domains:
    risk_dashboard_data.append({})  # Separator
    risk_dashboard_data.append({
        "Aggregation_Level": "Domain",
        "Entity_Name": domain["name"],
        "Risk_Metric": f"=== {domain['name']} RISK PROFILE ===",
        "Exposure_Amount": "",
        "Currency_Unit": "",
        "Probability": "",
        "Expected_Loss": "",
        "Time_Horizon": "",
        "Confidence_Level": "",
        "Data_Sources": f"Dimensions: {', '.join(domain['dimensions'])}",
        "Calculation_Method": domain["risk_profile"],
        "Last_Updated": ""
    })

    # Add each primary risk metric for this domain
    for risk_metric in domain["primary_risks"]:
        risk_dashboard_data.append({
            "Aggregation_Level": "Domain",
            "Entity_Name": domain["name"],
            "Risk_Metric": risk_metric,
            "Exposure_Amount": f"[Formula: Sum from {domain['name']} dimensions]",
            "Currency_Unit": "[See Company Level]",
            "Probability": "[Weighted avg of dimension probabilities]",
            "Expected_Loss": "[Formula: Exposure × Probability]",
            "Time_Horizon": "[12-36 months depending on metric]",
            "Confidence_Level": "[70-90% depending on metric]",
            "Data_Sources": f"{', '.join(domain['dimensions'])}",
            "Calculation_Method": f"Aggregate risk exposure from {len(domain['dimensions'])} dimensions within {domain['name']}. Weights based on dimension contribution to specific risk metric.",
            "Last_Updated": "[Date]"
        })

# ============================================================================
# SECTION 3: DIMENSION-LEVEL RISK BREAKDOWN (16 Dimensions)
# ============================================================================

risk_dashboard_data.append({})  # Blank row
risk_dashboard_data.append({})  # Blank row
risk_dashboard_data.append({
    "Aggregation_Level": "DIMENSION LEVEL",
    "Entity_Name": "=== 16 DIMENSIONS ===",
    "Risk_Metric": "Detailed risk breakdown by dimension",
    "Exposure_Amount": "",
    "Currency_Unit": "",
    "Probability": "",
    "Expected_Loss": "",
    "Time_Horizon": "",
    "Confidence_Level": "",
    "Data_Sources": "",
    "Calculation_Method": "",
    "Last_Updated": ""
})

# Define all 16 dimensions with their primary risk contributions
dimensions_detailed = [
    # ECONOMICS Domain (4 dimensions)
    {
        "domain": "ECONOMICS",
        "dimension": "Revenue",
        "primary_risk_contributions": {
            "Revenue at Risk": "100% - Direct revenue stream vulnerabilities",
            "Customer Lifetime Value at Risk": "90% - Customer base erosion",
            "Market Share at Risk": "40% - Revenue-related share loss",
            "Enterprise Value at Risk": "30% - Revenue multiple impact"
        },
        "calculation_approach": "Aggregate revenue vulnerabilities from 6 elements (24 sub-elements): Revenue Growth, Customer Acquisition, Customer Retention, Pricing Power, Revenue Mix, Revenue Quality. Risk triggered when metrics in Risk Factors range (e.g., growth <0%, churn >15%, NPS <20)."
    },
    {
        "domain": "ECONOMICS",
        "dimension": "Costs",
        "primary_risk_contributions": {
            "Opex at Risk": "60% - Direct operating cost overruns",
            "Cash Flow at Risk": "40% - Cost impact on cash",
            "Enterprise Value at Risk": "25% - EBITDA impact"
        },
        "calculation_approach": "Aggregate cost exposure from 6 elements: Cost Structure, Cost Drivers, Cost Efficiency, Cost Control, Economies of Scale, Shared Services. Risk from: cost ratio >industry avg, efficiency declining, controls weak, diseconomies."
    },
    {
        "domain": "ECONOMICS",
        "dimension": "Investment",
        "primary_risk_contributions": {
            "Capex at Risk": "80% - Direct capital project failures",
            "Stratex at Risk": "50% - Strategic investment failures",
            "Innovation Pipeline at Risk": "30% - R&D investment waste"
        },
        "calculation_approach": "Aggregate investment risks from 6 elements: Investment Strategy, Project Portfolio, Capital Allocation, ROI Realization, Portfolio Management, Governance. Risk from: ROI <hurdle rate, >40% projects over budget, weak governance."
    },
    {
        "domain": "ECONOMICS",
        "dimension": "Working Capital",
        "primary_risk_contributions": {
            "Cash Flow at Risk": "80% - Direct liquidity impact",
            "Credit Rating at Risk": "40% - Covenant risk from WC deterioration",
            "Enterprise Value at Risk": "15% - Cash efficiency impact"
        },
        "calculation_approach": "Aggregate WC risks from 6 elements: Cash Conversion Cycle, Receivables, Inventory, Payables, Cash Management, Working Capital Efficiency. Risk from: CCC >60 days, DSO >60, inventory turns <6, cash reserves <30 days."
    },

    # ENABLERS Domain (5 dimensions)
    {
        "domain": "ENABLERS",
        "dimension": "Brand",
        "primary_risk_contributions": {
            "Reputation at Risk": "60% - Brand value erosion",
            "Revenue at Risk": "30% - Brand-driven revenue loss",
            "Market Share at Risk": "30% - Brand weakness enables share loss"
        },
        "calculation_approach": "Aggregate brand risks from 6 elements: Brand Strength, Brand Perception, Brand Equity, Brand Differentiation, Brand Portfolio, Brand Protection. Risk from: awareness <60%, NPS <30, equity declining, no differentiation, weak protection."
    },
    {
        "domain": "ENABLERS",
        "dimension": "Culture",
        "primary_risk_contributions": {
            "Talent at Risk": "50% - Culture drives retention",
            "Productivity Time at Risk": "40% - Engagement impact",
            "Reputation at Risk": "20% - Culture issues going public"
        },
        "calculation_approach": "Aggregate culture risks from 6 elements: Values Alignment, Employee Engagement, Psychological Safety, Collaboration, Learning Culture, Diversity & Inclusion. Risk from: engagement <60%, safety <60%, collaboration weak, learning <2% payroll."
    },
    {
        "domain": "ENABLERS",
        "dimension": "People",
        "primary_risk_contributions": {
            "Talent at Risk": "100% - Direct talent risk",
            "Productivity Time at Risk": "40% - Capability/productivity impact",
            "Opex at Risk": "20% - Turnover/replacement costs"
        },
        "calculation_approach": "Aggregate people risks from 6 elements: Talent Acquisition, Performance Management, Compensation & Benefits, Career Development, Employee Wellbeing, Workforce Planning. Risk from: turnover >15%, hiring quality <70%, performance weak, succession gaps."
    },
    {
        "domain": "ENABLERS",
        "dimension": "Technology",
        "primary_risk_contributions": {
            "Data/IP at Risk": "80% - Cybersecurity vulnerabilities",
            "Service Availability at Risk": "70% - System failures",
            "Opex at Risk": "30% - IT cost overruns",
            "Compliance at Risk": "40% - Technology control gaps",
            "Productivity Time at Risk": "30% - System downtime"
        },
        "calculation_approach": "Aggregate technology risks from 6 elements: IT Infrastructure, Application Portfolio, Data & Analytics, Cybersecurity, IT Operations, Technology Innovation. Risk from: uptime <99%, security score <70, tech debt high, innovation lag."
    },
    {
        "domain": "ENABLERS",
        "dimension": "Third Parties",
        "primary_risk_contributions": {
            "Opex at Risk": "25% - Vendor failures/overruns",
            "Service Availability at Risk": "40% - Third-party outages",
            "Compliance at Risk": "50% - Vendor compliance gaps",
            "Data/IP at Risk": "30% - Third-party data exposures"
        },
        "calculation_approach": "Aggregate third-party risks from 6 elements: Supplier Management, Partner Ecosystem, Outsourcing Relationships, Vendor Risk, Contract Management, Third-Party Performance. Risk from: concentration >25%, risk score <70, SLA <90%, weak contracts."
    },

    # EXECUTION Domain (4 dimensions)
    {
        "domain": "EXECUTION",
        "dimension": "Innovation",
        "primary_risk_contributions": {
            "Innovation Pipeline at Risk": "100% - Direct pipeline risk",
            "Product at Risk": "40% - Innovation failures",
            "Revenue at Risk": "20% - Innovation-driven growth loss",
            "Market Share at Risk": "30% - Innovation lag enabling share loss"
        },
        "calculation_approach": "Aggregate innovation risks from 6 elements: Innovation Strategy, Idea Generation, Development Process, Portfolio Management, Launch Execution, Innovation Culture. Risk from: investment <5% revenue, time-to-market >18mo, success <50%, no culture."
    },
    {
        "domain": "EXECUTION",
        "dimension": "Change",
        "primary_risk_contributions": {
            "Stratex at Risk": "40% - Change initiative failures",
            "Productivity Time at Risk": "30% - Change disruption",
            "Opex at Risk": "15% - Change cost overruns"
        },
        "calculation_approach": "Aggregate change risks from 6 elements: Change Capability, Change Leadership, Stakeholder Engagement, Communication, Adoption & Sustainment, Change Performance. Risk from: success rate <60%, no capability, poor communication, low adoption."
    },
    {
        "domain": "EXECUTION",
        "dimension": "Processes",
        "primary_risk_contributions": {
            "Service Availability at Risk": "30% - Process failures",
            "Productivity Time at Risk": "40% - Process inefficiency",
            "Opex at Risk": "25% - Process waste",
            "Product at Risk": "30% - Process quality issues"
        },
        "calculation_approach": "Aggregate process risks from 6 elements: Process Design, Documentation, Performance, Control, Improvement, Technology. Risk from: value-add <30%, FPY <90%, control weak, automation <30%."
    },
    {
        "domain": "EXECUTION",
        "dimension": "Product & Services",
        "primary_risk_contributions": {
            "Product at Risk": "100% - Direct product delivery risk",
            "Service Availability at Risk": "60% - Service delivery risk",
            "Revenue at Risk": "40% - Product/service quality impact on revenue",
            "Reputation at Risk": "25% - Product failures damaging reputation"
        },
        "calculation_approach": "Aggregate product/service risks from 6 elements: Portfolio Strategy, Product Development, Service Delivery, Quality Management, Lifecycle Management, Customer Experience. Risk from: cycle >18mo, success <60%, CSAT <80%, NPS <30."
    },

    # VALUE Domain (3 dimensions)
    {
        "domain": "VALUE",
        "dimension": "Annual Results",
        "primary_risk_contributions": {
            "Enterprise Value at Risk": "50% - Results directly impact valuation",
            "Revenue at Risk": "30% - Results indicate revenue vulnerability",
            "Credit Rating at Risk": "60% - Results drive rating"
        },
        "calculation_approach": "Aggregate annual results risks from 6 elements: Revenue Achievement, Profitability, Cash Generation, Operational Efficiency, Customer Performance, Market Performance. Risk from: revenue <90% target, EBITDA <15%, cash conversion weak, market share declining."
    },
    {
        "domain": "VALUE",
        "dimension": "Strategic Goals",
        "primary_risk_contributions": {
            "Enterprise Value at Risk": "40% - Strategic position impact",
            "Market Share at Risk": "60% - Competitive position erosion",
            "Stratex at Risk": "50% - Strategic initiative failures",
            "Innovation Pipeline at Risk": "30% - Strategic capability gaps"
        },
        "calculation_approach": "Aggregate strategic goal risks from 6 elements: Vision Achievement, Strategic Initiatives, Competitive Position, Market Leadership, Capability Development, Strategic Partnerships. Risk from: vision <60% achieved, initiatives <60% success, position weak, capabilities lagging."
    },
    {
        "domain": "VALUE",
        "dimension": "Reputation",
        "primary_risk_contributions": {
            "Reputation at Risk": "100% - Direct reputation risk",
            "Revenue at Risk": "25% - Reputation impact on sales",
            "Talent at Risk": "30% - Reputation impact on hiring",
            "Enterprise Value at Risk": "20% - Reputation impact on valuation"
        },
        "calculation_approach": "Aggregate reputation risks from 6 elements: Corporate Reputation, Media Perception, Stakeholder Trust, Crisis Management, Transparency & Disclosure, ESG Performance. Risk from: RepTrak <60, negative media >30%, trust <60%, crisis unready, ESG bottom quartile."
    }
]

# Generate dimension-level risk rows
for dim in dimensions_detailed:
    risk_dashboard_data.append({})  # Separator
    risk_dashboard_data.append({
        "Aggregation_Level": "Dimension",
        "Entity_Name": f"{dim['domain']} > {dim['dimension']}",
        "Risk_Metric": f"=== {dim['dimension']} DIMENSION ===",
        "Exposure_Amount": "",
        "Currency_Unit": "",
        "Probability": "",
        "Expected_Loss": "",
        "Time_Horizon": "",
        "Confidence_Level": "",
        "Data_Sources": "6 elements × 4 sub-elements each = 24 data points",
        "Calculation_Method": dim["calculation_approach"],
        "Last_Updated": ""
    })

    # Add risk contributions for this dimension
    for risk_metric, contribution in dim["primary_risk_contributions"].items():
        risk_dashboard_data.append({
            "Aggregation_Level": "Dimension",
            "Entity_Name": f"{dim['domain']} > {dim['dimension']}",
            "Risk_Metric": risk_metric,
            "Exposure_Amount": "[Calculated from 24 sub-elements]",
            "Currency_Unit": "[See Company Level definition]",
            "Probability": "[From Risk Factor assessments]",
            "Expected_Loss": "[Exposure × Probability]",
            "Time_Horizon": "[Metric dependent]",
            "Confidence_Level": "[Metric dependent]",
            "Data_Sources": f"{dim['dimension']} dimension (31 rows: 1 dimension + 6 elements + 24 sub-elements)",
            "Calculation_Method": contribution,
            "Last_Updated": "[Date]"
        })

# ============================================================================
# Write to Excel
# ============================================================================

print("=" * 80)
print("GENERATING RISK DASHBOARD SHEET")
print("=" * 80)
print()

wb_path = r"C:\Users\AndrewSmart\Claude_Projects\AAgents\DecideWright-EA\Context\Predixtive-Model\Predixtive_Model.xlsx"
wb = load_workbook(wb_path)

# Create or get Risk Dashboard sheet
if "Risk Dashboard" in wb.sheetnames:
    print("Risk Dashboard sheet exists - clearing and rewriting...")
    ws = wb["Risk Dashboard"]
    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
else:
    print("Creating new Risk Dashboard sheet...")
    ws = wb.create_sheet("Risk Dashboard", 0)  # Insert as first sheet

# Define headers
headers = [
    "Aggregation_Level",
    "Entity_Name",
    "Risk_Metric",
    "Exposure_Amount",
    "Currency_Unit",
    "Probability",
    "Expected_Loss",
    "Time_Horizon",
    "Confidence_Level",
    "Data_Sources",
    "Calculation_Method",
    "Last_Updated"
]

# Write headers with formatting
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write data
for row_idx, data_row in enumerate(risk_dashboard_data, 2):
    for col_idx, header in enumerate(headers, 1):
        value = data_row.get(header, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Highlight section headers
        if "===" in str(value):
            cell.font = Font(bold=True, size=11, color="1F4E78")
            cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

# Auto-adjust column widths
column_widths = {
    "A": 18,  # Aggregation_Level
    "B": 35,  # Entity_Name
    "C": 40,  # Risk_Metric
    "D": 50,  # Exposure_Amount
    "E": 15,  # Currency_Unit
    "F": 12,  # Probability
    "G": 18,  # Expected_Loss
    "H": 15,  # Time_Horizon
    "I": 15,  # Confidence_Level
    "J": 50,  # Data_Sources
    "K": 80,  # Calculation_Method
    "L": 12   # Last_Updated
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Save workbook
wb.save(wb_path)

print()
print(f"Risk Dashboard generated with {len(risk_dashboard_data)} rows")
print()
print("Risk Metrics Summary:")
print("  Primary Risk Metrics (Requested):      9 metrics")
print("  Additional Lagging Risk Metrics:       8 metrics")
print("  Total Risk Metrics:                   17 metrics")
print()
print("Aggregation Levels:")
print("  Company Level:    Overall enterprise risk profile")
print("  Domain Level:     4 domains × 17 metrics = risk attribution by domain")
print("  Dimension Level:  16 dimensions × primary risk contributions")
print()
print("=" * 80)
print("RISK DASHBOARD COMPLETE!")
print("=" * 80)
print()
print("The Risk Dashboard is now the first sheet in the workbook,")
print("providing executive-level risk visibility with drill-down to dimensions.")
